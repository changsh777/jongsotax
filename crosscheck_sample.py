"""
경우의 수별 무작위 샘플링 교차검증
- 파싱결과.xlsx 값 vs PDF 재파싱 직접 비교
경우의 수:
  1) 간편 + 기준경비율 + 타소득X
  2) 간편 + 기준경비율 + 타소득O
  3) 간편 + 단순경비율 + 타소득X
  4) 간편 + 단순경비율 + 타소득O
  5) 복식 + 타소득X
  6) 복식 + 타소득O
"""
import sys, io, os, random
os.environ.setdefault("SEOTAX_ENV", "nas")
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, r"F:\종소세2026")

import openpyxl
from config import PARSE_RESULT_XLSX
from parse_to_xlsx import parse_anneam, collect_pdfs
from fee_calculator import count_other_income

random.seed(42)

OTHER_COLS = ["이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타"]

def main():
    # 1) 파싱결과.xlsx 로드
    wb = openpyxl.load_workbook(PARSE_RESULT_XLSX, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    saved = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not row[idx["PDF경로"]]: continue
        name = str(row[0]).strip()
        saved[name] = {h: row[i] for i, h in enumerate(headers) if h}

    # 2) 경우의 수 분류
    cases = {
        "간편+기준경비율+타소득X": [],
        "간편+기준경비율+타소득O": [],
        "간편+단순경비율+타소득X": [],
        "간편+단순경비율+타소득O": [],
        "복식+타소득X":            [],
        "복식+타소득O":            [],
    }
    for name, d in saved.items():
        ledger = str(d.get("기장의무") or "")
        rate   = str(d.get("추계시적용경비율") or "")
        has_other = any(str(d.get(c) or "").strip() == "O" for c in OTHER_COLS)
        if "복식" in ledger:
            key = "복식+타소득O" if has_other else "복식+타소득X"
        elif "간편" in ledger:
            if "단순" in rate:
                key = "간편+단순경비율+타소득O" if has_other else "간편+단순경비율+타소득X"
            else:
                key = "간편+기준경비율+타소득O" if has_other else "간편+기준경비율+타소득X"
        else:
            continue
        cases[key].append(name)

    print("=== 경우의 수별 모집단 ===")
    for k, v in cases.items():
        print(f"  {k}: {len(v)}명")

    # 3) 각 케이스에서 최대 2명씩 무작위 샘플
    samples = []
    for key, names in cases.items():
        n = min(2, len(names))
        pick = random.sample(names, n) if n > 0 else []
        samples.extend([(key, name) for name in pick])
    print(f"\n=== 샘플 {len(samples)}명 교차검증 ===\n")

    # PDF 경로 인덱스
    pdf_by_name = {name: str(d["PDF경로"]) for name, d in saved.items()}

    from pathlib import Path
    ok = err = 0
    for case, name in samples:
        pdf_path = Path(pdf_by_name.get(name, ""))
        if not pdf_path.exists():
            print(f"[{case}] {name}: PDF 없음 - 건너뜀")
            continue

        # 재파싱
        try:
            fresh = parse_anneam(pdf_path)
        except Exception as e:
            print(f"[{case}] {name}: 재파싱 실패 - {e}")
            continue

        saved_d = saved[name]
        checks = {
            "수입금액총계":    (saved_d.get("수입금액총계"), fresh.get("수입금액총계")),
            "기장의무":        (saved_d.get("기장의무"),     fresh.get("기장의무")),
            "추계시적용경비율":(saved_d.get("추계시적용경비율"), fresh.get("추계시적용경비율")),
            "이자":            (saved_d.get("이자"),          fresh.get("이자")),
            "배당":            (saved_d.get("배당"),          fresh.get("배당")),
            "근로(단일)":      (saved_d.get("근로(단일)"),    fresh.get("근로(단일)")),
            "연금":            (saved_d.get("연금"),          fresh.get("연금")),
            "기타":            (saved_d.get("기타"),          fresh.get("기타")),
        }
        mismatches = [(k, sv, fv) for k, (sv, fv) in checks.items() if str(sv) != str(fv)]

        if mismatches:
            err += 1
            print(f"[{case}] {name} ❌ 불일치:")
            for k, sv, fv in mismatches:
                print(f"    {k}: 저장={sv!r} / 재파싱={fv!r}")
        else:
            ok += 1
            income = saved_d.get('수입금액총계')
            income_str = f"{income:,}" if isinstance(income, int) else str(income)
            print(f"[{case}] {name} ✓  수입={income_str} / {saved_d.get('기장의무')} / {saved_d.get('추계시적용경비율')}")

    print(f"\n결과: {ok}명 일치 / {err}명 불일치 (총 {ok+err}명)")

if __name__ == "__main__":
    main()
