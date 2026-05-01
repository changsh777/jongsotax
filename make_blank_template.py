"""
박찬영(간편장부) / 오지혜(복식부기) 양식에서 노란색 셀만 비워서
빈 작업판 템플릿 2개 생성.
- 노란셀 = 입력셀 = 데이터 들어가는 곳
- 검정셀 = 수식 (그대로 둠)
- 회색/없음 = 라벨·고정 (그대로 둠)
"""
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
import warnings
warnings.filterwarnings("ignore")

SOURCES = {
    "간편장부": r"C:\Users\pc\OneDrive\문서\카카오톡 받은 파일\종소세자동화-박찬영 (2025신고대리간편장부).xlsx",
    "복식부기": r"C:\Users\pc\OneDrive\문서\카카오톡 받은 파일\종소세자동화-오지혜(2025신고대리복식).xlsx",
}
OUT_DIR = r"F:\종소세2026\templates"

import os
os.makedirs(OUT_DIR, exist_ok=True)


def is_yellow(cell):
    """셀이 노란색 배경인지 확인 (FFFF00 또는 비슷)"""
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.rgb is None:
        return False
    rgb = str(fg.rgb).upper()
    if not rgb or len(rgb) < 6:
        return False
    # 노랑 계열: R~FF, G~FF, B<80
    rgb6 = rgb[-6:]  # 알파 제외
    try:
        r = int(rgb6[0:2], 16)
        g = int(rgb6[2:4], 16)
        b = int(rgb6[4:6], 16)
    except Exception:
        return False
    return r >= 200 and g >= 200 and b < 120


GUIDE_KEYWORDS = ["입력", "셀", "금지", "수식", "노랑", "노란", "검정", "검은"]


def is_guide_text(value):
    """가이드/라벨 텍스트인지 (지우면 안 되는)"""
    if not isinstance(value, str):
        return False
    return any(k in value for k in GUIDE_KEYWORDS)


def make_blank(src_path, out_path, label):
    wb = openpyxl.load_workbook(src_path)
    yellow_cells_by_sheet = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cleared = []
        kept = []
        for row in ws.iter_rows():
            for cell in row:
                if is_yellow(cell) and cell.value is not None:
                    if is_guide_text(cell.value):
                        kept.append(f"{cell.coordinate}={cell.value!r}")
                        continue
                    cleared.append(f"{cell.coordinate}={cell.value!r}")
                    cell.value = None
        yellow_cells_by_sheet[sheet_name] = (cleared, kept)

    wb.save(out_path)

    print(f"\n=== {label} → {out_path} ===")
    for sheet_name, (cleared, kept) in yellow_cells_by_sheet.items():
        print(f"  [{sheet_name}] 비운 노란셀 {len(cleared)}개, 가이드 유지 {len(kept)}개")
        for c in cleared[:5]:
            print(f"    [지움] {c}")
        if len(cleared) > 5:
            print(f"    ... +{len(cleared)-5}개")


for label, src in SOURCES.items():
    out_path = os.path.join(OUT_DIR, f"빈양식_{label}.xlsx")
    make_blank(src, out_path, label)

print("\n[완료] 빈 양식 생성됨")
