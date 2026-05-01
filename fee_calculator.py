"""
수수료 자동 계산 + 견적 메시지 생성

룰 (장성환 정의):
1. 사업소득 + 장부유형 → 구간 정가 (간편/복식)
2. 타소득 1개당 + 100,000원
3. 합산 → 사전접수면 × 0.8 (20% 할인)

안전:
- 구간 경계 명확히 (X 이하인 가장 큰 row 적용)
- sanity check
- 견적 발송 자동화는 별도 (이 모듈은 계산 + 메시지 생성까지만)
"""
import sys
sys.path.insert(0, r"F:\종소세2026")
from gsheet_writer import get_credentials
import gspread


# 요율표 (구글시트 '수수료' 시트에서 한 번 읽어서 hardcode 가능)
# 사업소득 ≤ 이 row 금액 → 이 row 가격 적용
# (단, 다음 row 금액 미만일 때) - 즉 row 금액 ≤ 사업소득 < next row 금액
FEE_SCHEDULE = [
    # (사업소득_하한, 간편_정가, 복식_정가)
    (30_000_000, 100_000, 300_000),
    (40_000_000, 150_000, 300_000),
    (50_000_000, 200_000, 400_000),
    (75_000_000, 250_000, 500_000),
    (100_000_000, 300_000, 600_000),
    (150_000_000, 350_000, 700_000),
    (200_000_000, 400_000, 800_000),
    (250_000_000, 450_000, 900_000),
    (300_000_000, 500_000, 1_000_000),
    (350_000_000, 550_000, 1_100_000),
    (400_000_000, 600_000, 1_200_000),
    (450_000_000, 650_000, 1_300_000),
    (500_000_000, 700_000, 1_400_000),
    (550_000_000, 750_000, 1_500_000),
    (600_000_000, 800_000, 1_600_000),
    (650_000_000, 850_000, 1_700_000),
    (700_000_000, 900_000, 1_800_000),
    (750_000_000, 950_000, 1_900_000),
    (800_000_000, 1_000_000, 2_000_000),
    (850_000_000, 1_050_000, 2_100_000),
    (900_000_000, 1_100_000, 2_200_000),
    (950_000_000, 1_150_000, 2_300_000),
    (1_000_000_000, 1_200_000, 2_400_000),
]

OTHER_INCOME_PER_ITEM = 100_000  # 타소득 개당
ADVANCE_DISCOUNT = 0.8  # 사전접수 20% 할인 (정가의 80%)

# Sanity check
MIN_FEE = 50_000   # 5만원 미만 → 의심
MAX_FEE = 5_000_000  # 500만원 초과 → 의심


def get_base_price(income, ledger_type):
    """사업소득 구간별 정가 - "이하" 룰 (장성환 최종 확정)
    - income ≤ row 금액 인 row 중 가장 작은 row 적용
    - 정확히 row 금액과 같으면 그 row 적용 (예: 3000만원 = 30M row)
    - income > 1B인 경우 최대 row(1B) 적용
    예시:
      - 박수경 4,500,000 ≤ 30M → 30M row → 100,000원
      - 3000만원 정확 = 30M → 30M row → 100,000원
      - 4000만원 정확 = 40M → 40M row → 150,000원
      - 서인미 100,517,792 ≤ 150M → 150M row → 700,000원 (복식)
    """
    if income is None or income < 0:
        raise ValueError(f"사업소득 이상값: {income}")

    eligible = [r for r in FEE_SCHEDULE if income <= r[0]]
    if not eligible:
        # income > 1B → 최대 row(1B) 적용
        chosen = FEE_SCHEDULE[-1]
    else:
        chosen = min(eligible, key=lambda r: r[0])

    if ledger_type == "간편장부" or "간편" in (ledger_type or ""):
        return chosen[1]
    elif ledger_type == "복식부기" or "복식" in (ledger_type or ""):
        return chosen[2]
    else:
        raise ValueError(f"장부유형 알 수 없음: {ledger_type}")


def count_other_income(customer_data):
    """타소득 개수 (카테고리 단위):
    - 금융소득 (이자+배당 합쳐서 1개)
    - 근로소득 (근로 단일/복수 합쳐서 1개)
    - 연금소득
    - 기타소득
    최대 4개
    """
    categories = {
        "금융": ["이자", "배당"],
        "근로": ["근로(단일)", "근로(복수)"],
        "연금": ["연금"],
        "기타": ["기타"],
    }
    count = 0
    for cat_name, fields in categories.items():
        if any(customer_data.get(f) == "O" for f in fields):
            count += 1
    return count


def get_other_income_categories(customer_data):
    """O 표시된 카테고리 이름 리스트 (메시지용)"""
    categories = {
        "금융(이자/배당)": ["이자", "배당"],
        "근로": ["근로(단일)", "근로(복수)"],
        "연금": ["연금"],
        "기타": ["기타"],
    }
    out = []
    for cat_name, fields in categories.items():
        if any(customer_data.get(f) == "O" for f in fields):
            out.append(cat_name)
    return out


def calculate_fee(income, ledger_type, num_other_income, is_advance_booking=False):
    """수수료 계산.

    Returns:
        dict {
            'base_price': 정가 (구간),
            'other_income_fee': 타소득 가산,
            'total_full_price': 합산 정가,
            'discount_applied': 적용 할인,
            'final_fee': 최종 수수료,
            'sanity_warning': sanity check 경고,
        }
    """
    base = get_base_price(income, ledger_type)
    other_fee = num_other_income * OTHER_INCOME_PER_ITEM
    full_price = base + other_fee
    final = int(full_price * ADVANCE_DISCOUNT) if is_advance_booking else full_price
    discount = full_price - final

    # sanity check
    warnings = []
    if final < MIN_FEE:
        warnings.append(f"수수료가 {MIN_FEE:,}원 미만 (의심)")
    if final > MAX_FEE:
        warnings.append(f"수수료가 {MAX_FEE:,}원 초과 (의심)")

    return {
        "base_price": base,
        "other_income_fee": other_fee,
        "total_full_price": full_price,
        "discount_applied": discount,
        "final_fee": final,
        "sanity_warning": "; ".join(warnings) if warnings else "",
    }


def generate_quote_message(customer, fee_result, is_advance_booking=False):
    """카카오 알림톡 발송용 견적 메시지 (인간 검토용)"""
    name = customer.get("성명", "")
    income = customer.get("수입금액총계") or 0
    ledger = customer.get("기장의무", "")
    num_other = count_other_income(customer)

    other_labels = []
    for k in ["이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타"]:
        if customer.get(k) == "O":
            other_labels.append(k)

    discount_str = f" (사전접수 20% 할인 적용)" if is_advance_booking else ""

    msg = f"""[세무회계창연] 종합소득세 신고 견적 안내

{name}님 안녕하세요.
2024 귀속 종합소득세 신고 견적입니다.

▶ 사업소득: {income:,}원
▶ 장부 유형: {ledger}"""

    if other_labels:
        msg += f"\n▶ 타소득: {', '.join(other_labels)}"

    msg += f"""

[수수료 계산]
  정가 (구간): {fee_result['base_price']:,}원"""

    if fee_result['other_income_fee']:
        msg += f"\n  타소득 가산: +{fee_result['other_income_fee']:,}원 ({num_other}개 × 10만원)"

    msg += f"""
  합산 정가: {fee_result['total_full_price']:,}원"""

    if is_advance_booking:
        msg += f"\n  사전접수 할인: -{fee_result['discount_applied']:,}원"

    msg += f"""
  ─────────────────
  최종 수수료: {fee_result['final_fee']:,}원{discount_str}

수임 동의 후 아래 계좌로 입금 부탁드립니다.
[계좌번호 / 입금자명: {name}]
"""

    if fee_result['sanity_warning']:
        msg += f"\n\n⚠️ {fee_result['sanity_warning']}"

    return msg


# ===== 테스트 =====

def test_with_customers():
    """파싱결과.xlsx의 4명으로 견적 생성 테스트"""
    import openpyxl
    from pathlib import Path

    p = Path(r"F:\종소세2026\output\파싱결과.xlsx")
    if not p.exists():
        print(f"[에러] 파싱결과 없음: {p}")
        return

    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=i+1).value for i in range(len(headers))]
        customer = dict(zip(headers, row_values))

        if not customer.get("성명") or not customer.get("수입금액총계"):
            continue

        income = int(customer.get("수입금액총계") or 0)
        ledger = customer.get("기장의무", "")
        num_other = count_other_income(customer)

        try:
            for is_advance in [True, False]:  # 사전·일반 둘 다
                result = calculate_fee(income, ledger, num_other, is_advance)
                booking_label = "사전접수" if is_advance else "일반접수"
                print(f"\n{'='*60}")
                print(f"{customer['성명']} ({booking_label})")
                print(f"  사업소득: {income:,} / {ledger} / 타소득 {num_other}개")
                print(f"  정가: {result['base_price']:,} + 타소득 {result['other_income_fee']:,}")
                print(f"  합산: {result['total_full_price']:,} → 최종: {result['final_fee']:,}")
                if result['sanity_warning']:
                    print(f"  ⚠️ {result['sanity_warning']}")
        except Exception as e:
            print(f"\n{customer.get('성명')} - 계산 실패: {e}")


def test_quote_messages():
    """견적 메시지 1명 출력 (사전접수 가정)"""
    import openpyxl
    from pathlib import Path

    p = Path(r"F:\종소세2026\output\파싱결과.xlsx")
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    for row_idx in range(2, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=i+1).value for i in range(len(headers))]
        customer = dict(zip(headers, row_values))
        if not customer.get("성명") or not customer.get("수입금액총계"):
            continue

        income = int(customer.get("수입금액총계") or 0)
        ledger = customer.get("기장의무", "")
        num_other = count_other_income(customer)

        try:
            result = calculate_fee(income, ledger, num_other, is_advance_booking=True)
            msg = generate_quote_message(customer, result, is_advance_booking=True)
            print("\n" + "─" * 60)
            print(msg)
        except Exception as e:
            print(f"{customer.get('성명')} 메시지 실패: {e}")


if __name__ == "__main__":
    print("\n[1] 4명 수수료 계산 (사전·일반 둘 다)")
    test_with_customers()
    print("\n\n[2] 견적 메시지 샘플")
    test_quote_messages()
