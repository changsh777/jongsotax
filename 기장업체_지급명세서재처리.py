"""
기장업체_지급명세서재처리.py - 기장업체파싱.xlsx에서 지급명세서PDF=X or 간이용역=X인 대상만 재처리

처리: 지급명세서 PDF + 간이용역 xlsx (사업소득/기타소득) 다운로드
스킵: 지급명세서PDF=O AND 간이용역=O 이면 스킵
재개: python 기장업체_지급명세서재처리.py [시작번호]

전제조건:
  1. python launch_edge.py (Edge 디버그 창)
  2. 홈택스 세무사 계정 로그인
"""
import sys, io, os, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
os.environ.setdefault("SEOTAX_ENV", "nas")
sys.path.insert(0, r"F:\종소세2026")

from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.sync_api import sync_playwright

from 종합소득세안내문조회 import (
    _find_main_page,
    fill_jumin_and_search, normalize_jumin,
    download_jipgum_pdf,
    REPORT_HELP_URL,
)
from config import customer_folder

XLSX_PATH = Path(r"C:\Users\pc\OneDrive\문서\기장업체파싱.xlsx")


def load_customers(xlsx_path):
    """헤더 컬럼 위치 자동 감지 후 대표자+주민번호+처리상태+지급명세서PDF+간이용역 읽기"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    col = {}
    for i, h in enumerate(headers, 1):
        if h in ("대표자", "성명"):
            col["name"] = i
        elif h == "주민번호":
            col["jumin"] = i
        elif h == "처리상태":
            col["status"] = i
        elif h == "지급명세서PDF":
            col["jipgum"] = i
        elif h == "간이용역":
            col["ganiyiyong"] = i

    customers = []
    for row_idx in range(2, ws.max_row + 1):
        name  = str(ws.cell(row_idx, col["name"]).value or "").strip()
        jumin = str(ws.cell(row_idx, col["jumin"]).value or "").strip()
        if not name:
            continue
        jipgum    = str(ws.cell(row_idx, col.get("jipgum", 99)).value or "").strip()
        ganiyiyong = str(ws.cell(row_idx, col.get("ganiyiyong", 99)).value or "").strip()
        customers.append({
            "row_idx": row_idx,
            "name": name,
            "jumin_raw": jumin,
            "jipgum": jipgum,
            "ganiyiyong": ganiyiyong,
        })
    return wb, ws, col, customers


def main():
    start_idx = int(sys.argv[1]) - 1 if len(sys.argv) > 1 else 0

    if not XLSX_PATH.exists():
        print(f"[오류] 파일 없음: {XLSX_PATH}")
        sys.exit(1)

    wb, ws, col, customers = load_customers(XLSX_PATH)
    total = len(customers)

    # 처리 대상 필터: 지급명세서PDF ≠ O 인 경우만
    targets = [
        c for c in customers[start_idx:]
        if c["jipgum"] != "O"
    ]

    print(f"[기장업체 지급명세서 재처리] 총 {total}명 중 처리 대상 {len(targets)}명")
    print(f"  (지급명세서PDF=O 인 경우 스킵)\n")
    print(f"  Edge 디버그 창 + 세무사 계정 홈택스 로그인 확인 후 계속\n")

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://localhost:9222")
        ctx  = browser.contexts[0]
        page = _find_main_page(ctx)
        print(f"  [메인 페이지] {page.url[:80]}", flush=True)
        page.bring_to_front()
        page.on("dialog", lambda d: d.dismiss())

        for i, c in enumerate(targets, 1):
            name  = c["name"]
            jumin = str(c["jumin_raw"]).replace("-", "").replace(" ", "").strip()
            row_idx = c["row_idx"]

            print(f"[{i}/{len(targets)}] {name}  (xlsx행:{row_idx})", flush=True)

            if len(jumin) < 13:
                print(f"    [스킵] 주민번호 없음/불완전", flush=True)
                continue

            folder = customer_folder(name, jumin)

            try:
                front, back = normalize_jumin(jumin)
            except ValueError as e:
                print(f"    [스킵] 주민번호 오류: {e}", flush=True)
                continue

            # 홈택스 신고도움서비스 이동 + 주민번호 조회
            try:
                page.goto(REPORT_HELP_URL, wait_until="domcontentloaded", timeout=25000)
                time.sleep(2)
                fill_jumin_and_search(page, front, back)
                time.sleep(3)
            except Exception as e:
                print(f"    [에러] 조회 실패: {e}", flush=True)
                continue

            jip_ok = False

            # ─── 지급명세서 PDF ───
            try:
                result = download_jipgum_pdf(ctx, page, folder, name, jumin)
                jip_ok = bool(result)
                print(f"    [지급명세서] {'성공' if jip_ok else '자료없음/실패'}", flush=True)
            except Exception as e:
                print(f"    [에러] 지급명세서: {e}", flush=True)

            # ─── xlsx 업데이트 ───
            if col.get("jipgum"):
                ws.cell(row_idx, col["jipgum"], "O" if jip_ok else "X")
            ws.cell(row_idx, col.get("status", ws.max_column), "지급명세서재처리")
            wb.save(XLSX_PATH)
            print(f"    → 지급명세서:{jip_ok} 저장완료\n", flush=True)

    print(f"[완료] {len(targets)}명 처리 → {XLSX_PATH.name}")


if __name__ == "__main__":
    main()
