"""
jongsotaxbot.py - 종소세 작업 전용 텔레그램 봇 (@jongsotax_bot)

명령어:
  /work 강동수       NAS에서 파일 꺼내서 전송 (안내문+전년도+부가세+작업시트)
  /agree 강동수      진행 상태 조회
  /send 강동수       접수증+납부서 링크 발송 (게이트 포함)
  /pkg 강동수        출력패키지 PDF 재생성 (작업결과_이름.xls + 검증보고서 필요)

파일 업로드 자동 처리:
  25강동수신고서.pdf 업로드 → NAS 신고서.pdf 저장 + 교차검증 + 출력패키지 자동 생성
  작업결과_강동수.xls 업로드 → NAS 폴더 저장 + 업로드자+관리자 알림

자동 흐름 (신고서 업로드 시):
  [작업결과_이름.xls 있음] → 검증보고서(HTML) + 출력패키지 PDF (검증+소득+작업준비+안내문1p) 발송
  [작업결과_이름.xls 없음] → 검증보고서(HTML)만 발송 + "작업결과 엑셀 넣어주세요" 안내

실행: python3 ~/종소세2026/jongsotaxbot.py
"""

import asyncio
import os
import logging
import shutil
import tempfile
from pathlib import Path
from datetime import datetime
from telegram import Update
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters, ContextTypes
)
import sys as _sys
_sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import CUSTOMER_DIR as NAS_BASE
from print_sheet import nfc_glob

# ===== 설정 =====
TOKEN          = "REDACTED_TOKEN_1"
ADMIN_CHAT_ID  = 5980411081    # 세무사 (관리자)
NAS_URL        = "https://nas.taxenglab.com/종소세2026/고객"    # Cloudflare URL

ALLOWED_USERS: list[int] = []  # 빈 리스트 = 전체 허용

# 작업결과 엑셀 내 소득시트 목록 (이름.xls 에서 찾을 시트)
WORKPAN_SHEETS = {"프리", "사업자복식", "프리복식", "사업자+프리", "사업자+사업자", "프리+프리"}

LOG_FILE = os.path.expanduser("~/종소세2026/jongsotaxbot.log")

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
    handlers=[logging.FileHandler(LOG_FILE), logging.StreamHandler()],
)
logger = logging.getLogger(__name__)

# 동명이인 선택 대기 상태
user_pending: dict[int, dict] = {}


# ===== 유틸 =====
def nas_ok() -> bool:
    return NAS_BASE.exists()


def find_folders(name: str) -> list[Path]:
    """이름으로 고객 폴더 검색. 동명이인 있으면 여러 개 반환.
    NFD/NFC 양쪽 다 NFC로 정규화 후 비교 (Mac SMB 한글 인코딩 차이 회피)
    """
    import unicodedata
    name_nfc = unicodedata.normalize("NFC", name)
    prefix   = f"{name_nfc}_"
    try:
        return sorted(
            p for p in NAS_BASE.iterdir()
            if p.is_dir() and unicodedata.normalize("NFC", p.name).startswith(prefix)
        )
    except Exception:
        return []


def is_allowed(update: Update) -> bool:
    if not ALLOWED_USERS:
        return True
    return update.effective_user.id in ALLOWED_USERS


async def nas_fail(update: Update):
    await update.message.reply_text("⚠️ NAS 연결 끊김 — 관리자 확인 필요")
    logger.error("NAS 접근 실패: %s", NAS_BASE)


# ===== 동명이인 선택 =====
async def ask_choice(update: Update, user_id: int, folders: list[Path], action: str, extra=None):
    user_pending[user_id] = {"folders": folders, "action": action, "extra": extra}
    lines = "\n".join(f"{i+1}. {f.name}" for i, f in enumerate(folders))
    await update.message.reply_text(f"동명이인 확인:\n{lines}\n\n번호를 입력하세요.")


async def resolve_choice(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    user_id = update.effective_user.id
    if user_id not in user_pending:
        return False
    text = update.message.text.strip()
    try:
        idx = int(text) - 1
        pending = user_pending.pop(user_id)
        folders = pending["folders"]
        if not (0 <= idx < len(folders)):
            await update.message.reply_text("잘못된 번호입니다.")
            return True
        folder = folders[idx]
        action = pending["action"]
        extra  = pending["extra"]
        if action == "작업":
            await do_work(update, folder, force_jangbu=extra or "")
        elif action == "발송":
            await do_send(update, folder)
        elif action == "신고서":
            await do_save_singoser(update, context, folder, extra)
        elif action == "전기신고서":
            tg_file, fname = extra
            await do_save_전기신고서(update, folder, tg_file, fname)
        elif action == "수임동의":
            await do_status(update, folder)
        elif action == "출력패키지":
            await do_pkg(update, context, folder)
        elif action == "작업결과":
            tg_file, fname = extra
            await do_save_작업결과(update, context, folder, tg_file, fname)
    except ValueError:
        pass
    return True


JANGBU_KEYWORDS = ("간편장부", "복식부기")


def parse_name_arg(update: Update, context: ContextTypes.DEFAULT_TYPE) -> str:
    """명령어에서 이름 추출. /work강동수 또는 /work 강동수 둘 다 처리.
    장부유형 키워드(간편장부/복식부기)가 포함돼 있으면 제거 후 이름만 반환.
    """
    if context.args:
        tokens = [t for t in context.args if t not in JANGBU_KEYWORDS]
        return " ".join(tokens).strip()
    text = update.message.text or ""
    parts = text.split(None, 1)
    if len(parts) >= 2:
        remainder = parts[1].strip()
        tokens = [t for t in remainder.split() if t not in JANGBU_KEYWORDS]
        return " ".join(tokens).strip()
    cmd = parts[0].lstrip("/")
    for prefix in ("work", "작업", "agree", "수임동의", "send", "발송", "pkg", "출력패키지"):
        if cmd.startswith(prefix):
            return cmd[len(prefix):].strip()
    return ""


def parse_jangbu_arg(update: Update, context: ContextTypes.DEFAULT_TYPE) -> str:
    """명령어에서 장부유형 추출. '간편장부' 또는 '복식부기' 중 하나 반환. 없으면 빈 문자열."""
    if context.args:
        for token in context.args:
            if token in JANGBU_KEYWORDS:
                return token
        return ""
    text = update.message.text or ""
    for kw in JANGBU_KEYWORDS:
        if kw in text:
            return kw
    return ""


# ===== 출력패키지 생성 — 동기 함수들 (run_in_executor 용) =====

def _html_to_pdf_sync(html_path: Path, pdf_path: Path) -> bool:
    """playwright Chromium으로 HTML → A4 PDF 변환"""
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.goto(html_path.as_uri(), wait_until="networkidle")
            page.pdf(
                path=str(pdf_path),
                format="A4",
                print_background=True,
                margin={"top": "10mm", "bottom": "10mm",
                        "left": "10mm", "right": "10mm"},
            )
            browser.close()
        return True
    except Exception as e:
        logger.warning("[패키지] HTML→PDF 실패: %s", e)
        return False


def _sheet_to_pdf_sync(xls_path: Path, sheet_name: str, pdf_path: Path) -> bool:
    """xlwings로 특정 시트 → PDF 변환 (Windows + Excel 필요)"""
    app = None
    wb  = None
    try:
        import xlwings as xw
        app = xw.App(visible=False)
        app.display_alerts = False
        wb = app.books.open(str(xls_path))
        sheet = wb.sheets[sheet_name]
        sheet.api.ExportAsFixedFormat(
            Type=0,                     # xlTypePDF
            Filename=str(pdf_path),
            Quality=0,                  # xlQualityStandard
            IncludeDocProperties=True,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )
        return True
    except ImportError:
        logger.warning("[패키지] xlwings 없음 — Excel 시트 PDF 스킵 (Excel 설치 필요)")
        return False
    except Exception as e:
        logger.warning("[패키지] 시트 '%s' PDF 실패: %s", sheet_name, e)
        return False
    finally:
        if wb:
            try: wb.close()
            except Exception: pass
        if app:
            try: app.quit()
            except Exception: pass


def _sheet_to_pdf_libreoffice(xls_path: Path, sheet_name: str, pdf_path: Path) -> bool:
    """LibreOffice로 특정 시트 → PDF (macOS용, Excel 불필요)
    전략: 원본 그대로 전체 변환 → 해당 시트 인덱스 페이지 추출
    (시트 삭제 방식은 Named Range 깨짐 → #NAME? 오류 발생)
    """
    import subprocess, shutil as _sh, unicodedata as _ud
    try:
        import openpyxl, PyPDF2

        # 1. 시트 인덱스 파악 (NFC 정규화)
        wb = openpyxl.load_workbook(str(xls_path), read_only=True)
        nfc_names = [_ud.normalize("NFC", s) for s in wb.sheetnames]
        nfc_target = _ud.normalize("NFC", sheet_name)
        wb.close()
        if nfc_target not in nfc_names:
            logger.warning("[패키지] 시트 '%s' 없음 (목록: %s)", sheet_name, nfc_names)
            return False
        sheet_idx = nfc_names.index(nfc_target)

        # 2. 전체 xlsx → PDF (원본 그대로 — Named Range 보존)
        tmp_dir = Path(tempfile.mkdtemp(prefix="lo_pdf_"))
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", str(tmp_dir), str(xls_path)],
            capture_output=True, timeout=90, text=True
        )
        tmp_pdf = tmp_dir / (xls_path.stem + ".pdf")
        if not tmp_pdf.exists():
            logger.warning("[패키지] LibreOffice 변환 실패 (stderr: %s)", result.stderr[:300])
            _sh.rmtree(tmp_dir, ignore_errors=True)
            return False

        # 3. 해당 시트 페이지 추출 (시트 1개 = PDF 1페이지 가정)
        reader = PyPDF2.PdfReader(str(tmp_pdf))
        total = len(reader.pages)
        logger.info("[패키지] 전체 변환 %d페이지, 시트 '%s' 인덱스=%d", total, sheet_name, sheet_idx)

        if sheet_idx >= total:
            logger.warning("[패키지] 페이지 수(%d) < 시트 인덱스(%d) — 마지막 페이지 사용", total, sheet_idx)
            sheet_idx = total - 1

        writer = PyPDF2.PdfWriter()
        writer.add_page(reader.pages[sheet_idx])
        with open(pdf_path, "wb") as f:
            writer.write(f)

        _sh.rmtree(tmp_dir, ignore_errors=True)
        return True
    except Exception as e:
        logger.warning("[패키지] LibreOffice 시트 PDF 실패 (%s): %s", sheet_name, e)
        return False


def _extract_first_page_sync(pdf_in: Path, pdf_out: Path) -> bool:
    """PDF 첫 페이지만 추출"""
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(str(pdf_in))
        writer = PyPDF2.PdfWriter()
        writer.add_page(reader.pages[0])
        with open(pdf_out, "wb") as f:
            writer.write(f)
        return True
    except Exception as e:
        logger.warning("[패키지] PDF 첫 페이지 추출 실패: %s", e)
        return False


def _merge_pdfs_sync(pdf_list: list, out_path: Path) -> bool:
    """여러 PDF → 하나로 합치기"""
    try:
        import PyPDF2
        merger = PyPDF2.PdfMerger()
        for p in pdf_list:
            merger.append(str(p))
        with open(out_path, "wb") as f:
            merger.write(f)
        merger.close()
        return True
    except Exception as e:
        logger.warning("[패키지] PDF 합치기 실패: %s", e)
        return False


def _make_print_package_sync(folder: Path, name: str, html_path: Path, xls_path: Path):
    """
    출력패키지 PDF 생성 (동기 — run_in_executor 로 호출):
      1. 검증보고서 (HTML → PDF)
      2. 소득시트   (작업결과_이름.xls WORKPAN_SHEETS 시트 → PDF)
      3. 작업준비시트 (작업결과_이름.xls 작업준비_* 시트 → PDF)
      4. 안내문 1페이지
      5. 신고서.pdf (당기)
      → 합쳐서 출력패키지_{이름}_{날짜}.pdf 저장 후 경로 반환
    """
    ts     = datetime.now().strftime("%Y%m%d_%H%M")
    tmpdir = Path(tempfile.mkdtemp(prefix="print_pkg_"))
    pdf_parts: list[Path] = []

    try:
        # ─ 1. 검증보고서 HTML → PDF ─────────────────────────────
        pdf_check = tmpdir / "01_검증보고서.pdf"
        if _html_to_pdf_sync(html_path, pdf_check):
            pdf_parts.append(pdf_check)
            logger.info("[패키지] 검증보고서 PDF 완료")
        else:
            logger.warning("[패키지] 검증보고서 PDF 실패 — 스킵")

        # ─ 2. 작업결과 엑셀 시트 → PDF ──────────────────────────
        # Windows: xlwings + Excel / macOS: LibreOffice (Excel 불필요)
        import platform
        _is_mac = platform.system() == "Darwin"

        if _is_mac:
            # ── macOS: LibreOffice ──────────────────────────────
            try:
                import openpyxl, unicodedata as _ud
                wb_tmp = openpyxl.load_workbook(str(xls_path), read_only=True, data_only=True)
                # NFD→NFC 정규화 (macOS SMB 파일명 대응)
                sheet_names = [_ud.normalize("NFC", s) for s in wb_tmp.sheetnames]
                wb_tmp.close()
                logger.info("[패키지] 시트 목록(NFC): %s", sheet_names)

                workpan = next((s for s in sheet_names if s in WORKPAN_SHEETS), None)
                if workpan:
                    pdf_wp = tmpdir / "02_소득시트.pdf"
                    if _sheet_to_pdf_libreoffice(xls_path, workpan, pdf_wp):
                        pdf_parts.append(pdf_wp)
                        logger.info("[패키지] 소득시트 '%s' PDF 완료", workpan)
                    else:
                        logger.warning("[패키지] 소득시트 '%s' PDF 실패", workpan)
                else:
                    logger.warning("[패키지] 소득시트 없음 (시트 목록: %s)", sheet_names)

                junbi = next((s for s in sheet_names if s.startswith("작업준비_")), None)
                if junbi:
                    pdf_jj = tmpdir / "03_작업준비.pdf"
                    if _sheet_to_pdf_libreoffice(xls_path, junbi, pdf_jj):
                        pdf_parts.append(pdf_jj)
                        logger.info("[패키지] 작업준비 '%s' PDF 완료", junbi)
                    else:
                        logger.warning("[패키지] 작업준비 '%s' PDF 실패", junbi)
            except Exception as e:
                logger.warning("[패키지] macOS Excel PDF 처리 실패: %s", e)

        else:
            # ── Windows: xlwings + Excel ────────────────────────
            try:
                import xlwings as xw
                _app = xw.App(visible=False, add_book=False)
                _app.display_alerts = False
                try:
                    _wb = _app.books.open(str(xls_path))
                    sheet_names = [s.name for s in _wb.sheets]
                    logger.info("[패키지] 시트 목록: %s", sheet_names)

                    workpan = next((s for s in sheet_names if s in WORKPAN_SHEETS), None)
                    if workpan:
                        pdf_wp = tmpdir / "02_소득시트.pdf"
                        try:
                            _wb.sheets[workpan].api.ExportAsFixedFormat(
                                Type=0, Filename=str(pdf_wp),
                                Quality=0, IncludeDocProperties=True,
                                IgnorePrintAreas=False, OpenAfterPublish=False,
                            )
                            pdf_parts.append(pdf_wp)
                            logger.info("[패키지] 소득시트 '%s' PDF 완료", workpan)
                        except Exception as e:
                            logger.warning("[패키지] 소득시트 '%s' PDF 실패: %s", workpan, e)
                    else:
                        logger.warning("[패키지] 소득시트 없음 (시트 목록: %s)", sheet_names)

                    junbi = next((s for s in sheet_names if s.startswith("작업준비_")), None)
                    if junbi:
                        pdf_jj = tmpdir / "03_작업준비.pdf"
                        try:
                            _wb.sheets[junbi].api.ExportAsFixedFormat(
                                Type=0, Filename=str(pdf_jj),
                                Quality=0, IncludeDocProperties=True,
                                IgnorePrintAreas=False, OpenAfterPublish=False,
                            )
                            pdf_parts.append(pdf_jj)
                            logger.info("[패키지] 작업준비 '%s' PDF 완료", junbi)
                        except Exception as e:
                            logger.warning("[패키지] 작업준비 '%s' PDF 실패: %s", junbi, e)
                finally:
                    try: _wb.close()
                    except Exception: pass
                    try: _app.quit()
                    except Exception: pass
            except ImportError:
                logger.warning("[패키지] xlwings 없음 — Excel 시트 PDF 스킵")
            except Exception as e:
                logger.warning("[패키지] Windows Excel PDF 처리 실패: %s", e)

        # ─ 3. 안내문 첫 페이지 ──────────────────────────────────
        ann_files = sorted(nfc_glob(folder, "종소세안내문_*.pdf"),
                           key=lambda p: p.stat().st_mtime, reverse=True)
        if ann_files:
            pdf_ann = tmpdir / "04_안내문1p.pdf"
            if _extract_first_page_sync(ann_files[0], pdf_ann):
                pdf_parts.append(pdf_ann)
                logger.info("[패키지] 안내문 1페이지 완료")

        # ─ 4. 신고서.pdf (당기) ─────────────────────────────────
        singoser = folder / "신고서.pdf"
        if singoser.exists():
            pdf_sg = tmpdir / "05_신고서.pdf"
            try:
                import shutil as _sh
                _sh.copy2(str(singoser), str(pdf_sg))
                pdf_parts.append(pdf_sg)
                logger.info("[패키지] 신고서 추가 완료")
            except Exception as e:
                logger.warning("[패키지] 신고서 복사 실패: %s", e)

        # ─ 4. 합치기 ────────────────────────────────────────────
        if not pdf_parts:
            logger.warning("[패키지] 합칠 PDF 없음")
            return None

        # 기존 출력패키지 archive
        old_pkgs = nfc_glob(folder, "출력패키지_*.pdf")
        if old_pkgs:
            arch = folder / "_archive"
            arch.mkdir(exist_ok=True)
            for op in old_pkgs:
                try:
                    op.rename(arch / op.name)
                except Exception:
                    pass

        out_path = folder / f"출력패키지_{name}_{ts}.pdf"
        if _merge_pdfs_sync(pdf_parts, out_path):
            logger.info("[패키지] 저장 완료: %s (%d개 파트)", out_path.name, len(pdf_parts))
            return out_path
        return None

    finally:
        shutil.rmtree(tmpdir, ignore_errors=True)


# ===== 발송 헬퍼 (HTML 검증보고서 / 출력패키지 PDF) =====

async def _send_html_report(context: ContextTypes.DEFAULT_TYPE, update: Update,
                             html_path: Path, folder: Path, sender_id: int):
    """검증보고서 HTML을 발신자 + 관리자에게 전송"""
    caption = f"📊 {folder.name} 검증보고서"
    with open(html_path, "rb") as fp:
        await context.bot.send_document(chat_id=sender_id, document=fp,
                                        filename=html_path.name, caption=caption)
    if sender_id != ADMIN_CHAT_ID:
        with open(html_path, "rb") as fp:
            await context.bot.send_document(
                chat_id=ADMIN_CHAT_ID, document=fp, filename=html_path.name,
                caption=f"{caption} (직원: {update.effective_user.full_name})"
            )


async def _send_package(context: ContextTypes.DEFAULT_TYPE, update: Update,
                        pkg_path: Path, folder: Path, sender_id: int):
    """출력패키지 PDF를 발신자 + 관리자에게 전송"""
    try:
        import PyPDF2
        pages = len(PyPDF2.PdfReader(str(pkg_path)).pages)
    except Exception:
        pages = "?"
    caption = f"📋 {folder.name} 출력패키지 ({pages}p)"
    with open(pkg_path, "rb") as fp:
        await context.bot.send_document(chat_id=sender_id, document=fp,
                                        filename=pkg_path.name, caption=caption)
    if sender_id != ADMIN_CHAT_ID:
        with open(pkg_path, "rb") as fp:
            await context.bot.send_document(
                chat_id=ADMIN_CHAT_ID, document=fp, filename=pkg_path.name,
                caption=f"{caption} (직원: {update.effective_user.full_name})"
            )


# ===== /작업 =====
async def cmd_work(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    name     = parse_name_arg(update, context)
    jangbu   = parse_jangbu_arg(update, context)
    if not name:
        await update.message.reply_text(
            "사용법: /work 강동수  또는  /작업 강동수\n"
            "장부유형 지정: /work 강동수 복식부기  또는  /work 강동수 간편장부"
        )
        return
    if not nas_ok():
        await nas_fail(update); return

    folders = find_folders(name)

    if not folders:
        await update.message.reply_text(f"'{name}' 고객 자료가 없습니다. 홈택스 안내문 파싱이 완료됐나요?"); return
    if len(folders) > 1:
        # 동명이인 선택 대기 — jangbu 는 extra 로 전달
        user_pending[update.effective_user.id] = {
            "folders": folders, "action": "작업", "extra": jangbu
        }
        lines = "\n".join(f"{i+1}. {f.name}" for i, f in enumerate(folders))
        await update.message.reply_text(f"동명이인 확인:\n{lines}\n\n번호를 입력하세요.")
        return

    await do_work(update, folders[0], force_jangbu=jangbu)


async def do_work(update: Update, folder: Path, force_jangbu: str = ""):
    """안내문 + 전년도자료 + 작업판 + 지급명세서 + 간이용역소득 전송.

    force_jangbu: '간편장부' 또는 '복식부기' 강제 지정 가능. 없으면 안내문 자동 감지.
    항상 작업판을 재생성 후 전송.
    """
    import unicodedata
    def nfc(s): return unicodedata.normalize("NFC", str(s))

    # ── 작업판 항상 재생성 ─────────────────────────────────────────
    try:
        import sys as _sys
        _proj = str(Path(__file__).resolve().parent)
        if _proj not in _sys.path:
            _sys.path.insert(0, _proj)
        from jakupan_gen import make_jakupan

        parts  = folder.name.rsplit("_", 1)
        _name  = parts[0]
        jumin6 = parts[1][:6] if len(parts) > 1 else ""
        out = make_jakupan(_name, jumin6, force_jangbu=force_jangbu)
        if not out:
            await update.message.reply_text("⚠️ 작업판 생성 실패 — 기존 파일이 있으면 그대로 전송합니다.")
    except Exception as e:
        logger.error("[do_work] 작업판 재생성 오류: %s", e, exc_info=True)
        await update.message.reply_text(f"⚠️ 작업판 재생성 오류: {e}\n기존 파일이 있으면 그대로 전송합니다.")

    files_to_send = []
    root_files = sorted([f for f in folder.iterdir() if f.is_file()],
                        key=lambda f: f.stat().st_mtime, reverse=True)

    # 1. 안내문 (최신 1개)
    ann = [f for f in root_files if "종소세안내문" in nfc(f.name) and f.suffix == ".pdf"]
    if ann:
        files_to_send.append(ann[0])

    # 2. 전년도 자료 (전년도종소세신고내역 엑셀 + 수동 저장한 전년도 신고서 PDF)
    prev = [f for f in root_files if "전년도종소세신고내역" in nfc(f.name)]
    files_to_send.extend(prev)

    # 전년도 신고서 PDF (예: 20250513_2024신고서_홍길동.pdf — 신고서.pdf 제외)
    prev_pdf = [f for f in root_files
                if "신고서" in nfc(f.name) and f.suffix == ".pdf"
                and f.name != "신고서.pdf"]
    files_to_send.extend(prev_pdf)

    # ⚠️ 전년도 신고서 없음 알림 (신규 고객 — 수동 업로드 필요)
    if not prev_pdf:
        parts = folder.name.rsplit("_", 1)
        _name = parts[0]
        await update.message.reply_text(
            f"⚠️ *{_name}* 전년도 신고서 없음\n"
            f"신규 고객이면 전년도 신고서 PDF를 NAS `{folder.name}/` 폴더에 넣어주세요.",
            parse_mode="Markdown"
        )

    # 3. 작업판 엑셀 (최신 1개)
    wp = [f for f in root_files if nfc(f.name).startswith("작업판_") and f.suffix == ".xlsx"]
    if wp:
        files_to_send.append(wp[0])

    # 4. 지급명세서 + 간이용역소득 폴더
    for sub in ["지급명세서", "간이용역소득"]:
        d = next((p for p in folder.iterdir() if p.is_dir() and nfc(p.name) == sub), None)
        if d:
            files_to_send.extend(sorted(d.iterdir()))

    files_to_send = [f for f in files_to_send if f.is_file()]
    if not files_to_send:
        await update.message.reply_text(f"{folder.name}: 작업 파일 없음 (홈택스 안내문 파싱 완료됐나요?)"); return

    await update.message.reply_text(f"📁 {folder.name} — {len(files_to_send)}개 전송 중...")
    for f in files_to_send:
        try:
            # xlsx → xls 확장자 변환 (파일명만, 바이트 그대로)
            fname = f.name[:-1] if f.name.endswith(".xlsx") else f.name
            with open(f, "rb") as fp:
                await update.message.reply_document(document=fp, filename=fname)
        except Exception as e:
            await update.message.reply_text(f"❌ {f.name} 실패: {e}")
    await update.message.reply_text("✅ 전송 완료")


# ===== /수임동의 =====
async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    name = parse_name_arg(update, context)
    if not name:
        await update.message.reply_text("사용법: /agree 강동수  또는  /수임동의 강동수"); return
    if not nas_ok():
        await nas_fail(update); return

    folders = find_folders(name)

    if not folders:
        await update.message.reply_text(f"'{name}' 폴더 없음"); return
    if len(folders) > 1:
        await ask_choice(update, update.effective_user.id, folders, "수임동의"); return
    await do_status(update, folders[0])


async def do_status(update: Update, folder: Path):
    def chk(pattern):
        return bool(nfc_glob(folder, pattern))

    parts = folder.name.rsplit("_", 1)
    _name = parts[0]

    items = {
        "안내문 파싱":    chk("종소세안내문_*.pdf"),
        "신고서":         (folder / "신고서.pdf").exists(),
        "검증보고서":     chk("검증보고서_*.html"),
        "작업결과 엑셀":  chk("작업결과_*.xls") or chk("작업결과_*.xlsx"),
        "출력패키지":     chk("출력패키지_*.pdf"),
        "접수증":         (folder / "접수증.pdf").exists(),
        "납부서":         (folder / "납부서.pdf").exists(),
    }
    lines = "\n".join(f"{'✅' if v else '❌'} {k}" for k, v in items.items())
    await update.message.reply_text(f"📋 {folder.name}\n{lines}")


# ===== /발송 =====
async def cmd_send(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    name = parse_name_arg(update, context)
    if not name:
        await update.message.reply_text("사용법: /send 강동수  또는  /발송 강동수"); return
    if not nas_ok():
        await nas_fail(update); return

    folders = find_folders(name)

    if not folders:
        await update.message.reply_text(f"'{name}' 폴더 없음"); return
    if len(folders) > 1:
        await ask_choice(update, update.effective_user.id, folders, "발송"); return
    await do_send(update, folders[0])


async def do_send(update: Update, folder: Path):
    """접수증 + 납부서 링크 발송 — 게이트 ⑤⑥ 포함"""
    접수증 = folder / "접수증.pdf"
    납부서 = folder / "납부서.pdf"

    # 게이트 ⑥: 접수증 존재 확인
    if not 접수증.exists():
        await update.message.reply_text("❌ 접수증이 없습니다. 최종신고 완료 후 다시 시도하세요.")
        return

    # 게이트 ⑥: 당일 스크래핑 확인
    today = datetime.now().date()
    mtime = datetime.fromtimestamp(접수증.stat().st_mtime).date()
    if mtime != today:
        await update.message.reply_text(
            f"⚠️ 접수증이 오늘({today}) 파일이 아닙니다 (파일날짜: {mtime})\n"
            "오늘 최종신고 후 다시 시도하세요."
        )
        return

    url = f"{NAS_URL}/{folder.name}"
    lines = [f"📄 접수증: {url}/접수증.pdf"]

    # 게이트 ⑤: 납부서 존재 확인 (소득세>0인 경우 필수)
    if 납부서.exists():
        lines.append(f"💳 납부서: {url}/납부서.pdf")
    else:
        lines.append("ℹ️ 납부서 없음 (환급이거나 누락 — 세무사 확인)")

    await update.message.reply_text(
        f"✅ {folder.name} 발송 링크:\n\n" + "\n".join(lines) +
        "\n\n→ 솔라피 알림톡 발송하세요."
    )


# ===== /pkg (출력패키지 재생성) =====
async def cmd_pkg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    name = parse_name_arg(update, context)
    if not name:
        await update.message.reply_text("사용법: /pkg 강동수"); return
    if not nas_ok():
        await nas_fail(update); return

    folders = find_folders(name)
    if not folders:
        await update.message.reply_text(f"'{name}' 폴더 없음"); return
    if len(folders) > 1:
        await ask_choice(update, update.effective_user.id, folders, "출력패키지"); return
    await do_pkg(update, context, folders[0])


async def do_pkg(update: Update, context: ContextTypes.DEFAULT_TYPE, folder: Path):
    """출력패키지 PDF 재생성 — /pkg 명령어 또는 내부 호출"""
    parts = folder.name.rsplit("_", 1)
    _name = parts[0]
    xls_files = sorted(
        nfc_glob(folder, "작업결과_*.xls") + nfc_glob(folder, "작업결과_*.xlsx"),
        key=lambda p: p.stat().st_mtime, reverse=True
    )
    xls_path = xls_files[0] if xls_files else None

    if not xls_path:
        await update.message.reply_text(
            f"❌ 작업결과 엑셀 없음\n"
            f"NAS `{folder.name}/` 폴더에 *작업결과_{_name}.xls* 를 먼저 넣어주세요.",
            parse_mode="Markdown"
        )
        return

    # 검증보고서 HTML 찾기
    html_files = sorted(nfc_glob(folder, "검증보고서_*.html"),
                        key=lambda p: p.stat().st_mtime, reverse=True)
    if not html_files:
        await update.message.reply_text(
            f"❌ 검증보고서 없음\n"
            f"신고서.pdf를 먼저 업로드해 검증을 실행해주세요."
        )
        return

    html_path = html_files[0]
    sender_id = update.effective_chat.id
    await update.message.reply_text(f"📦 {_name} 출력패키지 생성 중 (잠시 대기)...")

    loop = asyncio.get_running_loop()
    pkg_path = await loop.run_in_executor(
        None, _make_print_package_sync, folder, _name, html_path, xls_path
    )

    if pkg_path and pkg_path.exists():
        await _send_package(context, update, pkg_path, folder, sender_id)
    else:
        await update.message.reply_text("⚠️ 출력패키지 생성 실패 — 로그 확인 필요")


# ===== 파일 수신 =====
async def handle_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return

    # 동명이인 선택 대기 중이면 먼저 처리
    if await resolve_choice(update, context):
        return

    doc = update.message.document
    if not doc:
        return

    import unicodedata
    fname = doc.file_name or ""
    fname_nfc   = unicodedata.normalize("NFC", fname)   # Mac NFD 대응
    fname_lower = fname_nfc.lower()

    # ── 작업결과 엑셀: 작업결과_이름.xls / .xlsx ──────────────────
    is_작업결과 = (
        fname_lower.startswith("작업결과_") and
        (fname_lower.endswith(".xls") or fname_lower.endswith(".xlsx"))
    )
    if is_작업결과:
        stem = Path(fname_nfc).stem             # 작업결과_홍길동
        name = stem[len("작업결과_"):]         # 홍길동
        if not name:
            await update.message.reply_text("파일명 오류: 작업결과_이름.xls 형식으로 올려주세요.")
            return
        if not nas_ok():
            await nas_fail(update); return
        folders = find_folders(name)
        if not folders:
            await update.message.reply_text(f"'{name}' 폴더 없음 — 이름 확인 필요"); return
        tg_file = await doc.get_file()
        if len(folders) > 1:
            await ask_choice(update, update.effective_user.id, folders, "작업결과", (tg_file, fname))
            return
        await do_save_작업결과(update, context, folders[0], tg_file, fname)
        return

    # ── 신고서 PDF: 25이름신고서.pdf / 2024전기신고서 ──────────────
    if not fname_lower.endswith(".pdf"):
        return

    # 25년 신고서 또는 2024 전기신고서 처리
    is_25_singoser = "25" in fname and "신고서" in fname_lower
    is_전기신고서  = "2024" in fname and "신고서" in fname_lower

    if not is_25_singoser and not is_전기신고서:
        return  # 조건 불만족 → 무시

    caption = (update.message.caption or "").strip()

    if is_25_singoser:
        # 고객명: 캡션 우선, 없으면 파일명 마지막 _뒤에서 추출
        # ex) "25년 종합소득세신고서_김지은" → "김지은"
        if caption:
            name = caption
        else:
            stem = Path(fname).stem.strip()
            name = stem.rsplit("_", 1)[-1] if "_" in stem else stem
    else:
        # 전기신고서: 캡션 우선, 없으면 _이름 패턴 우선, 없으면 키워드 제거
        # ex) "2024 종합소득세신고서_김지혁" → "김지혁"
        import re
        if caption:
            name = caption
        else:
            stem = Path(fname).stem
            if '_' in stem:
                name = stem.rsplit('_', 1)[-1].strip()
            else:
                cleaned = re.sub(r'2024|종합소득세|신고서', '', stem)
                name = cleaned.strip()

    if not nas_ok():
        await nas_fail(update); return

    folders = find_folders(name)
    if not folders:
        await update.message.reply_text(f"'{name}' 폴더 없음 — 이름 확인 필요"); return

    tg_file = await doc.get_file()

    if is_25_singoser:
        if len(folders) > 1:
            await ask_choice(update, update.effective_user.id, folders, "신고서", tg_file)
            return
        await do_save_singoser(update, context, folders[0], tg_file)
    else:
        if len(folders) > 1:
            await ask_choice(update, update.effective_user.id, folders, "전기신고서",
                             (tg_file, fname))
            return
        await do_save_전기신고서(update, folders[0], tg_file, fname)


async def do_save_singoser(update: Update, context: ContextTypes.DEFAULT_TYPE, folder: Path, tg_file):
    """
    신고서.pdf 저장 → 교차검증 → 출력패키지 생성(이름.xls 있을 때) → 발신자+관리자 발송
    [ULTRA CRITICAL] 기존 신고서는 반드시 _archive 이동 후 새 파일 저장
    """
    target  = folder / "신고서.pdf"
    archive = folder / "_archive"

    # 기존 파일 archive 이동
    if target.exists():
        archive.mkdir(exist_ok=True)
        ts = datetime.fromtimestamp(target.stat().st_mtime).strftime("%Y%m%d_%H%M%S")
        target.rename(archive / f"신고서_{ts}.pdf")
        logger.info("archive 이동: 신고서_%s.pdf", ts)

    await tg_file.download_to_drive(str(target))
    logger.info("신고서 저장: %s", target)
    await update.message.reply_text(f"✅ {folder.name}/신고서.pdf 저장\n⏳ 교차검증 실행 중...")

    # ── 교차검증 실행 ──────────────────────────────────────────────
    html_path = None
    try:
        import sys as _sys
        _bot_dir = str(Path(__file__).resolve().parent)
        if _bot_dir not in _sys.path:
            _sys.path.insert(0, _bot_dir)
        from tax_cross_verify import run as verify_run

        parts  = folder.name.rsplit("_", 1)
        _name  = parts[0]
        _jumin = parts[1] if len(parts) > 1 else ""

        html_path = verify_run(_name, _jumin, folder=folder)

    except Exception as e:
        logger.error("검증 오류: %s", e, exc_info=True)
        await update.message.reply_text(f"⚠️ 검증 오류: {e}\n(신고서는 저장됐습니다)")
        return

    if not html_path or not html_path.exists():
        await update.message.reply_text("⚠️ 검증보고서 생성 실패 — 수동 실행 필요")
        return

    # ── 작업결과 엑셀 확인 → 출력패키지 또는 검증보고서만 ──────────
    parts     = folder.name.rsplit("_", 1)
    _name     = parts[0]
    xls_files = sorted(
        nfc_glob(folder, "작업결과_*.xls") + nfc_glob(folder, "작업결과_*.xlsx"),
        key=lambda p: p.stat().st_mtime, reverse=True
    )
    xls_path  = xls_files[0] if xls_files else None
    sender_id = update.effective_chat.id

    # 검증보고서(HTML)는 항상 전송 (인터랙티브 확인용)
    await _send_html_report(context, update, html_path, folder, sender_id)

    if not xls_path:
        # 작업결과 엑셀 없음 → 안내 메시지
        await update.message.reply_text(
            f"📝 출력패키지를 만들려면 *작업결과_{_name}.xls* (작업결과 엑셀)를\n"
            f"텔레그램으로 업로드해주세요.\n"
            f"업로드 후 자동 저장되며, /pkg {_name} 명령어로 패키지 생성 가능합니다.",
            parse_mode="Markdown"
        )
    else:
        # 작업결과 엑셀 있음 → 출력패키지 자동 생성
        await update.message.reply_text(f"📦 {xls_path.name} 발견! 출력패키지 생성 중 (잠시 대기)...")
        loop = asyncio.get_running_loop()
        pkg_path = await loop.run_in_executor(
            None, _make_print_package_sync, folder, _name, html_path, xls_path
        )
        if pkg_path and pkg_path.exists():
            await _send_package(context, update, pkg_path, folder, sender_id)
        else:
            await update.message.reply_text(
                "⚠️ 출력패키지 생성 실패 — 검증보고서만 전송됐습니다\n"
                "Excel 설치 여부 및 시트명 확인 후 /pkg 명령어로 재시도하세요."
            )


async def do_save_전기신고서(update: Update, folder: Path, tg_file, fname: str):
    """전기신고서 PDF 폴더 저장 (교차검증 없음)"""
    target  = folder / fname
    archive = folder / "_archive"

    if target.exists():
        archive.mkdir(exist_ok=True)
        ts = datetime.fromtimestamp(target.stat().st_mtime).strftime("%Y%m%d_%H%M%S")
        target.rename(archive / f"{Path(fname).stem}_{ts}.pdf")
        logger.info("archive 이동: %s_%s.pdf", Path(fname).stem, ts)

    await tg_file.download_to_drive(str(target))
    logger.info("전기신고서 저장: %s", target)
    await update.message.reply_text(f"✅ {folder.name}/{fname} 저장 완료")


async def do_save_작업결과(update: Update, context: ContextTypes.DEFAULT_TYPE,
                           folder: Path, tg_file, fname: str):
    """
    작업결과_이름.xls 저장 → 업로드자 reply + 관리자 DM
    [ULTRA CRITICAL] 기존 작업결과_*.xls* → _archive 이동 후 저장
    """
    archive = folder / "_archive"

    # 기존 작업결과 파일 archive 이동
    old_files = nfc_glob(folder, "작업결과_*.xls") + nfc_glob(folder, "작업결과_*.xlsx")
    if old_files:
        archive.mkdir(exist_ok=True)
        for old in old_files:
            ts = datetime.fromtimestamp(old.stat().st_mtime).strftime("%Y%m%d_%H%M%S")
            new_name = f"{old.stem}_{ts}{old.suffix}"
            try:
                old.rename(archive / new_name)
                logger.info("archive 이동: %s → %s", old.name, new_name)
            except Exception as e:
                logger.warning("archive 이동 실패: %s — %s", old.name, e)

    target = folder / fname
    await tg_file.download_to_drive(str(target))
    logger.info("작업결과 저장: %s", target)

    parts = folder.name.rsplit("_", 1)
    _name = parts[0]
    sender_id = update.effective_chat.id
    sender_name = update.effective_user.full_name or update.effective_user.username or str(sender_id)

    # 업로드자에게 확인
    await update.message.reply_text(
        f"✅ *{_name}* 작업결과 저장 완료\n"
        f"📁 `{folder.name}/{fname}`\n"
        f"신고서 업로드 후 출력패키지가 자동 생성됩니다.",
        parse_mode="Markdown"
    )

    # 관리자에게 DM (업로드자가 관리자가 아닌 경우)
    if sender_id != ADMIN_CHAT_ID:
        await context.bot.send_message(
            chat_id=ADMIN_CHAT_ID,
            text=(
                f"📥 *{_name}* 작업결과 저장됨\n"
                f"업로드: {sender_name}\n"
                f"파일: `{fname}`"
            ),
            parse_mode="Markdown"
        )
        logger.info("관리자 알림 전송: %s 작업결과 (업로드: %s)", _name, sender_name)


# ===== 자정 배치: 혼입 검증 =====
async def job_integrity_check(context: ContextTypes.DEFAULT_TYPE):
    """매일 자정 verify_folder_integrity.py --fix 실행 → ADMIN에게 결과 전송"""
    import subprocess
    script = Path(__file__).resolve().parent / "verify_folder_integrity.py"
    try:
        result = subprocess.run(
            [_sys.executable, str(script), "--fix"],
            capture_output=True, text=True, encoding="utf-8", timeout=600
        )
        output = (result.stdout or result.stderr or "(출력 없음)").strip()
    except Exception as e:
        output = f"실행 오류: {e}"

    header = "혼입 검증 결과\n\n"
    msg = header + output
    MAX = 4096
    while msg:
        await context.bot.send_message(chat_id=ADMIN_CHAT_ID, text=msg[:MAX])
        msg = msg[MAX:]


# ===== 텍스트: 동명이인 번호 선택 =====
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update): return
    await resolve_choice(update, context)


# ===== 메인 =====
def main():
    if not nas_ok():
        logger.warning("NAS 미연결: %s", NAS_BASE)

    app = Application.builder().token(TOKEN).build()

    # 매일 자정(KST) 혼입 검증 — UTC 15:00 = KST 00:00
    from datetime import time as _dtime
    app.job_queue.run_daily(job_integrity_check, time=_dtime(hour=15, minute=0))

    app.add_handler(CommandHandler("work",    cmd_work))          # /work 강동수
    app.add_handler(CommandHandler("agree",   cmd_status))        # /agree 강동수
    app.add_handler(CommandHandler("send",    cmd_send))          # /send 강동수
    app.add_handler(CommandHandler("pkg",     cmd_pkg))           # /pkg 강동수 (출력패키지 재생성)
    app.add_handler(MessageHandler(filters.Document.ALL,            handle_file))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))

    logger.info("jongsotax_bot 시작")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
