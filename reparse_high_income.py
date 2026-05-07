"""
reparse_high_income.py — 수입금액 전체 재파싱 후 1억↑ 변경분 반영
세무회계창연 | 2026

실행:
    python3 ~/종소세2026/reparse_high_income.py

동작:
  1) 모든 고객 안내문 PDF 재파싱 (줄바꿈 버그 수정판 적용)
  2) 파싱결과.xlsx 수입금액총계 업데이트
  3) 접수명단 구글시트 수입 컬럼 업데이트
  4) 변경된 고객 목록 출력
"""

import sys, os, fnmatch
os.environ.setdefault("SEOTAX_ENV", "nas")
sys.path.insert(0, str(__import__("pathlib").Path(__file__).parent))

import unicodedata
import gspread
from pathlib import Path
from datetime import datetime
from gspread.utils import rowcol_to_a1
import openpyxl

from config import CUSTOMER_DIR, PARSE_RESULT_XLSX
from parse_to_xlsx import parse_anneam, COLUMNS
from fee_calculator import calculate_fee, count_other_income
from gsheet_writer import get_credentials
from jakupan_gen import make_jakupan

SPREADSHEET_ID = "1oh31k00Oa2lZWvu5fnBRVmurdlll1YEG8Fefi5FRfBI"
FORCE_NAMES    = ["김지혁"]   # 금액과 무관하게 강제 재처리할 고객

def _nfc(s):
    return unicodedata.normalize("NFC", s)

def _glob_nfc(folder, pattern):
    """macOS SMB NFC/NFD 대응 glob"""
    nfc_pat = _nfc(pattern)
    try:
        hits = [f for f in folder.iterdir()
                if f.is_file() and fnmatch.fnmatch(_nfc(f.name), nfc_pat)]
        return sorted(hits)
    except Exception:
        return []

def find_pdf(name):
    nfc = _nfc(name)
    for folder in CUSTOMER_DIR.iterdir():
        if not folder.is_dir():
            continue
        if _nfc(folder.name).startswith(f"{nfc}_") or _nfc(folder.name) == nfc:
            pdfs = _glob_nfc(folder, "종소세안내문_*.pdf")
            if pdfs:
                return max(pdfs, key=lambda p: p.stat().st_mtime)
    return None

def main():
    print(f"=== 수입금액 재파싱 시작: {datetime.now().strftime('%Y-%m-%d %H:%M')} ===")

    # ── 1) 파싱결과.xlsx 로드 ─────────────────────────────
    wb = openpyxl.load_workbook(PARSE_RESULT_XLSX)
    ws_xlsx = wb.active
    parse_headers = [c.value for c in ws_xlsx[1]]
    income_col_xlsx = parse_headers.index("수입금액총계") + 1

    name_to_xlsx_row = {}
    for row_idx in range(2, ws_xlsx.max_row + 1):
        n = ws_xlsx.cell(row=row_idx, column=1).value
        if n:
            name_to_xlsx_row[str(n).strip()] = row_idx

    # ── 2) 접수명단 gsheet 로드 ───────────────────────────
    creds = get_credentials()
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SPREADSHEET_ID)
    ws_gs = sh.worksheet("접수명단")
    all_vals = ws_gs.get_all_values()
    headers_gs = all_vals[0]
    name_col_gs  = headers_gs.index("성명") + 1
    income_col_gs = headers_gs.index("수입") + 1 if "수입" in headers_gs else None

    name_to_gs_row = {}
    current_income = {}
    for i, row in enumerate(all_vals[1:], start=2):
        if len(row) >= name_col_gs:
            n = row[name_col_gs - 1].strip()
            if n:
                name_to_gs_row[n] = i
                if income_col_gs and len(row) >= income_col_gs:
                    try:
                        current_income[n] = int(str(row[income_col_gs - 1]).replace(",", "") or 0)
                    except Exception:
                        current_income[n] = 0

    # ── 3) 재파싱 대상: 강제목록 + 현재 수입 1억↑ 또는 1천만↑(10x 오류 가능) ──
    targets = set(FORCE_NAMES)
    for name, inc in current_income.items():
        if inc >= 10_000_000:   # 1천만 이상 → 10배 오류 가능성 있는 구간 포함
            targets.add(name)
    # 수입 0이지만 PDF 있는 고객도 (파싱 실패 케이스)
    for name in name_to_xlsx_row:
        if current_income.get(name, 0) == 0:
            if find_pdf(name):
                targets.add(name)

    print(f"재파싱 대상: {len(targets)}명")

    changed = []
    xlsx_updates = []
    gs_updates   = []

    for name in sorted(targets):
        pdf = find_pdf(name)
        if not pdf:
            print(f"  [{name}] 안내문 PDF 없음 - 스킵")
            continue

        try:
            data = parse_anneam(pdf)
        except Exception as e:
            print(f"  [{name}] 파싱 오류: {e}")
            continue

        new_income = data.get("수입금액총계")
        if not new_income:
            print(f"  [{name}] 수입금액 파싱 실패 - 스킵")
            continue

        old_income = current_income.get(name, 0)
        diff = new_income - old_income

        if abs(diff) < 1000 and name not in FORCE_NAMES:
            continue   # 오차 1천원 미만이면 변경 없음으로 간주

        print(f"  [{name}] {old_income:,} → {new_income:,}  (차이 {diff:+,})")
        changed.append(name)

        # ── xlsx 업데이트 ───────────────────────────────
        row_idx = name_to_xlsx_row.get(name)
        if row_idx:
            ws_xlsx.cell(row=row_idx, column=income_col_xlsx).value = new_income
            xlsx_updates.append(name)

        # ── gsheet 수수료 재계산 ────────────────────────
        ledger    = data.get("기장의무", "")
        num_other = count_other_income(data)
        if new_income > 0 and ledger:
            try:
                fee     = calculate_fee(new_income, ledger, num_other, is_advance_booking=False)
                fee_adv = calculate_fee(new_income, ledger, num_other, is_advance_booking=True)
            except Exception:
                fee = fee_adv = None
        else:
            fee = fee_adv = None

        gs_row = name_to_gs_row.get(name)
        if gs_row and income_col_gs:
            col_letter = rowcol_to_a1(1, income_col_gs).rstrip("0123456789")
            gs_updates.append({"range": f"{col_letter}{gs_row}", "values": [[new_income]]})

            # 수수료 컬럼도 업데이트
            for col_name, val in [
                ("할인가", fee_adv["final_fee"]   if fee_adv else ""),
                ("수수료", fee["final_fee"]         if fee     else ""),
            ]:
                if col_name in headers_gs:
                    ci = headers_gs.index(col_name) + 1
                    cl = rowcol_to_a1(1, ci).rstrip("0123456789")
                    gs_updates.append({"range": f"{cl}{gs_row}", "values": [[val]]})

    # ── 4) 저장 ──────────────────────────────────────────
    if xlsx_updates:
        wb.save(PARSE_RESULT_XLSX)
        print(f"\n파싱결과.xlsx 저장 완료 ({len(xlsx_updates)}명)")

    if gs_updates:
        for i in range(0, len(gs_updates), 50):
            ws_gs.batch_update(gs_updates[i:i+50])
        print(f"접수명단 gsheet 업데이트 완료 ({len(gs_updates)}셀)")

    print(f"\n=== 완료: 변경 {len(changed)}명 ===")
    if changed:
        print("변경된 고객:", ", ".join(changed))

    # ── 5) 변경된 고객 작업판 재생성 ─────────────────────
    if changed:
        print(f"\n작업판 재생성 ({len(changed)}명)...")
        for name in changed:
            # jumin6: 폴더명에서 추출
            jumin6 = ""
            nfc = _nfc(name)
            for folder in CUSTOMER_DIR.iterdir():
                if folder.is_dir() and _nfc(folder.name).startswith(f"{nfc}_"):
                    parts = _nfc(folder.name).split("_")
                    jumin6 = parts[1] if len(parts) > 1 else ""
                    break
            try:
                result = make_jakupan(name, jumin6)
                if result:
                    print(f"  [{name}] 작업판 재생성 완료")
                else:
                    print(f"  [{name}] 작업판 생성 실패")
            except Exception as e:
                print(f"  [{name}] 작업판 오류: {e}")


if __name__ == "__main__":
    main()
