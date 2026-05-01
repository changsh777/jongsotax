"""
4명 모두 작업판 자동 채우기 - PoC

매핑 정책 (보수적):
- 확실한 셀만 채움
- 모르는 노란셀은 ⚠️ 확인필요 코멘트
- archive 패턴 저장

시나리오 매칭:
  오상연 (간편/프리만)        → 간편장부 / "프리"
  배성희 (간편/프리만)        → 간편장부 / "프리"
  박수경 (간편/사업자1+근로)   → 간편장부 / "사업자+프리" (프리 row 비움)
  서인미 (복식/프리만)        → 복식부기 / "프리복식"
"""
import openpyxl
from openpyxl.comments import Comment
from pathlib import Path
import warnings
warnings.filterwarnings("ignore")

import sys
sys.path.insert(0, r"F:\종소세2026")
from safe_save import safe_save_workbook

TEMPLATE_GANPYUN = Path(r"C:\Users\pc\OneDrive\문서\카카오톡 받은 파일\빈양식_간편장부.xlsx")
TEMPLATE_BOKSIK = Path(r"C:\Users\pc\OneDrive\문서\카카오톡 받은 파일\빈양식_복식부기.xlsx")
PDF_BASE = Path(r"F:\종소세2026\output\PDF")


CUSTOMERS = [
    {
        "성명": "오상연",
        "기장의무": "간편장부대상자",
        "사업장수": 0,
        "사업자번호": "",
        "업종코드": "940909",
        "수입금액": 52663077,
        "근로_단일": "X", "근로_복수": "X",
        "이자": "X", "배당": "X", "연금": "X", "기타": "X",
        "전년_총수입금액": 45383211,
        "전년_결정세액": 489416,
        "template": TEMPLATE_GANPYUN,
        "sheet": "프리",
    },
    {
        "성명": "배성희",
        "기장의무": "간편장부대상자",
        "사업장수": 0,
        "사업자번호": "",
        "업종코드": "940909",
        "수입금액": 43815000,
        "근로_단일": "X", "근로_복수": "X",
        "이자": "X", "배당": "X", "연금": "X", "기타": "X",
        "전년_총수입금액": "",
        "전년_결정세액": "",
        "template": TEMPLATE_GANPYUN,
        "sheet": "프리",
    },
    {
        "성명": "박수경",
        "기장의무": "간편장부대상자",
        "사업장수": 1,
        "사업자번호": "105-09-30253",
        "상호": "제이피디벨로프먼트",
        "업종코드": "742103",
        "수입금액": 4500000,
        "근로_단일": "O", "근로_복수": "X",
        "이자": "X", "배당": "X", "연금": "X", "기타": "X",
        "전년_총수입금액": 151450910,
        "전년_결정세액": 5443694,
        "부가세_매출": 4500000,
        "부가세_매입": 2909091,
        "부가세_납부": 159091,
        "template": TEMPLATE_GANPYUN,
        "sheet": "사업자+프리",
    },
    {
        "성명": "서인미",
        "기장의무": "복식부기의무자",
        "사업장수": 0,
        "사업자번호": "",
        "업종코드": "940906",
        "수입금액": 100517792,
        "근로_단일": "X", "근로_복수": "X",
        "이자": "X", "배당": "X", "연금": "X", "기타": "X",
        "전년_총수입금액": "",
        "전년_결정세액": "",
        "template": TEMPLATE_BOKSIK,
        "sheet": "프리복식",
    },
]


def is_yellow(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.rgb is None:
        return False
    rgb = str(fg.rgb).upper()
    if len(rgb) < 6: return False
    rgb6 = rgb[-6:]
    try:
        r, g, b = int(rgb6[0:2],16), int(rgb6[2:4],16), int(rgb6[4:6],16)
    except: return False
    return r >= 200 and g >= 200 and b < 120


def fill_pri_sheet(ws, c):
    """프리 시트: 사업소득(프리) + 근로/연금/이자배당/기타"""
    mapping = {
        "B2": ("성명", c["성명"]),
        "E11": ("프리 수입금액", c["수입금액"]),
        "F11": ("프리 업종코드", c["업종코드"]),
    }
    return mapping


def fill_saup_pri_sheet(ws, c):
    """사업자+프리: row 10=사업자, row 11=프리"""
    mapping = {
        "B2": ("성명", c["성명"]),
        # row 10: 사업자
        "E10": ("사업자 수입금액", c["수입금액"] if c["사업장수"] >= 1 else ""),
        "F10": ("사업자 업종코드", c["업종코드"] if c["사업장수"] >= 1 else ""),
        # row 11: 프리 (박수경은 프리 없음 - 비움)
        "E11": ("프리 수입금액", ""),
        "F11": ("프리 업종코드", ""),
    }
    return mapping


def fill_pri_boksik_sheet(ws, c):
    """프리복식: 복식부기 프리 (96개 중 핵심 일부만)"""
    # 박주환 원본 보면 D3=성명, D4=업종코드, D5=수입금액
    mapping = {
        "D3": ("성명", c["성명"]),
        "D4": ("업종코드", c["업종코드"]),
        "D5": ("수입금액", c["수입금액"]),
    }
    return mapping


SHEET_FILLERS = {
    "프리": fill_pri_sheet,
    "사업자+프리": fill_saup_pri_sheet,
    "프리복식": fill_pri_boksik_sheet,
}


def process_customer(c):
    print(f"\n========= {c['성명']} ({c['기장의무']}, 시트={c['sheet']}) =========")
    folder = PDF_BASE / c["성명"]
    folder.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(c["template"])
    if c["sheet"] not in wb.sheetnames:
        print(f"  [에러] 시트 '{c['sheet']}' 없음")
        return

    # 다른 시나리오 시트 삭제 (참조·업종코드는 유지)
    keep = {c["sheet"], "참조", "업종코드", "Sheet2", "2022귀속업종코드", "수지라"}
    for sn in list(wb.sheetnames):
        if sn not in keep:
            del wb[sn]

    ws = wb[c["sheet"]]

    # 노란셀 식별 + 매핑 적용
    filler = SHEET_FILLERS.get(c["sheet"])
    if filler is None:
        print(f"  [에러] 시트 '{c['sheet']}' filler 없음")
        return
    mapping = filler(ws, c)

    yellow_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if is_yellow(cell):
                yellow_cells.append(cell)

    filled, suspicious = [], []
    for cell in yellow_cells:
        coord = cell.coordinate
        if coord in mapping:
            label, val = mapping[coord]
            cell.value = val
            filled.append(f"{coord}={val} ({label})")
        else:
            cell.comment = Comment("⚠️ 자동 매핑 미정 - 직원 확인 필요", "Bot")
            suspicious.append(coord)

    print(f"  채움 {len(filled)}개:")
    for f in filled:
        print(f"    {f}")
    print(f"  미정 {len(suspicious)}개 (코멘트 부착)")
    if len(suspicious) <= 12:
        for s in suspicious:
            print(f"    {s}")
    else:
        print(f"    {suspicious[:8]}...")

    out_name = f"작업판_{c['성명']}_2024.xlsx"
    status, target = safe_save_workbook(wb, folder, out_name)
    print(f"  저장 [{status}]: {target}")


def main():
    for c in CUSTOMERS:
        process_customer(c)


if __name__ == "__main__":
    main()
