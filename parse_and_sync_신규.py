"""
신규 고객 PDF만 파싱해서 파싱결과.xlsx 업데이트 + 접수명단 동기화
전체 재파싱 없이 특정 고객만 처리
"""
import sys, os
os.environ.setdefault("SEOTAX_ENV", "nas")
sys.path.insert(0, r"F:\종소세2026")
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

from pathlib import Path
from datetime import datetime
import openpyxl
from gspread.utils import rowcol_to_a1
from config import PARSE_RESULT_XLSX, CUSTOMER_DIR
from parse_to_xlsx import (
    parse_anneam, parse_prev_income_xlsx, parse_vat_xlsx, COLUMNS
)
from fee_calculator import calculate_fee, count_other_income
from gsheet_writer import get_credentials
import gspread

# 처리할 신규 고객 목록
NEW_NAMES = ["탁설환"]  # 유영주는 홈택스 자료 없음

SPREADSHEET_ID = "1oh31k00Oa2lZWvu5fnBRVmurdlll1YEG8Fefi5FRfBI"

COL_MAP = {
    "수입":           "수입금액총계",
    "장부유형":       "기장의무",
    "추계시적용경비율": "추계시적용경비율",
    "할인가":         "사전접수할인가",
    "수수료":         "일반접수가",
    "이자":           "이자",
    "배당":           "배당",
    "근로(단일)":     "근로(단일)",
    "근로(복수)":     "근로(복수)",
    "연금":           "연금",
    "기타":           "기타",
}


def main():
    # 1) 신규 고객 PDF 파싱
    new_data = {}
    for name in NEW_NAMES:
        folder_candidates = list(CUSTOMER_DIR.glob(f"{name}_*")) + [CUSTOMER_DIR / name]
        folder = next((f for f in folder_candidates if f.is_dir()), None)
        if not folder:
            print(f"[{name}] 폴더 없음 - 스킵")
            continue

        pdf_candidates = list(folder.glob("종소세안내문_*.pdf"))
        if not pdf_candidates:
            print(f"[{name}] PDF 없음 - 스킵")
            continue
        pdf_path = max(pdf_candidates, key=lambda p: p.stat().st_mtime)

        print(f"[{name}] 파싱 중: {pdf_path.name}")
        try:
            data = parse_anneam(pdf_path)
        except Exception as e:
            print(f"[{name}] 파싱 실패: {e}")
            continue

        data.update(parse_prev_income_xlsx(folder))
        data.update(parse_vat_xlsx(folder))

        income = int(data.get("수입금액총계") or 0)
        ledger = data.get("기장의무", "")
        num_other = count_other_income(data)

        if income > 0 and ledger:
            try:
                fee     = calculate_fee(income, ledger, num_other, is_advance_booking=False)
                fee_adv = calculate_fee(income, ledger, num_other, is_advance_booking=True)
                data["사업장부 정가"] = fee["base_price"]
                data["타소득가산"]   = fee["other_income_fee"]
                data["합산정가"]     = fee["total_full_price"]
                data["사전접수할인가"] = fee_adv["final_fee"]
                data["일반접수가"]   = fee["final_fee"]
            except Exception as e:
                print(f"  [수수료 계산 실패] {name}: {e}")

        data["처리상태"] = "완료"
        data["처리일시"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        new_data[name] = data

        inc = data.get("수입금액총계", "")
        inc_str = f"{inc:,}" if isinstance(inc, int) else str(inc)
        print(f"  수입={inc_str} / {ledger} / {data.get('추계시적용경비율')} / 타소득{num_other}개")

    if not new_data:
        print("처리할 데이터 없음")
        return

    # 2) 파싱결과.xlsx 업데이트 (해당 고객 행만 추가/수정)
    wb = openpyxl.load_workbook(PARSE_RESULT_XLSX)
    ws_xlsx = wb.active
    parse_headers = [c.value for c in ws_xlsx[1]]

    # 기존 행 성명→row번호 인덱스
    name_to_xlsx_row = {}
    for row_idx in range(2, ws_xlsx.max_row + 1):
        n = ws_xlsx.cell(row=row_idx, column=1).value
        if n:
            name_to_xlsx_row[str(n).strip()] = row_idx

    for name, data in new_data.items():
        row_values = [data.get(h, "") for h in parse_headers]
        if name in name_to_xlsx_row:
            row_idx = name_to_xlsx_row[name]
            for col_idx, val in enumerate(row_values, start=1):
                ws_xlsx.cell(row=row_idx, column=col_idx).value = val
            print(f"[파싱결과.xlsx] {name} 행 업데이트 (행 {row_idx})")
        else:
            ws_xlsx.append(row_values)
            print(f"[파싱결과.xlsx] {name} 신규 행 추가")

    wb.save(PARSE_RESULT_XLSX)
    print(f"[파싱결과.xlsx] 저장 완료")

    # 3) 접수명단 Google Sheet 업데이트 (해당 고객 행만)
    creds = get_credentials()
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SPREADSHEET_ID)
    ws_gs = sh.worksheet("접수명단")

    all_vals = ws_gs.get_all_values()
    headers_gs = all_vals[0]

    # 컬럼 인덱스
    col_idx = {}
    for sheet_col in COL_MAP.keys():
        if sheet_col in headers_gs:
            col_idx[sheet_col] = headers_gs.index(sheet_col) + 1

    # 성명→행번호
    name_col_gs = headers_gs.index("성명") + 1
    name_to_gs_row = {}
    for i, row in enumerate(all_vals[1:], start=2):
        if len(row) >= name_col_gs:
            n = row[name_col_gs - 1].strip()
            if n:
                name_to_gs_row[n] = i

    updates = []
    for name, data in new_data.items():
        row_idx = name_to_gs_row.get(name)
        if not row_idx:
            print(f"  [접수명단] {name} 매칭 안됨 - 스킵")
            continue
        for sheet_col, parsed_key in COL_MAP.items():
            cidx = col_idx.get(sheet_col)
            if not cidx:
                continue
            val = data.get(parsed_key, "")
            col_letter = rowcol_to_a1(1, cidx).rstrip("0123456789")
            updates.append({"range": f"{col_letter}{row_idx}", "values": [[val]]})

    if updates:
        ws_gs.batch_update(updates)
        print(f"[접수명단] {list(new_data.keys())} 업데이트 완료 ({len(updates)}셀)")
    else:
        print("[접수명단] 업데이트 없음")

    # 4) 에어테이블 업데이트 (빈값 보호 로직 포함 — airtable_writer 참조)
    try:
        from airtable_writer import update_parsed_result
        for name, data in new_data.items():
            update_parsed_result(name=name, parsed=data)
    except Exception as e:
        print(f"[에어테이블 업데이트 오류] {e}")


if __name__ == "__main__":
    main()
