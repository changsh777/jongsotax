import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
os.environ.setdefault("SEOTAX_ENV", "nas")   # NAS 경로 우선
"""
PDF 안내문 → 엑셀 파싱 결과
- 고객/{성명_주민앞6}/종소세안내문_*.pdf 모두 스캔
- 11개 필드 + 메타정보 추출
- output/파싱결과.xlsx + 구글시트 동기화
"""
sys.path.insert(0, r"F:\종소세2026")
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import re
from pathlib import Path
from datetime import datetime
import warnings
import logging
from config import CUSTOMER_DIR, OUTPUT_DIR, RESULT_XLSX, PARSE_RESULT_XLSX

warnings.filterwarnings("ignore")
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)

PDF_BASE = CUSTOMER_DIR             # Z:\종소세2026\고객\ (SEOTAX_ENV=nas)
OUT_XLSX = PARSE_RESULT_XLSX        # Z:\종소세2026\output\파싱결과.xlsx

PREV_INC_COLS = [  # 전년도 종소세 (11개)
    "전년도_총수입금액", "전년도_필요경비", "전년도_종합소득금액", "전년도_소득공제",
    "전년도_과세표준", "전년도_산출세액", "전년도_세액감면공제", "전년도_결정세액",
    "전년도_가산세", "전년도_기납부세액", "전년도_납부할총세액",
]

VAT_COLS = [  # 부가세 (요약 - 2024년 2기 확정값 기준)
    "부가세_매출", "부가세_매입", "부가세_납부",
]

FEE_COLS = [  # 수수료 (5개)
    "사업장부 정가", "타소득가산", "합산정가",
    "사전접수할인가", "일반접수가",
]

COLUMNS = [
    "성명", "생년월일", "기장의무", "추계시적용경비율", "수입금액총계",
    "이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타",
] + FEE_COLS + PREV_INC_COLS + VAT_COLS + [
    "처리상태", "처리일시", "PDF경로",
]

INPUT_XLSX = RESULT_XLSX  # step4 결과.xlsx에서 처리상태 가져옴 (구글시트 대체)


def first_match(pattern, text, default=""):
    m = re.search(pattern, text)
    return m.group(1).strip() if m else default


def parse_prev_income_xlsx(folder):
    """전년도 종소세 xlsx → 11개 필드 dict"""
    f = folder / "전년도종소세신고내역.xls"
    if not f.exists():
        f = folder / "전년도종소세신고내역.xlsx"  # 이전 버전 호환
    if not f.exists():
        return {}
    try:
        import xlrd
        wb = xlrd.open_workbook(str(f))
        sheet = wb.sheet_by_index(0)
        if sheet.nrows < 2:
            return {}
        values = sheet.row_values(1)  # 데이터 행
        result = {}
        for i, col in enumerate(PREV_INC_COLS):
            if i < len(values):
                v = str(values[i]).replace(",", "").strip()
                try:
                    result[col] = int(v) if v and v.lstrip("-").isdigit() else v
                except Exception:
                    result[col] = v
        return result
    except Exception:
        return {}


def parse_vat_xlsx(folder):
    """부가세 xlsx → 매출/매입/납부 dict (여러 사업자번호 합산)
    각 셀 끝에 마지막 컬럼이 보통 2024년 2기 확정 = 가장 최신
    값 형식: '4,500,000 (0)' → 첫 숫자 시퀀스만 추출
    """
    import xlrd
    import openpyxl
    import re

    def extract_first_number(s):
        if not s:
            return 0
        m = re.search(r"-?[\d,]+", str(s))
        if not m:
            return 0
        try:
            return int(m.group(0).replace(",", ""))
        except Exception:
            return 0

    files = list(folder.glob("부가세신고내역_*.xlsx"))
    if not files:
        return {}

    매출_total = 0
    매입_total = 0
    납부_total = 0
    for f in files:
        try:
            rows = []
            try:
                wb = openpyxl.load_workbook(f, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            except Exception:
                wb = xlrd.open_workbook(str(f))
                sheet = wb.sheet_by_index(0)
                rows = [sheet.row_values(i) for i in range(sheet.nrows)]

            for row in rows:
                if not row:
                    continue
                cells = [str(c).strip() if c is not None else "" for c in row]
                first = cells[0]
                if "매출액" in first:
                    매출_total += extract_first_number(cells[-1] or cells[-2])
                elif "매입액" in first:
                    매입_total += extract_first_number(cells[-1] or cells[-2])
                elif "납부" in first or "환급" in first:
                    납부_total += extract_first_number(cells[-1] or cells[-2])
        except Exception:
            continue

    return {
        "부가세_매출": 매출_total,
        "부가세_매입": 매입_total,
        "부가세_납부": 납부_total,
    }


def parse_anneam(pdf_path):
    out = {c: "" for c in COLUMNS}
    with pdfplumber.open(pdf_path) as pdf:
        page1_text = pdf.pages[0].extract_text() or ""
        all_text = "\n".join((p.extract_text() or "") for p in pdf.pages)

    # 성명: 폴더명에서 추출 (폴더명 = 성명_주민앞6자리 or 성명)
    folder_name = pdf_path.parent.name
    if folder_name == "자료":  # /고객/홍길동_800101/자료/...
        folder_name = pdf_path.parent.parent.name
    out["성명"] = folder_name.split("_")[0]  # "홍길동_800101" → "홍길동"
    out["생년월일"] = first_match(r"생년월일\s+([\d.]+)", all_text)
    out["기장의무"] = first_match(r"기장의무\s+([^\n]+?)\s+추계시", all_text)
    out["추계시적용경비율"] = first_match(r"추계시적용경비율\s+(\S+)", all_text)

    # 수입금액 총계: 1페이지의 "총계" 라인 + 이후 줄까지 숫자 이어붙이기
    # 큰 숫자(1억↑)는 PDF 셀 너비 초과로 여러 줄로 쪼개짐 → 최대 4줄 룩어헤드
    lines = page1_text.split("\n")
    income_total = ""
    for i, line in enumerate(lines):
        if re.match(r"^\s*총\s?계", line):
            combined = re.sub(r"^\s*총\s?계", "", line)
            for j in range(1, 5):                        # 최대 4줄 더 확인
                if i + j >= len(lines):
                    break
                nxt = lines[i + j].strip()
                if re.match(r"^[\d,]+$", nxt):           # 숫자/콤마만 → 이어붙이기
                    combined += nxt
                else:
                    break
            m = re.search(r"[\d,]+", combined)
            if m:
                income_total = re.sub(r"[^\d]", "", m.group())
            break
    out["수입금액총계"] = int(income_total) if income_total else ""

    m = re.search(
        r"해당여부\s+([XO])\s+([XO])\s+([XO])\s+([XO])\s+([XO])\s+([XO])",
        all_text,
    )
    if m:
        labels = ["이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타"]
        for i, label in enumerate(labels):
            out[label] = m.group(i + 1)

    out["처리일시"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out["PDF경로"] = str(pdf_path)
    return out


def get_input_order():
    """고객 순서 + 처리상태 반환.

    순서: 구글시트 '접수명단' → 없으면 결과.xlsx 순서
    처리상태: 결과.xlsx(config 경로)에서 가져옴
    """
    # 1) 접수명단 순서 (구글시트)
    order = []
    try:
        from gsheet_writer import read_customers_from_gsheet
        customers = read_customers_from_gsheet()
        order = [c["name"] for c in customers]
    except Exception as e:
        print(f"  [접수명단] 구글시트 로드 실패, 결과.xlsx 순서 사용: {e}")

    # 2) 처리상태는 결과.xlsx에서 (config 경로 사용)
    status_map = {}
    if RESULT_XLSX.exists():
        wb2 = openpyxl.load_workbook(RESULT_XLSX, data_only=True)
        ws2 = wb2.active
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                name = str(row[0]).strip()
                status_map[name] = str(row[2] or "").strip()  # 처리상태 컬럼
                if not order:  # 구글시트 실패 시 결과.xlsx 순서 사용
                    order.append(name)

    return order, status_map


def collect_pdfs():
    """고객 폴더(성명_주민앞6자리)에서 가장 최근 종소세안내문 PDF 1개씩"""
    if not PDF_BASE.exists():
        return []
    pdfs = []
    for folder in sorted(PDF_BASE.iterdir()):
        if not folder.is_dir():
            continue
        # 자료 서브폴더도 스캔 (customer_folder는 하위에 /자료/ 생성)
        candidates = list(folder.glob("종소세안내문_*.pdf"))
        candidates += list(folder.glob("자료/종소세안내문_*.pdf"))
        if not candidates:
            continue
        latest = max(candidates, key=lambda p: p.stat().st_mtime)
        pdfs.append(latest)
    return pdfs


def write_xlsx(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "안내문파싱"

    # 헤더
    ws.append(COLUMNS)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터
    for row in rows:
        ws.append([row[c] for c in COLUMNS])

    # 수입금액 천단위 콤마
    income_col = COLUMNS.index("수입금액총계") + 1
    for row_idx in range(2, len(rows) + 2):
        cell = ws.cell(row=row_idx, column=income_col)
        cell.number_format = "#,##0"

    # 컬럼 너비 자동 조정 (간단 버전)
    widths = {
        "성명": 10, "생년월일": 12, "기장의무": 18, "추계시적용경비율": 18,
        "수입금액총계": 15, "처리일시": 20, "PDF경로": 50,
    }
    for i, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths.get(col, 10)

    # O/X 가운데 정렬
    ox_cols = ["이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타"]
    for col_name in ox_cols:
        col_idx = COLUMNS.index(col_name) + 1
        for row_idx in range(2, len(rows) + 2):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")

    wb.save(OUT_XLSX)


def _notify_errors(no_pdf: list, empty_income: list):
    """파싱 에러 텔레그램 알림 (종소택스봇)"""
    import os, json, urllib.request
    token   = os.environ.get("BOT_TOKEN", "")
    chat_id = os.environ.get("ADMIN_CHAT_ID", "")
    if not token or not chat_id:
        print("[텔레그램] 환경변수 미설정 — 알림 생략")
        return

    lines = ["[종소세 파싱 결과 알림]"]
    if no_pdf:
        lines.append(f"\n❌ PDF 없음 ({len(no_pdf)}명):")
        for name in no_pdf:
            lines.append(f"  • {name}")
    if empty_income:
        lines.append(f"\n⚠️ 수입금액 비어있음 ({len(empty_income)}명):")
        for name in empty_income:
            lines.append(f"  • {name}")
    if not no_pdf and not empty_income:
        lines.append("✅ 모든 고객 정상 파싱 완료")

    text = "\n".join(lines)
    url  = f"https://api.telegram.org/bot{token}/sendMessage"
    body = json.dumps({"chat_id": chat_id, "text": text}).encode("utf-8")
    req  = urllib.request.Request(
        url, data=body,
        headers={"Content-Type": "application/json"}, method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as r:
            r.read()
        print(f"[텔레그램] 알림 발송 완료 (에러 {len(no_pdf)}건, 수입비어있음 {len(empty_income)}건)")
    except Exception as e:
        print(f"[텔레그램] 알림 실패 (무시): {e}")


def main(sync_gsheet=True):
    # 입력 명단 순서 + 처리상태 (Track A 에러 케이스 포함)
    order, status_map = get_input_order()
    print(f"[입력 명단] {len(order)}명")

    # PDF 파싱한 결과를 성명 기준 dict로
    pdfs = collect_pdfs()
    parsed_by_name = {}
    for pdf in pdfs:
        try:
            data = parse_anneam(pdf)
            parsed_by_name[data["성명"]] = data
        except Exception as e:
            print(f"  ✗ {pdf.name} 파싱 실패: {e}")

    # 입력 순서대로 행 구성 (PDF 없으면 에러 케이스로 기록)
    from fee_calculator import calculate_fee, count_other_income

    rows = []
    for name in order:
        status = status_map.get(name, "")
        if name in parsed_by_name:
            data = parsed_by_name[name]
            data["처리상태"] = status or "완료"
            # 전년도 종소세 + 부가세 추가 파싱
            # 고객 폴더 탐색: 성명_주민앞6자리 패턴 우선, 없으면 성명만
            folder_candidates = list(PDF_BASE.glob(f"{name}_*")) + [PDF_BASE / name]
            folder = next((f for f in folder_candidates if f.is_dir()), PDF_BASE / name)
            data.update(parse_prev_income_xlsx(folder))
            data.update(parse_vat_xlsx(folder))
            # 수수료 계산
            try:
                income = int(data.get("수입금액총계") or 0)
                ledger = data.get("기장의무", "")
                num_other = count_other_income(data)
                if income > 0 and ledger:
                    fee = calculate_fee(income, ledger, num_other, is_advance_booking=False)
                    fee_adv = calculate_fee(income, ledger, num_other, is_advance_booking=True)
                    data["사업장부 정가"] = fee["base_price"]
                    data["타소득가산"] = fee["other_income_fee"]
                    data["합산정가"] = fee["total_full_price"]
                    data["사전접수할인가"] = fee_adv["final_fee"]
                    data["일반접수가"] = fee["final_fee"]
            except Exception as e:
                print(f"    [수수료 계산 실패] {name}: {e}")
            rows.append(data)
            _inc = data['수입금액총계']
            _inc_str = f"{_inc:,}" if isinstance(_inc, int) else str(_inc)
            print(f"  ✓ {name}  (수입 {_inc_str}) [{data['처리상태']}]")
        else:
            # PDF 없는 케이스 - 에러로 행 추가
            empty = {c: "" for c in COLUMNS}
            empty["성명"] = name
            empty["처리상태"] = status or "에러"
            empty["처리일시"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            rows.append(empty)
            print(f"  ✗ {name}  PDF 없음 [{empty['처리상태']}]")

    # 에러 분류
    no_pdf_names   = [r["성명"] for r in rows if not r.get("PDF경로")]
    empty_income   = [r["성명"] for r in rows if r.get("PDF경로") and not r.get("수입금액총계")]

    if rows:
        write_xlsx(rows)
        print(f"\n[로컬 엑셀] {OUT_XLSX} ({len(rows)}건)")

        if sync_gsheet:
            ok_rows = [r for r in rows if r.get("PDF경로")]
            # 1) 안내문파싱 시트 (전체 파싱 내역 보관)
            try:
                from gsheet_writer import write_all
                write_all(ok_rows)
                print(f"[구글시트] 안내문파싱 {len(ok_rows)}건 write_all 완료")
            except Exception as e:
                print(f"[구글시트] 안내문파싱 실패: {e}")
            # 2) 접수명단 시트에 수입·할인가·수수료 3개 컬럼만 upsert (에어테이블 순서 유지)
            try:
                from gsheet_writer import write_parsed_to_접수명단
                n = write_parsed_to_접수명단(ok_rows)
                print(f"[구글시트] 접수명단 {n}건 업데이트 (수입/할인가/수수료)")
            except Exception as e:
                print(f"[구글시트] 접수명단 업데이트 실패: {e}")

    # 텔레그램 알림
    _notify_errors(no_pdf_names, empty_income)


if __name__ == "__main__":
    main()
