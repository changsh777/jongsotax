import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
os.environ.setdefault("SEOTAX_ENV", "nas")
sys.path.insert(0, r"F:\종소세2026")

import openpyxl
from config import PARSE_RESULT_XLSX
from fee_calculator import calculate_fee, count_other_income

wb = openpyxl.load_workbook(PARSE_RESULT_XLSX, data_only=True)
ws = wb.active
headers = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

print(f"{'성명':<8} {'수입':>12} {'기장의무':<16} {'타소득':>4} {'시트_할인가':>10} {'시트_수수료':>10} {'계산_할인가':>10} {'계산_수수료':>10} 일치")
print("-" * 95)

total = 0
mismatch = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    income_raw = row[idx.get("수입금액총계", 0)]
    if not income_raw:
        continue
    name   = str(row[0]).strip()
    income = int(income_raw)
    ledger = str(row[idx["기장의무"]] or "").strip()

    # fee_calculator의 count_other_income 사용 (이자+배당=금융소득 1개)
    customer = {h: row[i] for i, h in enumerate(headers) if h}
    num_other = count_other_income(customer)

    sheet_adv  = int(row[idx["사전접수할인가"]] or 0)
    sheet_norm = int(row[idx["일반접수가"]]    or 0)

    try:
        fee_adv  = calculate_fee(income, ledger, num_other, is_advance_booking=True)
        fee_norm = calculate_fee(income, ledger, num_other, is_advance_booking=False)
        calc_adv  = fee_adv["final_fee"]
        calc_norm = fee_norm["final_fee"]
        ok = (calc_adv == sheet_adv and calc_norm == sheet_norm)
        total += 1
        if not ok:
            mismatch += 1
            print(f"{name:<8} {income:>12,} {ledger:<16} {num_other:>4} {sheet_adv:>10,} {sheet_norm:>10,} {calc_adv:>10,} {calc_norm:>10,} X")
    except Exception as e:
        print(f"{name:<8} 오류: {e}")

print(f"\n총 {total}건 / 불일치 {mismatch}건")
if mismatch == 0:
    print("전부 일치!")
