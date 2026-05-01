"""
step4_full.py
- 종소세 안내문 PDF + 전년도 종소세 엑셀 + 부가세 엑셀(사업자별)
- 이름별 폴더 구조: output/PDF/{성명}/...
- 사업자번호는 PDF에서 정규식으로 추출

사전 조건:
  1. python F:\종소세2026\launch_edge.py
  2. 엣지에서 홈택스 로그인
  3. 신고도움서비스 진입 가능 상태
"""
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from pathlib import Path
from datetime import datetime
import openpyxl
import pdfplumber
import time
import re
import warnings
import logging
import sys
sys.path.insert(0, r"F:\종소세2026")
from safe_save import safe_download, safe_popup_pdf, safe_save_workbook
from gsheet_writer import upsert_consent_row, read_customers_from_gsheet
from config import customer_folder, RESULT_XLSX, OUTPUT_DIR, CUSTOMER_DIR

warnings.filterwarnings("ignore")
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)

OUTPUT_XLSX = RESULT_XLSX          # config: NAS=Z:\종소세2026\output\결과.xlsx
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
CUSTOMER_DIR.mkdir(parents=True, exist_ok=True)

REPORT_HELP_URL = (
    "https://hometax.go.kr/websquare/websquare.html"
    "?w2xPath=/ui/pp/index_pp.xml"
    "&tmIdx=06&tm2lIdx=0601000000&tm3lIdx=0601200000"
)

BIZNO_PATTERN = re.compile(r"\d{3}-\d{2}-\d{5}")


# -------------------- 유틸 --------------------

def normalize_jumin(raw):
    s = str(raw).replace("-", "").replace(" ", "").strip()
    if len(s) != 13 or not s.isdigit():
        raise ValueError(f"주민번호 형식 이상: {raw}")
    return s[:6], s[6:]


def read_customers(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[2]
        jumin = row[4]
        phone = row[3]  # 자동전화번호 (10자리, leading 0 없을 수 있음)
        if not name or not jumin:
            continue
        out.append({
            "name": str(name).strip(),
            "jumin_raw": jumin,
            "phone_raw": str(phone).strip() if phone else "",
        })
    return out


def extract_biznos(pdf_path):
    biznos = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            biznos.extend(BIZNO_PATTERN.findall(text))
    seen, out = set(), []
    for b in biznos:
        if b not in seen:
            seen.add(b)
            out.append(b)
    return out


# -------------------- 페이지 조작 --------------------

def fill_jumin_and_search(page, front, back):
    """주민번호 입력 + 조회하기 클릭"""
    inputs = [el for el in page.locator(
        "xpath=//th[contains(.,'주민등록번호')]/following-sibling::td//input"
    ).all() if el.is_visible()]
    if len(inputs) < 2:
        raise RuntimeError("주민번호 입력 필드 2개 못 찾음")
    inputs[0].fill(front)
    inputs[1].fill(back)

    btns = [b for b in page.locator(
        "xpath=//th[contains(.,'주민등록번호')]/following-sibling::td//input[@value='조회하기']"
    ).all() if b.is_visible()]
    if not btns:
        raise RuntimeError("주민번호 행의 조회하기 버튼 못 찾음")
    btns[0].click()


def wait_preview_button(page, timeout_ms=10000):
    """미리보기 버튼이 보이면 그 locator 반환, 없으면 None"""
    try:
        btn = page.get_by_text("미리보기", exact=False).first
        btn.wait_for(timeout=timeout_ms, state="visible")
        return btn
    except Exception:
        return None


def save_anneam_pdf(ctx, page, preview_btn, save_path):
    """미리보기 클릭 → 팝업 → ClipReport4 PDF 저장 버튼 강제 활성화 → archive 패턴 저장"""
    with ctx.expect_page(timeout=15000) as popup_info:
        preview_btn.click()
    popup = popup_info.value
    popup.wait_for_load_state("networkidle", timeout=30000)
    time.sleep(3)

    with popup.expect_download(timeout=30000) as dl_info:
        clicked = popup.evaluate("""
            () => {
                const btn = document.querySelector('.report_menu_pdf_button');
                if (!btn) return false;
                btn.classList.remove('report_menu_pdf_button_svg_dis');
                btn.classList.add('report_menu_pdf_button_svg');
                btn.disabled = false;
                btn.click();
                return true;
            }
        """)
        if not clicked:
            popup.close()
            raise RuntimeError("PDF 저장 버튼 못 찾음")
    download = dl_info.value
    target = Path(save_path)
    status, _ = safe_download(download, target.parent, target.name)
    print(f"    [저장:{status}] {target}", flush=True)
    popup.close()


def wait_excel_button(page, timeout_ms=5000):
    """모달의 '엑셀 내려받기' 버튼이 보이는지 = 모달이 떴고 엑셀 다운로드 가능한 상태"""
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        try:
            visible = page.evaluate("""
                () => {
                    const all = document.querySelectorAll('button, input, a, span, div');
                    for (const el of all) {
                        const txt = (el.innerText || el.value || '').replace(/\\s+/g, '');
                        if (txt.includes('엑셀내려받기') && el.offsetParent !== null) return true;
                    }
                    return false;
                }
            """)
            if visible:
                return True
        except Exception:
            pass
        time.sleep(0.3)
    return False


def click_excel_download_in_modal(page, save_path, timeout_ms=8000):
    """엑셀 내려받기 클릭 → 다운로드 저장
    반환: True(성공) / False(자료없음·다운로드 미발생)
    """
    candidates = page.locator(
        "xpath=//button[contains(.,'엑셀 내려받기')]"
        " | //input[@type='button' and contains(@value,'엑셀 내려받기')]"
        " | //a[contains(.,'엑셀 내려받기')]"
    ).all()
    visible = [c for c in candidates if c.is_visible()]
    print(f"    [엑셀버튼] 후보 {len(candidates)}, 보이는 것 {len(visible)}", flush=True)

    if not visible:
        candidates = page.locator(
            "xpath=//*[normalize-space(text())='엑셀 내려받기']"
        ).all()
        visible = [c for c in candidates if c.is_visible()]
        print(f"    [엑셀버튼 fallback] 보이는 것 {len(visible)}", flush=True)

    if not visible:
        return False

    target = visible[0]
    try:
        with page.expect_download(timeout=timeout_ms) as dl_info:
            target.click()
        download = dl_info.value
        sp = Path(save_path)
        status, _ = safe_download(download, sp.parent, sp.name)
        print(f"    [엑셀저장:{status}] {sp}", flush=True)
        return True
    except PWTimeout:
        print(f"    [엑셀] 다운로드 미발생 - 자료없음으로 간주", flush=True)
        return False


def close_modal(page, verbose=True):
    """모달 닫기 - 항상 모든 전략 실행 (early return 제거)"""
    print(f"    [close_modal] 호출됨", flush=True)

    def is_popup_visible():
        """엑셀 내려받기 버튼이 보이면 모달 떠있음"""
        try:
            return page.evaluate("""
                () => {
                    const all = document.querySelectorAll('button, input, a, span, div');
                    for (const el of all) {
                        const txt = (el.innerText || el.value || '').replace(/\\s+/g, '');
                        if (txt.includes('엑셀내려받기') && el.offsetParent !== null) return true;
                    }
                    return false;
                }
            """)
        except Exception:
            return False

    # Strategy 1: 보이는 '닫기' 텍스트 중 마지막 것 클릭 (모달이 보통 가장 나중에 렌더됨)
    r1 = page.evaluate("""
        () => {
            const all = Array.from(document.querySelectorAll(
                'button, input[type=button], input[type=submit], a, span, div'
            ));
            const targets = all.filter(el => {
                const txt = (el.innerText || el.value || el.textContent || '').replace(/\\s+/g, '');
                return txt === '닫기' && el.offsetParent !== null;
            });
            if (targets.length === 0) return {n: 0};
            const t = targets[targets.length - 1];
            t.click();
            return {n: targets.length, tag: t.tagName, cls: String(t.className).slice(0, 50)};
        }
    """)
    print(f"    [close_modal s1] {r1}", flush=True)
    for _ in range(6):
        time.sleep(0.5)
        if not is_popup_visible():
            print(f"    [close_modal] s1 후 닫힘 확인", flush=True)
            return

    # Strategy 2: 클래스명에 close 포함된 요소 클릭 (X 버튼)
    r2 = page.evaluate("""
        () => {
            const all = Array.from(document.querySelectorAll('*'));
            const targets = all.filter(el => {
                const cls = (typeof el.className === 'string') ? el.className.toLowerCase() : '';
                return cls.includes('close') && el.offsetParent !== null;
            });
            if (targets.length === 0) return {n: 0};
            const t = targets[targets.length - 1];
            t.click();
            return {n: targets.length, tag: t.tagName, cls: String(t.className).slice(0, 50)};
        }
    """)
    print(f"    [close_modal s2] {r2}", flush=True)
    for _ in range(6):
        time.sleep(0.5)
        if not is_popup_visible():
            print(f"    [close_modal] s2 후 닫힘 확인", flush=True)
            return

    # Strategy 3: ESC
    page.keyboard.press("Escape")
    time.sleep(1)
    if not is_popup_visible():
        return

    # Strategy 4: 페이지 클릭 (외부 클릭으로 모달 닫는 패턴)
    page.mouse.click(10, 10)
    time.sleep(1)

    if is_popup_visible() and verbose:
        print(f"    [close_modal] !!! 모달 안 닫힘 - 후속 작업 실패 가능")


def download_prev_income_tax(page, save_path):
    """전년도 종합소득세 [조회하기] → 모달 → 엑셀 내려받기
    반환: True(다운로드 성공) / False(자료없음)
    """
    btns = [b for b in page.locator(
        "xpath=//*[normalize-space(text())='전년도 종합소득세']"
        "/ancestor::tr[1]//input[@value='조회하기']"
    ).all() if b.is_visible()]
    if not btns:
        btns = [b for b in page.locator(
            "xpath=//th[contains(.,'전년도 종합소득세')]/following-sibling::td//input[@value='조회하기']"
        ).all() if b.is_visible()]
    if not btns:
        raise RuntimeError("전년도 종소세 조회하기 버튼 못 찾음")

    btns[0].click()
    time.sleep(4)  # 모달 + 데이터 렌더링 여유

    # 데이터 행 등장 대기 (최대 10초)
    for _ in range(20):
        time.sleep(0.5)
        has_data = page.evaluate("""
            () => /총수입금액[\\s\\S]*?\\d/.test(document.body.innerText || '')
        """)
        if has_data:
            break

    if not wait_excel_button(page, timeout_ms=10000):
        return False

    time.sleep(2)
    ok = click_excel_download_in_modal(page, save_path)
    if ok:
        time.sleep(2)
    close_modal(page)
    return ok


def download_vat(page, bizno, save_path):
    """부가세: 사업자번호 입력 → 조회하기 → 모달 → 엑셀 내려받기"""
    print(f"    [부가세] 사업자번호 {bizno} 시도", flush=True)

    # 0. 이전 모달 잔존 검증 - 엑셀 내려받기 버튼이 보이면 이전 모달 안 닫힘
    leftover = page.evaluate("""
        () => {
            for (const el of document.querySelectorAll('button, input, a, span, div')) {
                const txt = (el.innerText || el.value || '').replace(/\\s+/g, '');
                if (txt.includes('엑셀내려받기') && el.offsetParent !== null) return true;
            }
            return false;
        }
    """)
    if leftover:
        print(f"    [부가세] !!! 이전 모달 잔존 - 강제 닫기", flush=True)
        close_modal(page)
        time.sleep(1)

    # 1. 사업자번호 입력
    bizno_inputs = [el for el in page.locator(
        "xpath=//input[@placeholder='사업자번호 입력' or contains(@title,'사업자번호')]"
    ).all() if el.is_visible()]
    print(f"    [부가세] 사업자번호 input 후보 {len(bizno_inputs)}개", flush=True)
    if not bizno_inputs:
        bizno_inputs = [el for el in page.locator(
            "xpath=//*[contains(text(),'당해년도 부가가치세')]"
            "/ancestor::tr[1]//input[@type='text']"
        ).all() if el.is_visible()]
    if not bizno_inputs:
        raise RuntimeError("부가세 사업자번호 입력란 못 찾음")
    bizno_input_el = bizno_inputs[0]
    bizno_input_el.fill("")
    bizno_input_el.fill(bizno)
    filled = bizno_input_el.input_value()
    print(f"    [부가세] 입력값 검증: '{filled}'", flush=True)

    # 2. 사업자번호 input 바로 다음 sibling 조회하기 클릭 (가장 확실)
    btn = bizno_input_el.locator("xpath=following::input[@value='조회하기'][1]").first
    print(f"    [부가세] sibling 조회하기 클릭 시도", flush=True)
    btn.click()
    time.sleep(4)  # 모달 뜨고 데이터 렌더링 여유

    # 3. 부가세 모달 감지 - 모든 frame 스캔 (최대 15초)
    #    홈택스는 모달 결과가 iframe 안에 렌더될 수 있으므로 page.frames() 전체 체크
    is_vat_modal = False
    has_data_row = False
    vat_frame = None

    _VAT_JS = """
        () => {
            const t = (document.body && document.body.innerText) || '';
            const vat = t.includes('매출액') || t.includes('매입액')
                     || t.includes('환급세액') || t.includes('과세기간');
            const dataRow = vat && /[\\d,]{3,}/.test(
                t.slice(Math.max(0, t.indexOf('매출액')))
            );
            return { vat, dataRow, preview: t.slice(0, 200) };
        }
    """
    for _i in range(30):  # 최대 15초
        time.sleep(0.5)
        for frame in page.frames:  # frames는 property (괄호 없음)
            try:
                state = frame.evaluate(_VAT_JS)
                if state.get("vat"):
                    if not is_vat_modal:  # 첫 감지 시에만 디버그 출력
                        print(f"    [부가세] 감지 frame: {frame.url[:80]}", flush=True)
                        print(f"    [부가세] preview: {state['preview'][:100]}", flush=True)
                    is_vat_modal = True
                    vat_frame = frame
                if state.get("dataRow"):
                    has_data_row = True
                    vat_frame = frame
                    break
            except Exception:
                pass
        if has_data_row:
            break
    print(f"    [부가세] 부가세 모달: {is_vat_modal}, 데이터: {has_data_row}", flush=True)
    if vat_frame:
        print(f"    [부가세] 감지 frame URL: {vat_frame.url[:100]}", flush=True)

    if not is_vat_modal:
        print(f"    [부가세] 부가세 모달 아님 - 다운로드 스킵", flush=True)
        close_modal(page)
        return False

    if not has_data_row:
        print(f"    [부가세] 데이터 행 미렌더링 - 안전을 위해 추가 5초 대기", flush=True)
        time.sleep(5)

    # 모달 DOM에서 표 직접 추출 - 감지된 frame 사용 (iframe 대응)
    time.sleep(2)
    _use_frame = vat_frame if vat_frame else page
    table_data = _use_frame.evaluate("""
        () => {
            const kws = ['매출액', '매입액', '과세기간', '부가가치세', '납부세액'];
            for (const table of document.querySelectorAll('table')) {
                if (table.offsetParent === null) continue;
                const txt = table.innerText || '';
                if (!kws.some(k => txt.includes(k))) continue;
                const rows = [];
                for (const tr of table.querySelectorAll('tr')) {
                    const cells = [];
                    for (const c of tr.querySelectorAll('td, th')) {
                        cells.push((c.innerText || '').trim().replace(/\\n/g, ' '));
                    }
                    if (cells.length) rows.push(cells);
                }
                return rows;
            }
            return null;
        }
    """)

    if not table_data:
        print(f"    [부가세] 모달에서 표 추출 실패", flush=True)
        close_modal(page)
        return False

    print(f"    [부가세] 모달 표 {len(table_data)}행 추출", flush=True)
    for r in table_data:
        print(f"      {r}")

    # 자체 xlsx → archive 패턴 저장
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "부가세신고내역"
    for row in table_data:
        ws.append(row)
    sp = Path(save_path)
    status, _ = safe_save_workbook(wb, sp.parent, sp.name)
    print(f"    [부가세 저장:{status}] {sp}", flush=True)

    close_modal(page)
    return True


# -------------------- 고객 1명 처리 --------------------

def process_one(ctx, page, customer):
    name = customer["name"]
    folder = customer_folder(name, customer.get("jumin_raw", ""))  # NAS 경로

    result = {
        "status": "에러",
        "error_msg": "",
        "anneam_pdf": "",
        "prev_income_xlsx": "",
        "vat_xlsx_count": 0,
        "biznos": "",
    }

    try:
        front, back = normalize_jumin(customer["jumin_raw"])
    except ValueError as e:
        result["error_msg"] = str(e)
        return result

    try:
        # 페이지 새로고침
        page.goto(REPORT_HELP_URL, wait_until="domcontentloaded")
        time.sleep(2)

        # 주민번호 + 조회
        fill_jumin_and_search(page, front, back)
        time.sleep(3)

        # 미리보기 확인 (없으면 에러)
        preview_btn = wait_preview_button(page, timeout_ms=10000)
        if preview_btn is None:
            result["error_msg"] = "미리보기 버튼 없음 (조회 결과 없음)"
            return result

        # 1) 안내문 PDF (이미 있으면 팝업만 닫고 스킵)
        anneam_path = folder / f"종소세안내문_{name}.pdf"
        if anneam_path.exists():
            print(f"    [PDF 스킵] 기존 파일 존재 - 팝업만 닫기", flush=True)
            # 팝업이 열렸을 수 있으니 닫기
            try:
                for p2 in list(ctx.pages):
                    if p2 != page:
                        p2.close()
            except Exception:
                pass
        else:
            try:
                save_anneam_pdf(ctx, page, preview_btn, anneam_path)
            except PermissionError:
                print(f"    [PDF] PermissionError - 파일 잠김, 팝업 정리 후 스킵", flush=True)
                try:
                    for p2 in list(ctx.pages):
                        if p2 != page:
                            p2.close()
                except Exception:
                    pass
        result["anneam_pdf"] = str(anneam_path)
        time.sleep(1)

        # 2) 전년도 종소세 엑셀
        try:
            prev_path = folder / "전년도종소세신고내역.xlsx"
            ok = download_prev_income_tax(page, prev_path)
            if ok:
                result["prev_income_xlsx"] = str(prev_path)
            else:
                result["prev_income_xlsx"] = "자료없음"
        except Exception as e:
            result["error_msg"] += f" [전년도종소세 실패: {type(e).__name__}: {str(e)[:100]}]"

        # 3) 사업자번호 추출
        biznos = extract_biznos(anneam_path)
        result["biznos"] = ",".join(biznos)

        # 4) 사업자별 부가세 엑셀
        vat_count = 0
        for bizno in biznos:
            try:
                vat_path = folder / f"부가세신고내역_{bizno}.xlsx"
                ok = download_vat(page, bizno, vat_path)
                if ok:
                    vat_count += 1
            except Exception as e:
                result["error_msg"] += f" [부가세 {bizno} 실패: {type(e).__name__}: {str(e)[:100]}]"
        result["vat_xlsx_count"] = vat_count

        result["status"] = "완료" if not result["error_msg"] else "부분완료"
        return result

    except Exception as e:
        result["error_msg"] = f"{type(e).__name__}: {str(e)[:200]}"
        return result


# -------------------- 결과 엑셀 --------------------

def ensure_output_workbook():
    if OUTPUT_XLSX.exists():
        wb = openpyxl.load_workbook(OUTPUT_XLSX)
        ws = wb.active
        return wb, ws
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "결과"
    ws.append([
        "성명", "주민번호", "처리상태",
        "안내문PDF", "전년도종소세엑셀", "사업자번호목록", "부가세엑셀건수",
        "에러메시지", "시도일시",
    ])
    wb.save(OUTPUT_XLSX)
    return wb, ws


# -------------------- 메인 --------------------

def main():
    customers = read_customers_from_gsheet()
    print(f"[시작] 총 {len(customers)}명 처리\n")

    wb, ws = ensure_output_workbook()

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://localhost:9222")
        ctx = browser.contexts[0]
        page = ctx.pages[0]
        page.bring_to_front()
        # alert/confirm 자동 닫기 (자료없음 alert 등 차단 방지)
        page.on("dialog", lambda d: d.dismiss())

        for i, c in enumerate(customers, 1):
            print(f"[{i}/{len(customers)}] {c['name']}")
            r = process_one(ctx, page, c)

            ws.append([
                c["name"], str(c["jumin_raw"]), r["status"],
                r["anneam_pdf"], r["prev_income_xlsx"],
                r["biznos"], r["vat_xlsx_count"],
                r["error_msg"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ])
            wb.save(OUTPUT_XLSX)

            if r["status"] == "완료":
                print(f"    완료 (사업자 {r['vat_xlsx_count']}개)")
            elif r["status"] == "부분완료":
                print(f"    부분완료: {r['error_msg']}")
            else:
                print(f"    에러: {r['error_msg']}")
                # 에러 = 미동의 → 미동의명단 시트에 복사
                try:
                    upsert_consent_row({
                        "성명": c["name"],
                        "주민번호": str(c["jumin_raw"]),
                        "핸드폰번호": c.get("phone_raw", ""),
                        "에러사유": r["error_msg"][:200],
                        "수임상태": "미동의",
                    })
                    print(f"    → 미동의명단 시트에 기록 완료")
                except Exception as ge:
                    print(f"    → 미동의명단 기록 실패 (무시): {ge}")
            print()

    print(f"[전체 완료] {OUTPUT_XLSX}")

    # ---- step4 완료 후 파싱 + 구글시트 자동 업데이트 ----
    print(f"\n[파싱 시작] parse_to_xlsx.main() 자동 호출")
    try:
        import parse_to_xlsx
        parse_to_xlsx.main(sync_gsheet=True)
    except Exception as pe:
        print(f"[파싱 실패] {pe}")
        print("  수동으로 실행: python F:\\종소세2026\\parse_to_xlsx.py")


if __name__ == "__main__":
    main()
