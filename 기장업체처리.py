"""
기장업체처리.py - 기장업체 종소세 안내문 전체 파이프라인

입력: 엑셀 파일 (성명, 주민번호 2컬럼)
처리: 세무사 계정 로그인 상태에서 주민번호 입력 → 조회
      안내문PDF + 전년도소득세 + 부가세 + 지급명세서 + 간이용역 다운로드
      PDF 파싱
출력: 입력 엑셀에 파싱 결과 컬럼 추가 저장
      NAS 고객 폴더에 파일 저장

전제조건:
  1. python launch_edge.py 실행 → Edge 디버그 창
  2. 홈택스에 세무사 계정 로그인 (3번)
  3. python 기장업체처리.py [엑셀경로] [시작번호]
     예) python 기장업체처리.py C:\\Users\\pc\\OneDrive\\문서\\기장업체파싱.xlsx
     예) python 기장업체처리.py ... 10   (10번째부터 재개)
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
os.environ.setdefault("SEOTAX_ENV", "nas")
sys.path.insert(0, r"F:\종소세2026")

from datetime import datetime
from pathlib import Path
import openpyxl
from playwright.sync_api import sync_playwright

from 종합소득세안내문조회 import process_one, ensure_output_workbook
from parse_to_xlsx import (
    parse_anneam, parse_prev_income_xlsx, parse_vat_xlsx,
    PREV_INC_COLS, VAT_COLS
)
from config import CUSTOMER_DIR, customer_folder

# ── 결과 컬럼 정의 ──────────────────────────────────────────────
RESULT_COLS = [
    "처리상태",          # C
    "수입금액총계",      # D
    "기장의무",          # E
    "추계시적용경비율",  # F
    "이자",              # G
    "배당",              # H
    "근로(단일)",        # I
    "근로(복수)",        # J
    "연금",              # K
    "기타",              # L
    "전년도_납부할총세액", # M
    "부가세_납부",       # N
    "사업자번호",        # O
    "지급명세서PDF",     # P
    "간이용역",          # Q
    "에러메시지",        # R
    "처리일시",          # S
]


def load_customers_from_xlsx(xlsx_path):
    """엑셀에서 성명+주민번호 읽기 (헤더 1행 제외)"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    customers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0] or "").strip()
        jumin = str(row[1] or "").strip()
        if not name:
            continue
        customers.append({"name": name, "jumin_raw": jumin, "phone_raw": ""})
    return customers


def write_result_row(ws, row_idx, result, parsed):
    """엑셀 행에 결과 기록"""
    vals = [
        result.get("status", "에러"),
        parsed.get("수입금액총계", ""),
        parsed.get("기장의무", ""),
        parsed.get("추계시적용경비율", ""),
        parsed.get("이자", ""),
        parsed.get("배당", ""),
        parsed.get("근로(단일)", ""),
        parsed.get("근로(복수)", ""),
        parsed.get("연금", ""),
        parsed.get("기타", ""),
        parsed.get("전년도_납부할총세액", ""),
        parsed.get("부가세_납부", ""),
        result.get("biznos", ""),
        "O" if result.get("anneam_pdf") and Path(result["anneam_pdf"]).parent.parent.glob("지급명세서/*.pdf") else "X",
        "O" if result.get("anneam_pdf") and Path(result["anneam_pdf"]).parent.parent.glob("간이용역소득/*.xlsx") else "X",
        result.get("error_msg", ""),
        datetime.now().strftime("%Y-%m-%d %H:%M"),
    ]
    for col_offset, val in enumerate(vals):
        ws.cell(row=row_idx, column=3 + col_offset, value=val)


def parse_customer(name, jumin_raw):
    """고객 폴더에서 파싱 결과 반환"""
    folder_candidates = list(CUSTOMER_DIR.glob(f"{name}_*")) + [CUSTOMER_DIR / name]
    folder = next((f for f in folder_candidates if f.is_dir()), None)
    if not folder:
        return {}

    parsed = {}

    # 안내문 PDF 파싱
    pdf_candidates = list(folder.glob("종소세안내문_*.pdf"))
    if pdf_candidates:
        pdf_path = max(pdf_candidates, key=lambda p: p.stat().st_mtime)
        print(f"    [파싱] {pdf_path.name}", flush=True)
        try:
            parsed.update(parse_anneam(pdf_path))
        except Exception as e:
            print(f"    [파싱실패] {e}", flush=True)

    # 전년도 소득세 파싱
    parsed.update(parse_prev_income_xlsx(folder))

    # 부가세 파싱
    parsed.update(parse_vat_xlsx(folder))

    return parsed


def ensure_header(ws):
    """엑셀 헤더에 결과 컬럼 추가 (없으면)"""
    existing = [ws.cell(row=1, column=c).value for c in range(3, 3 + len(RESULT_COLS))]
    if existing[0] != RESULT_COLS[0]:
        for i, col_name in enumerate(RESULT_COLS):
            ws.cell(row=1, column=3 + i, value=col_name)
        print("  [헤더] 결과 컬럼 추가", flush=True)


def main():
    # 인수 파싱
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\pc\OneDrive\문서\기장업체파싱.xlsx")
    start_idx = int(sys.argv[2]) - 1 if len(sys.argv) > 2 else 0  # 0-based

    if not xlsx_path.exists():
        print(f"[오류] 파일 없음: {xlsx_path}")
        sys.exit(1)

    print(f"[기장업체처리] {xlsx_path.name}")
    print(f"  Edge 디버그 창 + 세무사 계정 홈택스 로그인 확인 후 계속\n")

    # 엑셀 로드
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    ensure_header(ws)

    # 고객 목록 읽기
    customers = load_customers_from_xlsx(xlsx_path)
    total = len(customers)
    print(f"  총 {total}명, {start_idx+1}번부터 시작\n")

    # 결과 엑셀 (다운로드 현황용)
    result_wb, result_ws = ensure_output_workbook()

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://localhost:9222")
        ctx  = browser.contexts[0]
        # 메인 홈택스 페이지 찾기 (팝업이 pages[0]에 있을 수 있음)
        from 종합소득세안내문조회 import _find_main_page
        page = _find_main_page(ctx)
        print(f"  [메인 페이지] {page.url[:80]}", flush=True)
        page.bring_to_front()
        page.on("dialog", lambda d: d.dismiss())

        for i, c in enumerate(customers[start_idx:], start_idx + 1):
            name = c["name"]
            print(f"[{i}/{total}] {name}", flush=True)

            # 주민번호 없으면 스킵
            jumin = str(c.get("jumin_raw", "")).replace("-", "").replace(" ", "").strip()
            if not jumin or len(jumin) < 12:
                print(f"    [스킵] 주민번호 없음/불완전", flush=True)
                ws.cell(row=i + 1, column=3, value="주민번호없음")
                ws.cell(row=i + 1, column=3 + len(RESULT_COLS) - 1, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
                wb.save(xlsx_path)
                continue

            # 다운로드 실행
            r = process_one(ctx, page, c)

            # 파싱
            parsed = parse_customer(name, c["jumin_raw"])

            # 지급명세서/간이용역 유무 체크
            folder_candidates = list(CUSTOMER_DIR.glob(f"{name}_*")) + [CUSTOMER_DIR / name]
            folder = next((f for f in folder_candidates if f.is_dir()), None)
            jipgum = "O" if folder and list(folder.glob("지급명세서/*.pdf")) else "X"
            ganiyong = "O" if folder and list(folder.glob("간이용역소득/*.xlsx")) else "X"

            # 엑셀 기록
            row_idx = i + 1
            vals = [
                r["status"],
                parsed.get("수입금액총계", ""),
                parsed.get("기장의무", ""),
                parsed.get("추계시적용경비율", ""),
                parsed.get("이자", ""),
                parsed.get("배당", ""),
                parsed.get("근로(단일)", ""),
                parsed.get("근로(복수)", ""),
                parsed.get("연금", ""),
                parsed.get("기타", ""),
                parsed.get("전년도_납부할총세액", ""),
                parsed.get("부가세_납부", ""),
                r["biznos"],
                jipgum,
                ganiyong,
                r["error_msg"],
                datetime.now().strftime("%Y-%m-%d %H:%M"),
            ]
            for col_offset, val in enumerate(vals):
                ws.cell(row=row_idx, column=3 + col_offset, value=val)
            wb.save(xlsx_path)

            # 결과.xlsx에도 기록
            result_ws.append([
                name, str(c["jumin_raw"]), r["status"],
                r["anneam_pdf"], r["prev_income_xlsx"],
                r["biznos"], r["vat_xlsx_count"],
                r["error_msg"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ])
            from config import OUTPUT_DIR
            result_wb.save(str(OUTPUT_DIR / "결과_기장.xlsx"))

            _수입 = parsed.get('수입금액총계', '')
            _수입_str = f"{_수입:,}" if isinstance(_수입, (int, float)) else (str(_수입) if _수입 else '?')
            print(f"    → {r['status']} | 수입:{_수입_str} | 지급명세서:{jipgum} | {r['error_msg'] or ''}\n", flush=True)

    print(f"[완료] {total}명 처리 → {xlsx_path.name}")


if __name__ == "__main__":
    main()
