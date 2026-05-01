"""
빈 양식의 노란셀을 우리 파싱 데이터로 자동 채우기 - PoC

대상: 오상연 (간편장부, 프리만)
시나리오: 빈양식_간편장부.xlsx → "프리" 시트
출력: 작업판_오상연_2024.xlsx

원칙:
- 노란셀만 채움 (주황·검정·라벨 안 건드림)
- 확실한 매핑만 채움
- 추측 셀은 빈 채로 두고 메모 코멘트 추가
- archive 패턴 적용
"""
import openpyxl
from openpyxl.comments import Comment
from copy import copy
from pathlib import Path
import shutil
import warnings
warnings.filterwarnings("ignore")

import sys
sys.path.insert(0, r"F:\종소세2026")
from safe_save import safe_save_workbook

SRC_TEMPLATE = Path(r"C:\Users\pc\OneDrive\문서\카카오톡 받은 파일\빈양식_간편장부.xlsx")
OUT_DIR = Path(r"F:\종소세2026\output\PDF\오상연")
OUT_FILENAME = "작업판_오상연_2024.xlsx"

CUSTOMER = {
    "성명": "오상연",
    "생년월일": "84.12.12",
    "기장의무": "간편장부대상자",
    "추계시적용경비율": "기준경비율",
    "사업자번호": "",  # 없음
    "업종코드": "940909",  # 안내문에서
    "수입금액": 52663077,
    "이자": "X", "배당": "X", "근로_단일": "X", "근로_복수": "X",
    "연금": "X", "기타": "X",
}


def is_yellow(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.rgb is None:
        return False
    rgb = str(fg.rgb).upper()
    if len(rgb) < 6: return False
    rgb6 = rgb[-6:]
    try:
        r, g, b = int(rgb6[0:2],16), int(rgb6[2:4],16), int(rgb6[4:6],16)
    except:
        return False
    return r >= 200 and g >= 200 and b < 120


def fill_pri_sheet(ws, customer):
    """프리 시트: 사업소득(프리) + 근로/연금/기타 해당여부"""
    filled = []
    suspicious = []

    # 노란셀 모두 식별
    yellow_cells = []
    for row in ws.iter_rows():
        for c in row:
            if is_yellow(c):
                yellow_cells.append(c)

    # 매핑 정의 (보수적 - 확실한 것만)
    mapping = {
        "B2": ("성명", customer["성명"]),
        "E11": ("프리 수입금액", customer["수입금액"]),
        "F11": ("프리 업종코드", customer["업종코드"]),
        # 근로/연금/기타 해당여부 (안내문 타소득)
        "E12": ("근로 해당여부", customer["근로_단일"] if customer["근로_단일"] == "O" else ""),
        "E13": ("연금 해당여부", customer["연금"] if customer["연금"] == "O" else ""),
        "E14": ("기타 해당여부", customer["기타"] if customer["기타"] == "O" else ""),
    }

    for c in yellow_cells:
        coord = c.coordinate
        if coord in mapping:
            label, val = mapping[coord]
            c.value = val
            filled.append(f"{coord}={val} ({label})")
        else:
            # 자동 채우지 못한 노란셀에 코멘트
            c.comment = Comment("⚠️ 자동 채움 매핑 미정 - 확인 필요", "Bot")
            suspicious.append(coord)

    return filled, suspicious


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # 빈양식 복사 → 작업
    print(f"[1] 빈양식 로드: {SRC_TEMPLATE.name}")
    wb = openpyxl.load_workbook(SRC_TEMPLATE)

    target_sheet = "프리"
    if target_sheet not in wb.sheetnames:
        print(f"[에러] 시트 '{target_sheet}' 없음")
        return
    ws = wb[target_sheet]
    print(f"[2] '{target_sheet}' 시트 작업 시작 ({ws.max_row}행 x {ws.max_column}열)")

    filled, suspicious = fill_pri_sheet(ws, CUSTOMER)

    print(f"\n[3] 채운 셀 ({len(filled)}개):")
    for f in filled:
        print(f"    {f}")
    print(f"\n[4] 매핑 미정 노란셀 ({len(suspicious)}개) - 확인필요 코멘트 부착:")
    for s in suspicious:
        print(f"    {s}")

    # 다른 시트는 삭제 (오상연용은 프리만 필요)
    for sn in list(wb.sheetnames):
        if sn != target_sheet and sn not in ("참조", "업종코드"):
            del wb[sn]

    # archive 패턴 저장
    print(f"\n[5] 저장")
    status, target = safe_save_workbook(wb, OUT_DIR, OUT_FILENAME)
    print(f"    [{status}] {target}")


if __name__ == "__main__":
    main()
