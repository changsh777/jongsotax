"""
양식의 노란셀 전체를 표 형태로 정리한 '회의용 양식' 생성.
직원 회의에서 각 셀에 대해 분류(A/B) 정하기 위한 작업표.

컬럼:
  양식종류 | 시트명 | 셀좌표 | 원본값(샘플) | 인접라벨 | 분류(A/B) | 데이터출처/메모
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

SOURCES = {
    "간편장부": r"C:\Users\pc\OneDrive\문서\카카오톡 받은 파일\종소세자동화-박찬영 (2025신고대리간편장부).xlsx",
    "복식부기": r"C:\Users\pc\OneDrive\문서\카카오톡 받은 파일\종소세자동화-오지혜(2025신고대리복식).xlsx",
}
OUT = r"F:\종소세2026\templates\회의용_셀분류표.xlsx"

GUIDE_KEYWORDS = ["입력", "셀", "금지", "수식", "노랑", "노란", "검정", "검은"]


def is_yellow(cell):
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    fg = fill.fgColor
    if fg is None or fg.rgb is None:
        return False
    rgb = str(fg.rgb).upper()
    if not rgb or len(rgb) < 6:
        return False
    rgb6 = rgb[-6:]
    try:
        r = int(rgb6[0:2], 16)
        g = int(rgb6[2:4], 16)
        b = int(rgb6[4:6], 16)
    except Exception:
        return False
    return r >= 200 and g >= 200 and b < 120


def is_guide_text(value):
    if not isinstance(value, str):
        return False
    return any(k in value for k in GUIDE_KEYWORDS)


def find_adjacent_label(ws, row, col):
    """셀 좌측·상단 인접 라벨(텍스트) 찾기 - 최대 3칸 떨어진 곳까지"""
    labels = []
    # 같은 행 좌측
    for offset in range(1, 4):
        if col - offset < 1:
            break
        v = ws.cell(row=row, column=col - offset).value
        if v and isinstance(v, str) and not v.replace(",", "").replace(".", "").replace("-", "").isdigit():
            labels.append(f"←{v.strip()}")
            break
    # 같은 열 상단
    for offset in range(1, 4):
        if row - offset < 1:
            break
        v = ws.cell(row=row - offset, column=col).value
        if v and isinstance(v, str) and not v.replace(",", "").replace(".", "").replace("-", "").isdigit():
            labels.append(f"↑{v.strip()}")
            break
    return " / ".join(labels) if labels else "(라벨 없음)"


def collect_yellow_cells():
    rows = []
    for label, src in SOURCES.items():
        wb = openpyxl.load_workbook(src)
        for sheet_name in wb.sheetnames:
            if sheet_name in ("참조", "업종코드", "Sheet2", "2022귀속업종코드", "수지라"):
                continue  # 참조 데이터/업종표는 스킵
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if not is_yellow(cell):
                        continue
                    val = cell.value
                    if val is None:
                        continue
                    if is_guide_text(val):
                        continue
                    adj = find_adjacent_label(ws, cell.row, cell.column)
                    rows.append({
                        "양식": label,
                        "시트": sheet_name,
                        "좌표": cell.coordinate,
                        "샘플값": str(val)[:50],
                        "인접라벨": adj,
                    })
    return rows


def write_meeting_doc(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "셀분류"

    headers = ["양식", "시트", "셀좌표", "샘플값", "인접라벨", "분류(A/B)", "데이터출처/메모"]
    ws.append(headers)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([
            r["양식"],
            r["시트"],
            r["좌표"],
            r["샘플값"],
            r["인접라벨"],
            "",  # 분류 - 회의에서 채울 칸
            "",  # 데이터출처 - 회의에서 채울 칸
        ])

    # 컬럼 너비
    widths = [12, 16, 10, 30, 40, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 분류 컬럼은 노란색 강조 (직원이 채울 곳)
    fill_yellow = PatternFill("solid", fgColor="FFFF00")
    for row_idx in range(2, len(rows) + 2):
        ws.cell(row=row_idx, column=6).fill = fill_yellow
        ws.cell(row=row_idx, column=7).fill = fill_yellow

    # 안내 시트 (가장 앞)
    info = wb.create_sheet("안내", 0)
    info.append(["회의용 셀 분류표 - 사용 가이드"])
    info["A1"].font = Font(bold=True, size=14)
    guide = [
        "",
        "[목적]",
        "양식의 노란색 입력셀(N=총 N개) 하나하나에 대해",
        "1) 봇이 자동 채울지 (분류 A)",
        "2) 직원이 매입자료 보고 직접 채울지 (분류 B)",
        "결정한 다음 자동화 매핑을 확정합니다.",
        "",
        "[분류 기준]",
        "A (자동): 안내문 PDF·전년도 종소세 엑셀·부가세 엑셀에서 추출 가능한 정보",
        "    예: 성명, 사업자번호, 업종코드, 수입금액, 전년도 결정세액, 근로/연금 해당여부",
        "B (수동): 매입자료, 신용카드 사용액, 임차료, 지급이자 등 외부 자료 필요",
        "    예: 신용카드 매입, 급여, 임차료, 접대비, 기부금, 감가상각비",
        "",
        "[작성 방법]",
        "F열 '분류'에 A 또는 B 입력",
        "G열에 데이터 출처(예: '안내문 PDF 1페이지') 또는 메모(예: '카카오뱅크 거래내역') 입력",
        "",
        f"[전체 노란셀 개수] {len(rows)}개",
    ]
    for line in guide:
        info.append([line])
    info.column_dimensions["A"].width = 90

    wb.save(OUT)


rows = collect_yellow_cells()
write_meeting_doc(rows)

print(f"[완료] {OUT}")
print(f"  총 노란셀 {len(rows)}개")
# 양식별 통계
from collections import Counter
by_form_sheet = Counter((r["양식"], r["시트"]) for r in rows)
for (form, sheet), n in sorted(by_form_sheet.items()):
    print(f"  - {form} / {sheet}: {n}개")
