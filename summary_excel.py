import sys, io, os
os.environ.setdefault("SEOTAX_ENV", "nas")
_saved_stdout = sys.stdout

"""
216명 고객별 5가지 자료 → 1행 1고객 요약 엑셀
출력: Z:\종소세2026\output\종소세요약.xlsx
"""

sys.path.insert(0, r"F:\종소세2026")

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

from config import CUSTOMER_DIR, OUTPUT_DIR
from parse_to_xlsx import parse_anneam, parse_prev_income_xlsx, parse_vat_xlsx
sys.stdout = _saved_stdout  # parse_to_xlsx가 래핑한 stdout 복원
from gsheet_writer import get_credentials

import gspread
import warnings, logging

warnings.filterwarnings("ignore")
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)

# ── 설정 ────────────────────────────────────────────
GSHEET_ID  = "1oh31k00Oa2lZWvu5fnBRVmurdlll1YEG8Fefi5FRfBI"
SHEET_NAME = "접수명단"
COL_NAME   = 2   # B열 = 성명
COL_JUMIN  = 4   # D열 = 주민번호 (앞 6자리)

OUT_XLSX = OUTPUT_DIR / "종소세요약.xlsx"

# 출력 컬럼 정의
SUMMARY_COLS = [
    "순번", "성명", "생년월일",
    # 안내문
    "수입금액총계", "기장의무", "추계시적용경비율",
    "이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타",
    # 지급명세서·간이용역
    "지급명세서", "간이용역",
    # 전년도 소득세
    "전년도_납부할총세액",
    # 부가세
    "부가세_납부합계",
    # 비고
    "비고",
]

NUMERIC_COLS = {"수입금액총계", "전년도_납부할총세액", "부가세_납부합계"}

# 스타일 상수
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
FILL_GRAY   = PatternFill("solid", fgColor="BFBFBF")   # 자료없음 행
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # O
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")    # X


# ── 구글시트에서 고객 목록 읽기 ─────────────────────
def load_customers_from_gsheet() -> list[dict]:
    """접수명단 시트 B열(성명), D열(주민번호) 읽기 → [{"name": str, "jumin6": str}]"""
    creds = get_credentials()
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(GSHEET_ID)
    ws = sh.worksheet(SHEET_NAME)

    all_rows = ws.get_all_values()  # 헤더 포함
    customers = []
    for row in all_rows[1:]:  # 헤더 제외
        if len(row) < COL_NAME:
            continue
        name = row[COL_NAME].strip()   # 0-based
        if not name:
            continue
        jumin_raw = row[COL_JUMIN].strip() if len(row) > COL_JUMIN else ""
        # 주민번호 앞 6자리 추출 (850101 or 850101-1234567)
        jumin6 = jumin_raw.replace("-", "")[:6] if jumin_raw else ""
        customers.append({"name": name, "jumin6": jumin6})

    print(f"[구글시트] 접수명단 {len(customers)}명 로드")
    return customers


# ── 고객 폴더 탐색 ──────────────────────────────────
def find_customer_folder(name: str, jumin6: str) -> Path | None:
    """성명_* 패턴으로 탐색. 주민번호 앞6자리가 있으면 정확히 매칭."""
    candidates = list(CUSTOMER_DIR.glob(f"{name}_*"))
    if not candidates:
        # 이름만으로 된 폴더도 시도
        plain = CUSTOMER_DIR / name
        if plain.is_dir():
            return plain
        return None

    if jumin6:
        # 정확한 jumin6 매칭 우선
        exact = [c for c in candidates if c.name.endswith(f"_{jumin6}") and c.is_dir()]
        if exact:
            return exact[0]

    # 그냥 첫 번째 폴더
    dirs = [c for c in candidates if c.is_dir()]
    return dirs[0] if dirs else None


# ── 파일 존재 여부 체크 ────────────────────────────
def has_jibup(folder: Path) -> bool:
    """지급명세서 폴더에 PDF 있으면 O"""
    jibup_dir = folder / "지급명세서"
    if not jibup_dir.is_dir():
        return False
    return bool(list(jibup_dir.glob("*.pdf")))


def has_ganyi(folder: Path) -> bool:
    """간이용역소득 폴더에 xlsx 있으면 O"""
    ganyi_dir = folder / "간이용역소득"
    if not ganyi_dir.is_dir():
        return False
    return bool(list(ganyi_dir.glob("*.xlsx")))


# ── 안내문 PDF 탐색 ─────────────────────────────────
def find_anneam_pdf(folder: Path) -> Path | None:
    candidates = list(folder.glob("종소세안내문_*.pdf"))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


# ── 고객 1명 처리 ────────────────────────────────────
def process_one(name: str, jumin6: str) -> dict:
    row = {c: "" for c in SUMMARY_COLS}
    row["성명"] = name

    folder = find_customer_folder(name, jumin6)
    if folder is None:
        row["비고"] = "폴더없음"
        return row

    # 안내문 파싱
    pdf = find_anneam_pdf(folder)
    if pdf:
        try:
            parsed = parse_anneam(pdf)
            row["생년월일"]       = parsed.get("생년월일", "")
            row["수입금액총계"]   = parsed.get("수입금액총계", "")
            row["기장의무"]       = parsed.get("기장의무", "")
            row["추계시적용경비율"] = parsed.get("추계시적용경비율", "")
            row["이자"]           = parsed.get("이자", "")
            row["배당"]           = parsed.get("배당", "")
            row["근로(단일)"]     = parsed.get("근로(단일)", "")
            row["근로(복수)"]     = parsed.get("근로(복수)", "")
            row["연금"]           = parsed.get("연금", "")
            row["기타"]           = parsed.get("기타", "")
        except Exception as e:
            row["비고"] = f"안내문파싱오류:{e}"
    else:
        # 비고에 안내문없음만 표시 (나머지는 계속 파싱)
        row["비고"] = "안내문없음"

    # 지급명세서 / 간이용역 존재 여부
    row["지급명세서"] = "O" if has_jibup(folder) else "X"
    row["간이용역"]   = "O" if has_ganyi(folder) else "X"

    # 전년도 소득세
    prev = parse_prev_income_xlsx(folder)
    if prev:
        row["전년도_납부할총세액"] = prev.get("전년도_납부할총세액", "")

    # 부가세 (여러 사업자 합산 - parse_vat_xlsx가 이미 합산해서 반환)
    vat = parse_vat_xlsx(folder)
    if vat:
        row["부가세_납부합계"] = vat.get("부가세_납부", "")

    return row


# ── 엑셀 출력 ────────────────────────────────────────
def write_summary_xlsx(rows: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "종소세요약"

    # 헤더
    ws.append(SUMMARY_COLS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터
    for i, row in enumerate(rows, start=2):
        ws_row = [row[c] for c in SUMMARY_COLS]
        ws.append(ws_row)

        # 자료없음 행 (안내문없음 + 폴더없음): 회색
        bigo = row.get("비고", "")
        if bigo in ("폴더없음", "안내문없음"):
            for cell in ws[i]:
                cell.fill = FILL_GRAY

        # 지급명세서 컬럼 색상
        jibup_col_idx = SUMMARY_COLS.index("지급명세서") + 1
        jibup_cell = ws.cell(row=i, column=jibup_col_idx)
        jibup_cell.alignment = Alignment(horizontal="center")
        if jibup_cell.value == "O":
            jibup_cell.fill = FILL_GREEN
        elif jibup_cell.value == "X":
            jibup_cell.fill = FILL_RED

        # 간이용역 컬럼 색상
        ganyi_col_idx = SUMMARY_COLS.index("간이용역") + 1
        ganyi_cell = ws.cell(row=i, column=ganyi_col_idx)
        ganyi_cell.alignment = Alignment(horizontal="center")
        if ganyi_cell.value == "O":
            ganyi_cell.fill = FILL_GREEN
        elif ganyi_cell.value == "X":
            ganyi_cell.fill = FILL_RED

        # 숫자 컬럼 천단위 콤마
        for col_name in NUMERIC_COLS:
            col_idx = SUMMARY_COLS.index(col_name) + 1
            cell = ws.cell(row=i, column=col_idx)
            if isinstance(cell.value, (int, float)) and cell.value != "":
                cell.number_format = "#,##0"

    # 컬럼 너비
    col_widths = {
        "순번": 6, "성명": 10, "생년월일": 12,
        "수입금액총계": 16, "기장의무": 18, "추계시적용경비율": 16,
        "이자": 6, "배당": 6, "근로(단일)": 9, "근로(복수)": 9,
        "연금": 6, "기타": 6,
        "지급명세서": 10, "간이용역": 10,
        "전년도_납부할총세액": 18, "부가세_납부합계": 15,
        "비고": 20,
    }
    for i, col_name in enumerate(SUMMARY_COLS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = col_widths.get(col_name, 10)

    # 순번 채우기 (1부터)
    num_col_idx = SUMMARY_COLS.index("순번") + 1
    for i, _ in enumerate(rows, start=2):
        ws.cell(row=i, column=num_col_idx).value = i - 1

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


# ── 메인 ─────────────────────────────────────────────
def main():
    customers = load_customers_from_gsheet()
    total = len(customers)

    rows = []
    for idx, c in enumerate(customers, start=1):
        name   = c["name"]
        jumin6 = c["jumin6"]

        try:
            row = process_one(name, jumin6)
        except Exception as e:
            row = {c_: "" for c_ in SUMMARY_COLS}
            row["성명"] = name
            row["비고"] = f"처리오류:{e}"

        rows.append(row)

        if idx % 100 == 0 or idx == total:
            print(f"  진행: {idx}/{total}명 완료")

    write_summary_xlsx(rows)

    folder_err = sum(1 for r in rows if r.get("비고") == "폴더없음")
    pdf_err    = sum(1 for r in rows if r.get("비고") == "안내문없음")
    print(f"\n완료: {total}명 처리")
    print(f"  - 폴더없음: {folder_err}명")
    print(f"  - 안내문없음: {pdf_err}명")
    print(f"저장 경로: {OUT_XLSX}")


if __name__ == "__main__":
    main()
