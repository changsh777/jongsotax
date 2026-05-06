"""
jakupan_gen.py - 작업판 엑셀에 고객 데이터 자동 입력

처리:
  1. 작업판 템플릿(참조/업종코드/프리 시트) 로드
  2. 프리 시트 → 자동 입력 가능한 노랑셀 채우기
  3. 작업준비 시트 추가 (①②③④ 레이아웃)
  4. 고객 폴더에 저장: 작업판_{이름}.xlsx

사용법:
  python jakupan_gen.py 황순영 800315
  python jakupan_gen.py --all
"""
import sys, io, os
_saved_stdout = sys.stdout
os.environ.setdefault("SEOTAX_ENV", "nas")
_proj = str(__import__('pathlib').Path(__file__).resolve().parent)
if _proj not in sys.path:
    sys.path.insert(0, _proj)

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import xlrd, re

from parse_to_xlsx import parse_anneam, parse_prev_income_xlsx
sys.stdout = _saved_stdout

from config import CUSTOMER_DIR, OUTPUT_DIR, TEMPLATES_DIR
from gsheet_writer import get_credentials
import gspread

# ── 템플릿 경로 ──────────────────────────────────────────────────
TEMPLATE_간편 = TEMPLATES_DIR / "빈양식_간편장부.xlsx"
TEMPLATE_복식 = TEMPLATES_DIR / "빈양식_복식부기.xlsx"


def select_template_and_sheet(biz_rows: list, 기장의무: str) -> tuple:
    """안내문 파싱 결과로 템플릿·시트 선택
    반환: (template_path, sheet_name)
    """
    reg_count  = sum(1 for b in biz_rows if b.get("사업자번호"))
    free_count = sum(1 for b in biz_rows if not b.get("사업자번호"))
    is_복식    = "복식" in 기장의무

    if is_복식:
        template = TEMPLATE_복식
        sheet    = "사업자복식" if reg_count > 0 else "프리복식"
    else:
        template = TEMPLATE_간편
        if reg_count >= 2:
            sheet = "사업자+사업자"
        elif reg_count == 1 and free_count > 0:
            sheet = "사업자+프리"
        elif free_count >= 2:
            sheet = "프리+프리"
        else:
            sheet = "프리"   # 기본값 (프리 1개 또는 biz_rows 없음)

    return template, sheet

GSHEET_ID  = "1oh31k00Oa2lZWvu5fnBRVmurdlll1YEG8Fefi5FRfBI"
SHEET_NAME = "접수명단"
COL_NAME   = 2
COL_JUMIN  = 4

# ── print_sheet 에서 필요한 함수들 재사용 ───────────────────────
from print_sheet import (
    find_folder, find_anneam, parse_anneam_biz, parse_jibup_pdf,
    read_ganyi, read_vat_raw, parse_prev_income_xlsx,
    build_sheet, to_num,
    FILL_TITLE, FILL_SEC, FILL_HDR, FILL_PCT, FILL_WARN, FILL_OK, FILL_NO,
    FONT_TITLE, FONT_SEC, FONT_HDR, FONT_DATA, FONT_PCT, FONT_WARN,
    B, CENTER, LEFT, RIGHT, NUM_FMT, NCOLS,
    sc, merge_sec, merge_span,
)


# ── 입력셀 초기화 & 채우기 ───────────────────────────────────────
# 초기화할 입력셀 목록 (수식셀 제외, 구조 레이블 제외)
CLEAR_CELLS = [
    # 성명
    "B2",
    # 장부유형 체크 (D열)
    "D3", "D4", "D5", "D6", "D7", "D8", "D9",
    # 소득종류 체크
    "D10", "D11", "D12", "D13", "D14", "D15", "D16",
    # 금액 입력
    "E10", "E11", "E12", "E13", "E14", "E15", "E16",
    # 업종코드
    "F11",
    # 경비 입력 (사람이 채울 영역 - 빈칸으로)
    "C19", "C20",
    "F19", "F20",
    "C29", "C30", "C31", "C32", "C33", "C34",
    "C35", "C36", "C37", "C38", "C39", "C40",
    "C41", "C42", "C43", "C44",
    "F29", "F30", "F31", "F32", "F33", "F34",
    "F35", "F36", "F37", "F38", "F39", "F40",
    "F41", "F42", "F43", "F44",
    # 소득공제
    "C49", "C50", "C51", "C52", "C53",
    "C54", "C55", "C56", "C57", "C58",
    # 세액공제
    "C60", "C61",
    # 오상연 특화 데이터 (J30:L31)
    "J30", "K30", "L30", "J31", "K31", "L31",
]


def detect_income_from_files(folder: Path) -> set:
    """지급명세서/간이용역소득 폴더 파일명에서 소득종류 감지
    반환: {"사업소득", "기타소득", "연금소득", "근로소득", "이자소득", "배당소득"} 중 해당 항목
    """
    import unicodedata
    found = set()
    keywords = {
        "사업소득": "사업소득",
        "기타소득": "기타소득",
        "연금소득": "연금소득",
        "근로소득": "근로소득",
        "이자소득": "이자소득",
        "배당소득": "배당소득",
    }
    for sub in ["지급명세서", "간이용역소득"]:
        d = folder / sub
        if not d.is_dir():
            continue
        for f in d.iterdir():
            fname = unicodedata.normalize("NFC", f.name)
            for kw in keywords:
                if kw in fname:
                    found.add(kw)
    return found


def fill_puri(ws_puri, name, ann, biz_rows, ganyi_rows, ann_raw, file_income_types=None):
    """프리 시트 노랑셀 자동 입력"""

    # 초기화
    for addr in CLEAR_CELLS:
        ws_puri[addr].value = None

    # 성명
    ws_puri["B2"].value = name

    # ── 장부유형 ──────────────────────────────────────────────
    gijang    = str(ann.get("기장의무", "")).strip()
    choegye   = str(ann.get("추계시적용경비율", "")).strip()
    수입합계  = to_num(ann.get("수입금액총계", 0)) or 0

    # 복식부기의무자
    if "복식부기" in gijang:
        ws_puri["D3"].value = "해당"
        if 수입합계 >= 60_000_000:
            ws_puri["D5"].value = "해당"   # 간장복식 (6천만이상)
        else:
            ws_puri["D5"].value = "N/A"
    else:
        ws_puri["D3"].value = None
        ws_puri["D5"].value = "N/A"

    # 간편장부대상자
    if "간편" in gijang:
        ws_puri["D6"].value = "해당"

    # 단순경비율
    if "단순" in choegye:
        ws_puri["D9"].value = "해당"
    elif "기준" in choegye:
        ws_puri["D9"].value = None  # 기준경비율 → G유형 아님

    # ── 소득종류 ──────────────────────────────────────────────
    # 사업자등록 여부 (안내문 사업장별수입금액에 사업자번호 있음)
    has_bizno_reg = any(bz.get("사업자번호") for bz in biz_rows)
    if has_bizno_reg:
        ws_puri["D10"].value = "해당"
        # 사업자 수입금액 (업종별 첫 번째)
        ws_puri["E10"].value = biz_rows[0]["수입금액"] if biz_rows else None

    # 프리랜서 (미등록 사업소득 or 간이용역소득)
    has_freelance = any(not bz.get("사업자번호") for bz in biz_rows) or bool(ganyi_rows)
    if has_freelance:
        ws_puri["D11"].value = "해당"
        # 수입금액: 미등록 사업소득 합산 or 간이용역소득 합산
        freelance_amt = sum(
            bz["수입금액"] for bz in biz_rows if not bz.get("사업자번호")
        )
        if not freelance_amt and ganyi_rows:
            freelance_amt = sum(
                to_num(g.get("총지급액", 0)) or 0 for g in ganyi_rows
            )
        ws_puri["E11"].value = freelance_amt if freelance_amt else None

        # 업종코드: 첫 번째
        upjong = next(
            (bz["업종코드"] for bz in biz_rows if not bz.get("사업자번호")), None
        )
        if not upjong and ganyi_rows:
            # 간이용역소득에서 업종코드 직접 없음 → 안내문 biz_rows 참조
            upjong = biz_rows[0]["업종코드"] if biz_rows else None
        ws_puri["F11"].value = str(upjong) if upjong else None
    else:
        ws_puri["D11"].value = None

    fit = file_income_types or set()  # 파일명 기반 소득종류

    # 근로소득
    근로단일 = str(ann_raw.get("근로(단일)", "X"))
    근로복수 = str(ann_raw.get("근로(복수)", "X"))
    if "O" in (근로단일, 근로복수) or "근로소득" in fit:
        ws_puri["D12"].value = "해당"
    else:
        ws_puri["D12"].value = "N/A"

    # 연금
    연금 = str(ann_raw.get("연금", "X"))
    ws_puri["D13"].value = "해당" if ("O" in 연금 or "연금소득" in fit) else None

    # 기타
    기타 = str(ann_raw.get("기타", "X"))
    ws_puri["D14"].value = "해당" if ("O" in 기타 or "기타소득" in fit) else "N/A"
    ws_puri["E14"].value = 0

    # 금융 (이자+배당)
    이자 = str(ann_raw.get("이자", "X"))
    배당 = str(ann_raw.get("배당", "X"))
    ws_puri["D15"].value = "해당" if ("O" in 이자 or "O" in 배당
                                      or "이자소득" in fit or "배당소득" in fit) else None

    # 주택임대 (안내문에 별도 표시 없음 → 기본 0)
    ws_puri["E16"].value = 0


# ── 작업준비 시트 추가 ──────────────────────────────────────────
def add_jakupjunbi_sheet(wb, name, jumin6, folder):
    """작업판 워크북에 작업준비 시트 추가"""
    sname = f"작업준비_{name}"[:31]

    # 기존 시트가 있으면 삭제
    if sname in wb.sheetnames:
        del wb[sname]

    ws_new = wb.create_sheet(sname, 0)  # 맨 앞에 삽입

    # A4 인쇄 설정
    ws_new.page_setup.paperSize   = ws_new.PAPERSIZE_A4
    ws_new.page_setup.orientation = "portrait"
    ws_new.page_setup.fitToPage   = True
    ws_new.page_setup.fitToHeight = 1
    ws_new.page_setup.fitToWidth  = 1
    ws_new.page_margins.left   = 0.5
    ws_new.page_margins.right  = 0.5
    ws_new.page_margins.top    = 0.5
    ws_new.page_margins.bottom = 0.5

    last_row = build_sheet(ws_new, name, jumin6, folder)
    ws_new.print_area = f"A1:{get_column_letter(NCOLS)}{last_row - 1}"

    return ws_new


# ── 메인 처리 ────────────────────────────────────────────────────
def make_jakupan(name, jumin6="", force_jangbu: str = ""):
    """작업판 엑셀 생성.

    force_jangbu: "간편장부" 또는 "복식부기" 를 전달하면 안내문 기장의무 대신
                  해당 값을 기장의무로 강제 적용합니다.
                  빈 문자열이면 기존 자동 감지 로직을 사용합니다.
    """
    import unicodedata
    name = unicodedata.normalize("NFC", str(name))   # Mac NFD → NFC
    jumin6 = unicodedata.normalize("NFC", str(jumin6))
    folder = find_folder(name, jumin6)
    if folder is None:
        print(f"  [오류] 폴더 없음: {name}")
        return None

    # 데이터 파싱
    anneam_pdf = find_anneam(folder)
    ann_raw    = {}
    if anneam_pdf:
        try:
            ann_raw = parse_anneam(anneam_pdf) or {}
        except Exception:
            ann_raw = {}

    biz_rows   = parse_anneam_biz(anneam_pdf) if anneam_pdf else []
    ganyi_rows = read_ganyi(folder, jumin6)

    # 지급명세서/간이용역소득 폴더 파일명에서 소득종류 감지
    file_income_types = detect_income_from_files(folder)

    # 장부유형: force_jangbu 지정 시 강제, 아니면 안내문 자동 감지
    if force_jangbu in ("간편장부", "복식부기"):
        기장의무 = force_jangbu
        print(f"  [장부유형] 강제 지정: {force_jangbu}")
    else:
        기장의무 = str(ann_raw.get("기장의무", "")).strip()

    # 템플릿·시트 자동 선택
    template_path, sheet_name = select_template_and_sheet(biz_rows, 기장의무)
    print(f"  [템플릿] {template_path.name}  /  시트: {sheet_name}")

    if not template_path.exists():
        print(f"  [오류] 템플릿 없음: {template_path}")
        return None

    # 템플릿 로드
    wb = load_workbook(template_path)

    if sheet_name not in wb.sheetnames:
        print(f"  [오류] 시트 없음: {sheet_name}  (존재: {wb.sheetnames})")
        return None

    # 사용하지 않는 작업판 시트 삭제 (참조·업종코드·수지라·기본 등은 유지)
    ALL_WORKPAN = {"프리", "프리+프리", "사업자+프리", "사업자+사업자",
                   "프리복식", "사업자복식"}
    for sn in list(wb.sheetnames):
        if sn in ALL_WORKPAN and sn != sheet_name:
            del wb[sn]

    # 작업판 시트 채우기
    ws_puri = wb[sheet_name]
    fill_puri(ws_puri, name, ann_raw, biz_rows, ganyi_rows, ann_raw, file_income_types)

    # 작업준비 시트 추가
    add_jakupjunbi_sheet(wb, name, jumin6, folder)

    # 저장
    out = folder / f"작업판_{name}.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        print(f"  [스킵] {out.name} - Excel에서 열려있음. 닫고 다시 실행하세요.")
        return None
    return out


# ── 구글시트 고객 목록 ────────────────────────────────────────────
def load_customers():
    creds = get_credentials()
    gc = gspread.authorize(creds)
    ws = gc.open_by_key(GSHEET_ID).worksheet(SHEET_NAME)
    rows = ws.get_all_values()
    customers = []
    for row in rows[1:]:
        name  = row[COL_NAME].strip()  if len(row) > COL_NAME  else ""
        jumin = row[COL_JUMIN].strip() if len(row) > COL_JUMIN else ""
        if not name:
            continue
        customers.append({"name": name, "jumin6": jumin.replace("-", "")[:6]})
    return customers


# ── 엔트리포인트 ─────────────────────────────────────────────────
def main():
    args = sys.argv[1:]

    if args and args[0] == "--all":
        customers = load_customers()
        total = len(customers)
        ok = 0
        for i, c in enumerate(customers, 1):
            out = make_jakupan(c["name"], c["jumin6"])
            if out:
                ok += 1
            if i % 50 == 0 or i == total:
                print(f"  {i}/{total} 완료")
        print(f"\n[완료] {ok}/{total}명 생성")
    else:
        name         = args[0] if len(args) > 0 else "황순영"
        jumin6       = args[1] if len(args) > 1 else "800315"
        force_jangbu = args[2] if len(args) > 2 else ""
        out = make_jakupan(name, jumin6, force_jangbu)
        if out:
            print(f"저장: {out}")


if __name__ == "__main__":
    main()
