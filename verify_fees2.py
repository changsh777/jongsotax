import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
os.environ.setdefault("SEOTAX_ENV", "nas")
sys.path.insert(0, r"F:\종소세2026")

import openpyxl
from config import PARSE_RESULT_XLSX
from fee_calculator import calculate_fee

wb = openpyxl.load_workbook(PARSE_RESULT_XLSX, data_only=True)
ws = wb.active
headers = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(headers)}

other_cols = ["이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타"]
targets = {"김병수", "오민영", "김혜정"}

for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    name = str(row[0]).strip()
    if name not in targets:
        continue

    income = int(row[idx["수입금액총계"]] or 0)
    ledger = str(row[idx["기장의무"]] or "").strip()
    others = {c: str(row[idx.get(c, 0)] or "").strip() for c in other_cols}
    num_other = sum(1 for v in others.values() if v == "O")
    sheet_adv  = int(row[idx["사전접수할인가"]] or 0)
    sheet_norm = int(row[idx["일반접수가"]]    or 0)

    print(f"=== {name} ===")
    print(f"  수입: {income:,} / 기장의무: {ledger} / 타소득O: {num_other}개")
    print(f"  타소득 상세: {others}")
    print(f"  시트 수수료: 할인가={sheet_adv:,}  일반={sheet_norm:,}")
    for n in [0, 1, 2]:
        fa = calculate_fee(income, ledger, n, is_advance_booking=True)
        fn = calculate_fee(income, ledger, n, is_advance_booking=False)
        print(f"  타소득{n}개 기준: 할인가={fa['final_fee']:,}  일반={fn['final_fee']:,}")
    print()
