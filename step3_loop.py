"""
step3_loop.py
- 엑셀 명단 전체를 순회하며 종소세 안내문 PDF 다운로드
- 각 고객 사이에는 페이지 새로고침(goto)으로 상태 초기화 = 안정성 우선
- 결과는 매 건마다 저장 (중단 시 재개 가능)

사전 조건:
  1. launch_edge.bat 실행
  2. 홈택스 로그인 완료
  3. 신고도움서비스 진입 상태
"""
from playwright.sync_api import sync_playwright
from pathlib import Path
from datetime import datetime
import openpyxl
import time

INPUT_XLSX = Path(r"F:\종소세2026\input\종소세신고도움서비스테스트.xlsx")
OUTPUT_XLSX = Path(r"F:\종소세2026\output\결과.xlsx")
PDF_DIR = Path(r"F:\종소세2026\output\PDF")
PDF_DIR.mkdir(parents=True, exist_ok=True)

REPORT_HELP_URL = (
    "https://hometax.go.kr/websquare/websquare.html"
    "?w2xPath=/ui/pp/index_pp.xml"
    "&tmIdx=06&tm2lIdx=0601000000&tm3lIdx=0601200000"
)


def normalize_jumin(raw):
    """주민번호 정규화: 숫자/하이픈 섞인 입력을 (앞6, 뒤7)로"""
    s = str(raw).replace("-", "").replace(" ", "").strip()
    if len(s) != 13 or not s.isdigit():
        raise ValueError(f"주민번호 형식 이상: {raw}")
    return s[:6], s[6:]


def read_customers(xlsx_path):
    """엑셀에서 (성명, 주민번호) 뽑기. 열 순서: Name, 고객구분, 성명, 핸드폰, 주민번호, ..."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    customers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[2]        # 성명
        jumin_raw = row[4]   # 주민번호
        if not name or not jumin_raw:
            continue
        customers.append({"name": str(name).strip(), "jumin_raw": jumin_raw})
    return customers


def process_one(ctx, page, customer):
    """한 고객 처리 → dict(status, pdf_path, error_msg)"""
    name = customer["name"]

    try:
        front, back = normalize_jumin(customer["jumin_raw"])
    except ValueError as e:
        return {"status": "에러", "pdf_path": "", "error_msg": str(e)}

    try:
        # 페이지 새로고침(= 상태 초기화)
        page.goto(REPORT_HELP_URL, wait_until="domcontentloaded")
        time.sleep(2)

        # 주민번호 입력 (보이는 input 2개)
        all_inputs = page.locator(
            "xpath=//th[contains(normalize-space(.),'주민등록번호')]"
            "/following-sibling::td//input"
        ).all()
        visible = [el for el in all_inputs if el.is_visible()]
        if len(visible) < 2:
            return {"status": "에러", "pdf_path": "",
                    "error_msg": "주민번호 필드 2개 찾지 못함"}
        visible[0].fill(front)
        visible[1].fill(back)

        # 조회하기 (input[value=조회하기])
        btns = page.locator(
            "xpath=//th[contains(.,'주민등록번호')]"
            "/following-sibling::td//input[@value='조회하기']"
        ).all()
        visible_btns = [b for b in btns if b.is_visible()]
        if not visible_btns:
            return {"status": "에러", "pdf_path": "",
                    "error_msg": "조회하기 버튼 못 찾음"}
        visible_btns[0].click()

        time.sleep(3)

        # 미리보기 버튼 대기 (없으면 곧 에러 처리)
        preview = page.get_by_text("미리보기", exact=False).first
        try:
            preview.wait_for(timeout=10000, state="visible")
        except Exception:
            return {"status": "에러", "pdf_path": "",
                    "error_msg": "미리보기 버튼 없음 (조회 결과 없음 추정)"}

        # 미리보기 → 팝업 → PDF 저장
        with ctx.expect_page(timeout=15000) as popup_info:
            preview.click()
        popup = popup_info.value
        popup.wait_for_load_state("networkidle", timeout=30000)
        time.sleep(2)

        pdf_path = PDF_DIR / f"종소세안내문_{name}.pdf"
        popup.pdf(
            path=str(pdf_path),
            format="A4",
            print_background=True,
            margin={"top": "10mm", "bottom": "10mm", "left": "10mm", "right": "10mm"},
        )
        popup.close()

        return {"status": "완료", "pdf_path": str(pdf_path), "error_msg": ""}

    except Exception as e:
        return {"status": "에러", "pdf_path": "",
                "error_msg": f"{type(e).__name__}: {str(e)[:200]}"}


def ensure_output_workbook():
    if OUTPUT_XLSX.exists():
        wb = openpyxl.load_workbook(OUTPUT_XLSX)
        ws = wb.active
        return wb, ws
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "결과"
    ws.append(["성명", "주민번호", "처리상태", "PDF경로", "에러메시지", "시도일시"])
    wb.save(OUTPUT_XLSX)
    return wb, ws


def main():
    customers = read_customers(INPUT_XLSX)
    print(f"[시작] 총 {len(customers)}명 처리")

    wb, ws = ensure_output_workbook()

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://localhost:9222")
        ctx = browser.contexts[0]
        page = ctx.pages[0]
        page.bring_to_front()

        for i, customer in enumerate(customers, 1):
            print(f"\n[{i}/{len(customers)}] {customer['name']} 처리 중...")
            result = process_one(ctx, page, customer)

            ws.append([
                customer["name"],
                str(customer["jumin_raw"]),
                result["status"],
                result["pdf_path"],
                result["error_msg"],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ])
            wb.save(OUTPUT_XLSX)  # 매 건 저장

            if result["status"] == "완료":
                print(f"    완료 → {result['pdf_path']}")
            else:
                print(f"    에러 → {result['error_msg']}")

    print(f"\n[전체 완료] 결과 파일: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
