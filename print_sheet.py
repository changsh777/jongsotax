"""
print_sheet.py - 고객별 A4 1장 작업준비 시트 생성

사용법:
  python print_sheet.py 황순영 800315          (1명)
  python print_sheet.py --all                  (216명 전체)
"""
import sys, io, os
_saved_stdout = sys.stdout
os.environ.setdefault("SEOTAX_ENV", "nas")
sys.path.insert(0, r"F:\종소세2026")

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import xlrd

from parse_to_xlsx import parse_anneam, parse_prev_income_xlsx
sys.stdout = _saved_stdout  # parse_to_xlsx stdout 이중래핑 복원

from config import CUSTOMER_DIR, OUTPUT_DIR
from gsheet_writer import get_credentials
import gspread

# ── 스타일 상수 ─────────────────────────────────────────────────
FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
FILL_SEC   = PatternFill("solid", fgColor="2E75B6")
FILL_HDR   = PatternFill("solid", fgColor="D9E1F2")
FILL_PCT   = PatternFill("solid", fgColor="EBF0DE")
FILL_WARN  = PatternFill("solid", fgColor="FFEB9C")
FILL_OK    = PatternFill("solid", fgColor="C6EFCE")
FILL_NO    = PatternFill("solid", fgColor="FFC7CE")

FONT_TITLE = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF")
FONT_SEC   = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
FONT_HDR   = Font(name="맑은 고딕", size=8,  bold=True)
FONT_DATA  = Font(name="맑은 고딕", size=9)
FONT_PCT   = Font(name="맑은 고딕", size=8)
FONT_WARN  = Font(name="맑은 고딕", size=10, bold=True, color="9C5700")

THIN = Side(style="thin", color="BBBBBB")
B    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

NUM_FMT = "#,##0"

NCOLS = 12  # A~L

GSHEET_ID  = "1oh31k00Oa2lZWvu5fnBRVmurdlll1YEG8Fefi5FRfBI"
SHEET_NAME = "접수명단"
COL_NAME   = 2
COL_JUMIN  = 4


# ── 유틸 ────────────────────────────────────────────────────────
def to_num(v):
    if v is None or v == '':
        return None
    s = str(v).replace(',', '').replace(' ', '').strip()
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except Exception:
        return None


def sc(ws, row, col, val=None, font=None, fill=None, align=None,
       fmt=None, border=None):
    """셀 값·스타일 일괄 설정"""
    cell = ws.cell(row=row, column=col, value=val)
    if font:   cell.font   = font
    if fill:   cell.fill   = fill
    if align:  cell.alignment = align
    if fmt:    cell.number_format = fmt
    if border: cell.border = border
    return cell


def merge_sec(ws, row, val, font=FONT_SEC, fill=FILL_SEC, height=18):
    """섹션 헤더 행 (A1:L1 병합)"""
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=NCOLS)
    sc(ws, row, 1, val, font=font, fill=fill, align=CENTER, border=B)
    ws.row_dimensions[row].height = height


def merge_span(ws, row, c1, c2, val=None, font=None, fill=None,
               align=CENTER, fmt=None, border=B):
    """여러 열 병합 후 스타일"""
    if c2 > c1:
        ws.merge_cells(start_row=row, start_column=c1,
                       end_row=row,   end_column=c2)
    sc(ws, row, c1, val, font=font, fill=fill, align=align,
       fmt=fmt, border=border)
    # 병합 구간 나머지 셀 border 적용
    for cc in range(c1 + 1, c2 + 1):
        ws.cell(row, cc).border = border or B


def apply_border(ws, r1, r2, c1, c2):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = B


# ── 데이터 읽기 ─────────────────────────────────────────────────
def find_folder(name, jumin6):
    import unicodedata
    def nfc(s): return unicodedata.normalize("NFC", str(s))
    name_nfc = nfc(name)
    # glob 대신 iterdir + NFC 비교 (Mac SMB NFD 대응)
    candidates = [
        p for p in CUSTOMER_DIR.iterdir()
        if p.is_dir() and nfc(p.name).startswith(f"{name_nfc}_")
    ]
    if not candidates:
        p = CUSTOMER_DIR / name
        return p if p.is_dir() else None
    if jumin6:
        exact = [c for c in candidates if nfc(c.name).endswith(f"_{jumin6}")]
        if exact:
            return exact[0]
    return candidates[0] if candidates else None


def find_anneam(folder):
    import unicodedata
    ps = [p for p in folder.iterdir()
          if p.is_file()
          and unicodedata.normalize("NFC", p.name).startswith("종소세안내문_")
          and unicodedata.normalize("NFC", p.name).endswith(".pdf")]
    return max(ps, key=lambda p: p.stat().st_mtime) if ps else None


def parse_anneam_biz(pdf_path):
    """안내문 PDF 1페이지 → 사업장별수입금액 list of dict
    반환: [{사업자번호, 업종코드, 수입금액}, ...]
    """
    import pdfplumber, re
    try:
        with pdfplumber.open(pdf_path) as pdf:
            text = pdf.pages[0].extract_text() or ""
    except Exception:
        return []

    # "사업장별수입금액" ~ "총계" 구간 추출
    m_sec = re.search(r'사업장별수입금액(.+?)총계', text, re.DOTALL)
    if not m_sec:
        return []

    section = m_sec.group(1)
    # 라인 단위로 붙여가며 데이터 행 탐지
    # 데이터 행: 6자리 업종코드 + 수입금액(쉼표 포함 5자리 이상) 가 있는 라인
    combined = " ".join(section.split())  # 줄바꿈 통합

    records = []
    # 업종코드(6자리) 위치를 앵커로 분리
    pattern = re.compile(r'(\d{3}-\d{2}-\d{5})?\s*\S*\s*(\d{6})\s+\S+\s+\S+\s+\S+\s+([\d,]{4,})')
    for m in pattern.finditer(combined):
        bizno  = m.group(1) or ''
        upjong = m.group(2)
        amt    = int(m.group(3).replace(',', ''))
        records.append({'사업자번호': bizno, '업종코드': upjong, '수입금액': amt})

    # fallback: 더 느슨한 패턴
    if not records:
        for m in re.finditer(r'(\d{6})\b.*?([\d,]{5,})', combined):
            upjong = m.group(1)
            amt    = int(m.group(2).replace(',', ''))
            records.append({'사업자번호': '', '업종코드': upjong, '수입금액': amt})

    return records


def parse_jibup_pdf(folder):
    """지급명세서 PDF → list of dict {사업자번호, 징수의무자, 지급총액, 소득세, 지방소득세}"""
    import pdfplumber, re
    jdir = folder / "지급명세서"
    if not jdir.is_dir():
        return []
    pdfs = list(jdir.glob("*.pdf"))
    if not pdfs:
        return []

    records = []
    try:
        with pdfplumber.open(pdfs[0]) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                m_bizno = re.search(r'①사업자등록번호\s+([\d\-]+)', text)
                m_name  = re.search(r'②법인명 또는 상호\s+(.+?)(?:\n|③)', text)
                m_data  = re.search(
                    r'(\d{4})\s+\d{4}\s+([\d,]+)\s+\d+\s+([\d,]+)\s+([\d,]+)', text)
                if not m_data:
                    continue
                records.append({
                    '사업자번호':  m_bizno.group(1).strip() if m_bizno else '',
                    '징수의무자':  m_name.group(1).strip()  if m_name  else '',
                    '지급총액':   int(m_data.group(2).replace(',', '')),
                    '소득세':     int(m_data.group(3).replace(',', '')),
                    '지방소득세': int(m_data.group(4).replace(',', '')),
                })
    except Exception:
        pass
    return records


def read_ganyi(folder, jumin6=""):
    """간이용역소득 xlsx → list of dict
    - 비밀번호 있는 경우 jumin6(주민번호 앞 6자리)로 해제 시도
    """
    import msoffcrypto, io
    gdir = folder / "간이용역소득"
    if not gdir.is_dir():
        return []
    rows = []
    for xf in gdir.glob("*.xlsx"):
        try:
            # 먼저 암호화 여부 확인
            with open(xf, 'rb') as f:
                raw = f.read()
            try:
                of = msoffcrypto.OfficeFile(io.BytesIO(raw))
                is_enc = of.is_encrypted()
            except Exception:
                is_enc = False

            if is_enc and jumin6:
                # 주민번호(앞6자리 또는 13자리)로 해제 시도
                passwords = [jumin6]
                if len(jumin6) == 6:
                    # 생년월일 6자리 → 완전한 주민번호 없으면 앞6자리만 시도
                    passwords.append(jumin6)
                buf = None
                for pw in passwords:
                    try:
                        of2 = msoffcrypto.OfficeFile(io.BytesIO(raw))
                        decrypted = io.BytesIO()
                        of2.load_key(password=pw)
                        of2.decrypt(decrypted)
                        buf = decrypted.getvalue()
                        break
                    except Exception:
                        continue
                if buf is None:
                    print(f"  [경고] 비밀번호 해제 실패: {xf.name}")
                    continue
                wb = xlrd.open_workbook(file_contents=buf)
            else:
                wb = xlrd.open_workbook(str(xf))

            sheet = wb.sheet_by_index(0)
            if sheet.nrows < 2:
                continue
            hdrs = sheet.row_values(0)
            for i in range(1, sheet.nrows):
                vals = sheet.row_values(i)
                rows.append(dict(zip(hdrs, vals)))
        except Exception as e:
            print(f"  [경고] {xf.name} 읽기 실패: {e}")
    return rows


def read_vat_raw(folder):
    """부가세 xlsx → {bizno: [(row...), ...]} — 실제 데이터 있는 것만"""
    result = {}
    for f in folder.glob("부가세신고내역_*.xlsx"):
        bizno = f.stem.replace("부가세신고내역_", "")
        try:
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            has_real = any(
                any(kw in str(c) for kw in ['매출액', '매입액', '납부', '환급'])
                for row in rows for c in row if c
            )
            if has_real:
                result[bizno] = rows
        except Exception:
            pass
    return result


# ── 시트 레이아웃 구성 ───────────────────────────────────────────
def build_sheet(ws, name, jumin6, folder):
    # 컬럼 너비 (A-L)
    col_widths = [4, 11, 11, 10, 9, 9, 9, 9, 9, 9, 9, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1

    # ══ 타이틀 ══════════════════════════════════════════════════
    merge_sec(ws, r,
              f"  {name}  ({jumin6})     ·     2025년 종합소득세 작업준비",
              font=FONT_TITLE, fill=FILL_TITLE, height=26)
    r += 1

    # spacer
    ws.row_dimensions[r].height = 4
    r += 1

    # ══ ① 종소세 안내문 ══════════════════════════════════════════
    merge_sec(ws, r, "① 종소세 안내문")
    r += 1

    anneam_pdf = find_anneam(folder)
    ann = {}
    if anneam_pdf:
        try:
            ann = parse_anneam(anneam_pdf) or {}
        except Exception:
            ann = {}

    # 헤더: 9항목을 12열에 배치 (col spans 합계=12)
    AN_HDRS = ["수입금액합계",  "기장의무",  "추계경비율",
               "이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타"]
    AN_KEYS = ["수입금액총계", "기장의무", "추계시적용경비율",
               "이자", "배당", "근로(단일)", "근로(복수)", "연금", "기타"]
    AN_SPAN = [2, 2, 2, 1, 1, 1, 1, 1, 1]   # 합=12

    ws.row_dimensions[r].height = 28
    col = 1
    for hdr, span in zip(AN_HDRS, AN_SPAN):
        merge_span(ws, r, col, col + span - 1, hdr,
                   font=FONT_HDR, fill=FILL_HDR, align=CENTER)
        col += span
    r += 1

    ws.row_dimensions[r].height = 22
    col = 1
    for key, span in zip(AN_KEYS, AN_SPAN):
        raw = ann.get(key, "")
        if key == "수입금액총계":
            val = to_num(raw)
            fmt = NUM_FMT
        else:
            val = raw
            fmt = None
        merge_span(ws, r, col, col + span - 1, val,
                   font=FONT_DATA, align=CENTER, fmt=fmt)
        col += span

    if not anneam_pdf:
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r,   end_column=NCOLS)
        sc(ws, r, 1, "⚠  안내문 없음",
           font=FONT_WARN, fill=FILL_WARN, align=CENTER, border=B)
    r += 1

    # ── 사업장별 수입금액 (안내문 상세) ─────────────────────────
    biz_rows = parse_anneam_biz(anneam_pdf) if anneam_pdf else []
    if biz_rows:
        # 헤더: 사업자번호(3) | 업종코드(2) | 수입금액(4) | (공백3)
        BZ_HDRS = ["사업자번호", "업종코드", "수입금액"]
        BZ_SPAN = [4, 3, 5]   # 합=12
        ws.row_dimensions[r].height = 22
        col = 1
        for hdr, span in zip(BZ_HDRS, BZ_SPAN):
            merge_span(ws, r, col, col + span - 1, hdr,
                       font=FONT_HDR, fill=FILL_HDR, align=CENTER)
            col += span
        r += 1
        for bz in biz_rows:
            ws.row_dimensions[r].height = 18
            vals = [bz['사업자번호'] or '(미등록)',
                    bz['업종코드'],
                    bz['수입금액']]
            col = 1
            for val, span in zip(vals, BZ_SPAN):
                align = RIGHT if isinstance(val, (int, float)) else CENTER
                fmt   = NUM_FMT if isinstance(val, (int, float)) else None
                merge_span(ws, r, col, col + span - 1, val,
                           font=FONT_DATA, align=align, fmt=fmt)
                col += span
            r += 1

    # spacer
    ws.row_dimensions[r].height = 5
    r += 1

    # ══ ② 지급명세서 / 간이용역소득 대조 ══════════════════════════
    merge_sec(ws, r, "② 지급명세서 / 간이용역소득 대조")
    r += 1

    jibup_rows = parse_jibup_pdf(folder)          # 지급명세서 PDF 파싱
    ganyi_rows  = read_ganyi(folder, jumin6)      # 간이용역소득 xlsx (비밀번호 처리 포함)

    # 컬럼: 구분(1) | 사업자번호(2) | 징수의무자(2) | 업종(1) | 지급총액(2) | 소득세(2) | 지방소득세(2)
    G_HDRS = ["구분", "사업자번호", "징수의무자 / 소득종류", "업종", "지급총액", "소득세", "지방소득세"]
    G_SPAN = [1, 2, 3, 1, 2, 2, 1]   # 합=12

    FILL_DIFF  = PatternFill("solid", fgColor="FFF2CC")
    FILL_MISM  = PatternFill("solid", fgColor="FFC7CE")
    FONT_DIFF  = Font(name="맑은 고딕", size=8, italic=True, color="7F7F7F")
    FONT_MISM  = Font(name="맑은 고딕", size=8, bold=True, color="9C0006")

    # 헤더 행
    ws.row_dimensions[r].height = 24
    col = 1
    for hdr, span in zip(G_HDRS, G_SPAN):
        merge_span(ws, r, col, col + span - 1, hdr,
                   font=FONT_HDR, fill=FILL_HDR, align=CENTER)
        col += span
    r += 1

    # 사업자번호 기준으로 양쪽 합치기
    all_biznos = list(dict.fromkeys(
        [d['사업자번호'] for d in jibup_rows] +
        [d.get('사업자(주민)등록번호', '') for d in ganyi_rows]
    ))

    def write_data_row(ws, r, label, fill, bizno, company, upjong, amt, tax, loc):
        ws.row_dimensions[r].height = 18
        vals = [label, bizno, company, upjong, amt, tax, loc]
        col = 1
        for val, span in zip(vals, G_SPAN):
            align = RIGHT if isinstance(val, (int, float)) else CENTER
            fmt   = NUM_FMT if isinstance(val, (int, float)) else None
            merge_span(ws, r, col, col + span - 1, val,
                       font=FONT_DATA, fill=fill, align=align, fmt=fmt)
            col += span

    if not all_biznos:
        ws.row_dimensions[r].height = 16
        merge_span(ws, r, 1, NCOLS, "지급명세서 / 간이용역소득 자료 없음",
                   font=FONT_DATA, fill=FILL_NO, align=CENTER)
        r += 1
    else:
        for bizno in all_biznos:
            # 지급명세서 row
            jrow = next((d for d in jibup_rows if d['사업자번호'] == bizno), None)
            # 간이용역소득 row
            grow = next((d for d in ganyi_rows
                         if d.get('사업자(주민)등록번호', '') == bizno), None)

            # 지급명세서 행
            if jrow:
                write_data_row(ws, r, "지급명세서",
                               FILL_OK if jrow else FILL_NO,
                               bizno, jrow['징수의무자'], "",
                               jrow['지급총액'], jrow['소득세'], jrow['지방소득세'])
            else:
                ws.row_dimensions[r].height = 16
                merge_span(ws, r, 1, 1, "지급명세서",
                           font=FONT_DATA, fill=FILL_NO, align=CENTER)
                merge_span(ws, r, 2, 3, bizno,
                           font=FONT_DATA, fill=FILL_NO, align=LEFT)
                merge_span(ws, r, 4, NCOLS, "지급명세서 없음",
                           font=FONT_DATA, fill=FILL_NO, align=LEFT)
            r += 1

            # 간이용역소득 행
            if grow:
                write_data_row(ws, r, "간이용역소득", None,
                               bizno,
                               grow.get('징수의무자', ''),
                               grow.get('업종구분', ''),
                               to_num(grow.get('총지급액', '')),
                               to_num(grow.get('소득세', '')),
                               to_num(grow.get('지방소득세', '')))
            else:
                ws.row_dimensions[r].height = 16
                merge_span(ws, r, 1, 1, "간이용역소득",
                           font=FONT_DATA, fill=FILL_NO, align=CENTER)
                merge_span(ws, r, 2, 3, bizno,
                           font=FONT_DATA, fill=FILL_NO, align=LEFT)
                merge_span(ws, r, 4, NCOLS, "간이용역소득 없음",
                           font=FONT_DATA, fill=FILL_NO, align=LEFT)
            r += 1

            # 차이 행
            ws.row_dimensions[r].height = 14
            merge_span(ws, r, 1, 1, "차이",
                       font=FONT_DIFF, fill=FILL_DIFF, align=CENTER)
            merge_span(ws, r, 2, 3, "", fill=FILL_DIFF)
            merge_span(ws, r, 4, 4, "", fill=FILL_DIFF)
            col = 5  # 지급총액 시작
            for j_val, g_key in [
                (jrow['지급총액'] if jrow else None, '총지급액'),
                (jrow['소득세']   if jrow else None, '소득세'),
                (jrow['지방소득세'] if jrow else None, '지방소득세'),
            ]:
                g_val = to_num(grow.get(g_key, '')) if grow else None
                if j_val is not None and g_val is not None:
                    diff = j_val - g_val
                    diff_fill = FILL_MISM if diff != 0 else FILL_DIFF
                    diff_font = FONT_MISM if diff != 0 else FONT_DIFF
                else:
                    diff = "확인필요"
                    diff_fill = FILL_DIFF
                    diff_font = FONT_DIFF
                span = G_SPAN[col - 4]   # 지급총액=2, 소득세=2, 지방소득세=1
                merge_span(ws, r, col, col + span - 1, diff,
                           font=diff_font, fill=diff_fill, align=RIGHT,
                           fmt=NUM_FMT if isinstance(diff, int) else None)
                col += span
            r += 1

    # spacer
    ws.row_dimensions[r].height = 5
    r += 1

    # ══ ③ 전년도 소득세 ══════════════════════════════════════════
    merge_sec(ws, r, "③ 전년도 소득세 (2024년)")
    r += 1

    prev = parse_prev_income_xlsx(folder) or {}

    PREV_KEYS = [
        "전년도_총수입금액", "전년도_필요경비", "전년도_종합소득금액",
        "전년도_소득공제", "전년도_과세표준", "전년도_산출세액",
        "전년도_세액감면공제", "전년도_결정세액", "전년도_가산세",
        "전년도_기납부세액", "전년도_납부할총세액",
    ]
    PREV_HDRS = [
        "총수입금액", "필요경비", "종합소득금액",
        "소득공제", "과세표준", "산출세액",
        "세액감면공제", "결정세액", "가산세",
        "기납부세액", "납부총세액",
    ]

    # 헤더: A열=항목레이블, B~L=11개 컬럼
    ws.row_dimensions[r].height = 28
    sc(ws, r, 1, "항목", font=FONT_HDR, fill=FILL_HDR, align=CENTER, border=B)
    for i, hdr in enumerate(PREV_HDRS, 2):
        sc(ws, r, i, hdr, font=FONT_HDR, fill=FILL_HDR, align=CENTER, border=B)
    r += 1

    # 금액 행
    ws.row_dimensions[r].height = 20
    sc(ws, r, 1, "금액", font=FONT_DATA, align=CENTER, border=B)
    nums = {}
    for i, key in enumerate(PREV_KEYS, 2):
        val = to_num(prev.get(key, ''))
        nums[key] = val
        sc(ws, r, i, val, font=FONT_DATA, align=RIGHT, fmt=NUM_FMT, border=B)
    r += 1

    # % 행 (분모=총수입금액)
    ws.row_dimensions[r].height = 16
    sc(ws, r, 1, "%", font=FONT_PCT, fill=FILL_PCT, align=CENTER, border=B)
    denom = nums.get("전년도_총수입금액")
    for i, key in enumerate(PREV_KEYS, 2):
        val = nums.get(key)
        pct = (val / denom) if (val is not None and denom and denom != 0) else None
        sc(ws, r, i, pct, font=FONT_PCT, fill=FILL_PCT,
           align=CENTER, fmt="0.0%", border=B)
    r += 1

    if not prev:
        # 없으면 경고로 덮어쓰기
        for rr in range(r - 3, r):
            ws.merge_cells(start_row=rr, start_column=1,
                           end_row=rr,   end_column=NCOLS)
        sc(ws, r - 3, 1, "⚠  전년도 소득세 자료 없음",
           font=FONT_WARN, fill=FILL_WARN, align=CENTER, border=B)

    # spacer
    ws.row_dimensions[r].height = 5
    r += 1

    # ══ ④ 부가세 ════════════════════════════════════════════════
    merge_sec(ws, r, "④ 부가세 (2024년)")
    r += 1

    vat_data = read_vat_raw(folder)

    if not vat_data:
        # 자료 없음 → 수동 조회 안내
        ws.row_dimensions[r].height = 26
        merge_span(ws, r, 1, NCOLS,
                   "⚠  부가세 자료 없음  —  고객 아이디로 직접 조회 필요",
                   font=FONT_WARN, fill=FILL_WARN, align=CENTER)
        r += 1
    else:
        # 사업자번호별 1기/2기 표
        VAT_HDRS = ["사업자번호", "구분", "매출액", "매입액", "납부(환급)세액"]
        VAT_SPAN = [3, 2, 3, 2, 2]  # 합=12
        ws.row_dimensions[r].height = 22
        col = 1
        for hdr, span in zip(VAT_HDRS, VAT_SPAN):
            merge_span(ws, r, col, col + span - 1, hdr,
                       font=FONT_HDR, fill=FILL_HDR, align=CENTER)
            col += span
        r += 1

        for bizno, rows in vat_data.items():
            for vrow in rows:
                if not any(vrow):
                    continue
                cells = [str(c).strip() if c else '' for c in vrow]
                label = cells[0]
                if not label:
                    continue
                ws.row_dimensions[r].height = 16
                vals = [bizno, label] + [to_num(c) for c in cells[1:4]]
                col = 1
                for val, span in zip(vals + [''] * 5, VAT_SPAN):
                    align = RIGHT if isinstance(val, (int, float)) else LEFT
                    fmt   = NUM_FMT if isinstance(val, (int, float)) else None
                    merge_span(ws, r, col, col + span - 1, val,
                               font=FONT_DATA, align=align, fmt=fmt)
                    col += span
                r += 1

    return r


# ── Excel 파일 생성 ──────────────────────────────────────────────
def make_sheet(name, jumin6=""):
    folder = find_folder(name, jumin6)
    if folder is None:
        print(f"  [오류] 폴더 없음: {name}")
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]

    # A4 인쇄 설정
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToHeight = 1
    ws.page_setup.fitToWidth  = 1
    ws.page_margins.left   = 0.5
    ws.page_margins.right  = 0.5
    ws.page_margins.top    = 0.5
    ws.page_margins.bottom = 0.5

    last_row = build_sheet(ws, name, jumin6, folder)
    ws.print_area = f"A1:{get_column_letter(NCOLS)}{last_row - 1}"

    out = folder / f"작업준비_{name}.xlsx"
    wb.save(out)
    return out


# ── 구글시트 고객 목록 ────────────────────────────────────────────
def load_customers():
    creds = get_credentials()
    gc = gspread.authorize(creds)
    ws = gc.open_by_key(GSHEET_ID).worksheet(SHEET_NAME)
    rows = ws.get_all_values()
    customers = []
    for row in rows[1:]:
        name  = row[COL_NAME].strip()  if len(row) > COL_NAME  else ""
        jumin = row[COL_JUMIN].strip() if len(row) > COL_JUMIN else ""
        if not name:
            continue
        jumin6 = jumin.replace("-", "")[:6]
        customers.append({"name": name, "jumin6": jumin6})
    return customers


# ── 메인 ─────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]

    if args and args[0] == "--all":
        customers = load_customers()
        total = len(customers)
        ok = 0
        for i, c in enumerate(customers, 1):
            out = make_sheet(c["name"], c["jumin6"])
            if out:
                ok += 1
            if i % 50 == 0 or i == total:
                print(f"  {i}/{total} 완료")
        print(f"\n[완료] {ok}/{total}명 생성")
        print(f"저장 위치: {OUTPUT_DIR / '작업준비'}")

    else:
        name   = args[0] if len(args) > 0 else "황순영"
        jumin6 = args[1] if len(args) > 1 else "800315"
        out = make_sheet(name, jumin6)
        if out:
            print(f"저장: {out}")


if __name__ == "__main__":
    main()
