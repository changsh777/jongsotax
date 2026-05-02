import sys, time
sys.path.insert(0, r"F:\종소세2026")
from playwright.sync_api import sync_playwright
from 신규고객처리 import logout_hometax, login_hometax_id, INDIVIDUAL_PREVIEW_BTN_ID
from 종합소득세안내문조회 import save_anneam_pdf
from config import customer_folder

name       = "한효성"
hometax_id = "hyopary1351"
hometax_pw = "9107487hs*"
jumin_raw  = "720720-2123213"

with sync_playwright() as p:
    browser = p.chromium.connect_over_cdp("http://localhost:9222")
    ctx = browser.contexts[0]
    page = ctx.pages[0]
    page.bring_to_front()
    page.on("dialog", lambda d: d.dismiss())

    print("[1] 로그아웃", flush=True)
    logout_hometax(page)

    print("[2] 로그인 (2차인증 포함)", flush=True)
    ok = login_hometax_id(page, hometax_id, hometax_pw, jumin_raw=jumin_raw)
    print("결과:", ok, flush=True)
    if not ok:
        page.screenshot(path=r"F:\종소세2026\output\FAIL_han_login.png")
        sys.exit(1)
    time.sleep(1)

    print("[3] 신고도움서비스 이동", flush=True)
    page.evaluate("document.getElementById('menuAtag_4103080000').onclick()")
    time.sleep(4)

    txt = page.evaluate("() => document.body.innerText")
    if "로그인 정보가 없습니다" in txt:
        print("접근 실패!", flush=True)
        page.screenshot(path=r"F:\종소세2026\output\FAIL_han_report.png")
        sys.exit(1)

    print("[4] 미리보기 → PDF 저장", flush=True)
    preview_btn = page.locator(INDIVIDUAL_PREVIEW_BTN_ID)
    print("미리보기 visible:", preview_btn.is_visible(timeout=5000), flush=True)

    folder = customer_folder(name, jumin_raw)
    anneam_path = folder / f"종소세안내문_{name}.pdf"
    save_anneam_pdf(ctx, page, preview_btn, anneam_path)
    print("PDF 저장 완료:", anneam_path, flush=True)

# 파싱 + 구글시트
print("[5] 파싱 + 구글시트 업데이트", flush=True)
from pathlib import Path
from datetime import datetime
import openpyxl
from gspread.utils import rowcol_to_a1
from config import PARSE_RESULT_XLSX, CUSTOMER_DIR
from parse_to_xlsx import parse_anneam, parse_prev_income_xlsx, parse_vat_xlsx
from fee_calculator import calculate_fee, count_other_income
from gsheet_writer import get_credentials
import gspread

SPREADSHEET_ID = "1oh31k00Oa2lZWvu5fnBRVmurdlll1YEG8Fefi5FRfBI"
COL_MAP = {
    "수입":           "수입금액총계",
    "장부유형":       "기장의무",
    "추계시적용경비율": "추계시적용경비율",
    "할인가":         "사전접수할인가",
    "수수료":         "일반접수가",
    "이자":           "이자",
    "배당":           "배당",
    "근로(단일)":     "근로(단일)",
    "근로(복수)":     "근로(복수)",
    "연금":           "연금",
    "기타":           "기타",
}

folder_candidates = list(CUSTOMER_DIR.glob(f"{name}_*")) + [CUSTOMER_DIR / name]
folder = next((f for f in folder_candidates if f.is_dir()), None)
if not folder:
    print(f"[{name}] 폴더 없음")
    sys.exit(1)

pdf_candidates = list(folder.glob("종소세안내문_*.pdf"))
pdf_path = max(pdf_candidates, key=lambda p: p.stat().st_mtime)
print(f"파싱: {pdf_path.name}", flush=True)

data = parse_anneam(pdf_path)
data.update(parse_prev_income_xlsx(folder))
data.update(parse_vat_xlsx(folder))

income = int(data.get("수입금액총계") or 0)
ledger = data.get("기장의무", "")
num_other = count_other_income(data)

if income > 0 and ledger:
    fee     = calculate_fee(income, ledger, num_other, is_advance_booking=False)
    fee_adv = calculate_fee(income, ledger, num_other, is_advance_booking=True)
    data["사전접수할인가"] = fee_adv["final_fee"]
    data["일반접수가"]   = fee["final_fee"]

inc_str = f"{income:,}" if income else str(income)
print(f"  수입={inc_str} / {ledger} / {data.get('추계시적용경비율')} / 타소득{num_other}개", flush=True)

# 파싱결과.xlsx
wb = openpyxl.load_workbook(PARSE_RESULT_XLSX)
ws_xlsx = wb.active
parse_headers = [c.value for c in ws_xlsx[1]]
name_to_row = {str(ws_xlsx.cell(r,1).value).strip(): r for r in range(2, ws_xlsx.max_row+1) if ws_xlsx.cell(r,1).value}

row_values = [data.get(h, "") for h in parse_headers]
if name in name_to_row:
    ri = name_to_row[name]
    for ci, v in enumerate(row_values, 1):
        ws_xlsx.cell(ri, ci).value = v
    print(f"[파싱결과.xlsx] {name} 업데이트 (행 {ri})", flush=True)
else:
    ws_xlsx.append(row_values)
    print(f"[파싱결과.xlsx] {name} 신규 추가", flush=True)
wb.save(PARSE_RESULT_XLSX)

# 구글시트
creds = get_credentials()
gc = gspread.authorize(creds)
sh = gc.open_by_key(SPREADSHEET_ID)
ws_gs = sh.worksheet("접수명단")
all_vals = ws_gs.get_all_values()
headers_gs = all_vals[0]
col_idx = {sc: headers_gs.index(sc)+1 for sc in COL_MAP if sc in headers_gs}
name_col_gs = headers_gs.index("성명") + 1
name_to_gs_row = {row[name_col_gs-1].strip(): i+2 for i, row in enumerate(all_vals[1:]) if len(row)>=name_col_gs and row[name_col_gs-1].strip()}

row_idx = name_to_gs_row.get(name)
if not row_idx:
    print(f"[접수명단] {name} 없음 - 스킵", flush=True)
    sys.exit(1)

updates = []
for sheet_col, parsed_key in COL_MAP.items():
    cidx = col_idx.get(sheet_col)
    if not cidx:
        continue
    val = data.get(parsed_key, "")
    col_letter = rowcol_to_a1(1, cidx).rstrip("0123456789")
    updates.append({"range": f"{col_letter}{row_idx}", "values": [[val]]})

ws_gs.batch_update(updates)
print(f"[접수명단] {name} 업데이트 완료 ({len(updates)}셀)", flush=True)
print("전체 완료!", flush=True)
